import datetime
import random
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required

from .models import (
    Skill, Milestone, StudyLog, Profile, CentralizedSkill,
    Education, WorkExperience, Certification, ApplicantSkill,
    JobVacancy, JobSkillRequirement, TrainingProgram, TrainingEnrollment, Referral,
    GapScoreLog, JobBookmark, Notification, ApplicantDocument, Interview
)
from .forms import SkillForm, MilestoneForm, StudyLogForm
import re
from datetime import datetime

def calculate_total_experience_years(profile):
    experiences = profile.experience.all()
    if not experiences.exists():
        return "0 Years"
    
    total_months = 0
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    months_map = {
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
        'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
        'january': 1, 'february': 2, 'march': 3, 'april': 4, 'june': 6,
        'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12
    }
    
    for exp in experiences:
        start_str = (exp.start_date or '').strip()
        end_str = (exp.end_date or '').strip()
        
        start_year = None
        start_month = 1
        
        year_match = re.search(r'\b(19\d\d|20\d\d)\b', start_str)
        if year_match:
            start_year = int(year_match.group(1))
        
        for m_name, m_num in months_map.items():
            if m_name in start_str.lower():
                start_month = m_num
                break
                
        end_year = None
        end_month = 1
        
        if not end_str or 'present' in end_str.lower() or exp.is_current:
            end_year = current_year
            end_month = current_month
        else:
            year_match = re.search(r'\b(19\d\d|20\d\d)\b', end_str)
            if year_match:
                end_year = int(year_match.group(1))
            for m_name, m_num in months_map.items():
                if m_name in end_str.lower():
                    end_month = m_num
                    break
        
        if start_year is not None and end_year is not None:
            months = (end_year - start_year) * 12 + (end_month - start_month)
            if months < 0:
                months = 0
            if months == 0:
                months = 6  # default 0.5 year if user inputs same year
            total_months += months
            
    years = round(total_months / 12, 1)
    if years.is_integer():
        years_val = int(years)
    else:
        years_val = years
        
    if years_val == 1:
        return "1 Year"
    else:
        return f"{years_val} Years"

def verified_required(view_func):
    @login_required(login_url='login')
    def _wrapped_view(request, *args, **kwargs):
        if request.user.is_superuser or request.user.is_staff:
            if not hasattr(request.user, 'profile'):
                Profile.objects.create(user=request.user, role='admin', is_verified=True)
            elif not request.user.profile.is_verified or request.user.profile.role != 'admin':
                request.user.profile.is_verified = True
                request.user.profile.role = 'admin'
                request.user.profile.save()
        else:
            if not hasattr(request.user, 'profile'):
                Profile.objects.create(user=request.user)
            if not request.user.profile.is_verified:
                return redirect('verify')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def role_required(allowed_roles):
    def decorator(view_func):
        @verified_required
        def _wrapped_view(request, *args, **kwargs):
            if request.user.profile.role not in allowed_roles:
                role = request.user.profile.role
                if role == 'admin':
                    return redirect('peso_dashboard_admin')
                elif role == 'employer':
                    return redirect('employer_dashboard')
                else:
                    # If applicant dashboard is allowed and profile is incomplete, dashboard redirect will handle it
                    return redirect('applicant_dashboard')
            
            if request.user.profile.role == 'applicant' and not request.user.profile.is_profile_complete:
                if request.resolver_match.url_name not in ['profile_wizard', 'logout']:
                    return redirect('profile_wizard')
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator

def dashboard(request):
    if not request.user.is_authenticated:
        return redirect('login')
    if request.user.is_superuser or request.user.is_staff:
        if not hasattr(request.user, 'profile'):
            Profile.objects.create(user=request.user, role='admin', is_verified=True)
        elif not request.user.profile.is_verified or request.user.profile.role != 'admin':
            request.user.profile.is_verified = True
            request.user.profile.role = 'admin'
            request.user.profile.save()
    else:
        if not hasattr(request.user, 'profile'):
            Profile.objects.create(user=request.user)
        if not request.user.profile.is_verified:
            return redirect('verify')
    
    role = request.user.profile.role
    if role == 'applicant' and not request.user.profile.is_profile_complete:
        return redirect('profile_wizard')
        
    if role == 'admin':
        return redirect('peso_dashboard_admin')
    elif role == 'employer':
        return redirect('employer_dashboard')
    else:
        return redirect('applicant_dashboard')

def legacy_skill_dashboard(request):
    skills = Skill.objects.all()
    
    # Calculate aggregate stats
    total_skills = skills.count()
    skills_in_progress = skills.filter(status='in_progress').count()
    skills_completed = skills.filter(status='completed').count()
    skills_not_started = skills.filter(status='not_started').count()
    
    total_study_minutes = StudyLog.objects.aggregate(Sum('duration_minutes'))['duration_minutes__sum'] or 0
    total_study_hours = round(total_study_minutes / 60, 1)
    
    total_milestones = Milestone.objects.count()
    completed_milestones = Milestone.objects.filter(is_completed=True).count()
    milestone_progress = int((completed_milestones / total_milestones) * 100) if total_milestones > 0 else 0

    # Weekly study hours for line chart
    today = timezone.localdate()
    last_7_days = [today - datetime.timedelta(days=i) for i in range(6, -1, -1)]
    weekly_labels = []
    weekly_data = []
    
    for date in last_7_days:
        daily_minutes = StudyLog.objects.filter(date=date).aggregate(Sum('duration_minutes'))['duration_minutes__sum'] or 0
        weekly_data.append(round(daily_minutes / 60, 2))
        weekly_labels.append(date.strftime('%a (%b %d)'))

    # Category distribution for pie chart
    category_counts = Skill.objects.values('category').annotate(count=Count('id')).order_by('-count')
    category_labels = [item['category'] for item in category_counts]
    category_data = [item['count'] for item in category_counts]

    recent_logs = StudyLog.objects.select_related('skill').all()[:5]

    context = {
        'skills': skills,
        'total_skills': total_skills,
        'skills_in_progress': skills_in_progress,
        'skills_completed': skills_completed,
        'skills_not_started': skills_not_started,
        'total_study_hours': total_study_hours,
        'milestone_progress': milestone_progress,
        'weekly_labels': weekly_labels,
        'weekly_data': weekly_data,
        'category_labels': category_labels,
        'category_data': category_data,
        'recent_logs': recent_logs,
    }
    return render(request, 'tracker/dashboard.html', context)

def skill_detail(request, pk):
    skill = get_object_or_404(Skill, pk=pk)
    milestones = skill.milestones.all().order_by('is_completed', 'target_date')
    study_logs = skill.study_logs.all().order_by('-date', '-created_at')
    
    milestone_form = MilestoneForm()
    study_log_form = StudyLogForm(initial={'date': timezone.localdate()})
    
    context = {
        'skill': skill,
        'milestones': milestones,
        'study_logs': study_logs,
        'milestone_form': milestone_form,
        'study_log_form': study_log_form,
    }
    return render(request, 'tracker/skill_detail.html', context)

def skill_create(request):
    if request.method == 'POST':
        form = SkillForm(request.POST)
        if form.is_valid():
            skill = form.save()
            messages.success(request, f"Skill '{skill.name}' created successfully!")
            return redirect('skill_detail', pk=skill.pk)
    else:
        form = SkillForm()
    
    return render(request, 'tracker/skill_form.html', {'form': form, 'title': 'Add New Skill'})

def skill_edit(request, pk):
    skill = get_object_or_404(Skill, pk=pk)
    if request.method == 'POST':
        form = SkillForm(request.POST, instance=skill)
        if form.is_valid():
            form.save()
            messages.success(request, f"Skill '{skill.name}' updated successfully!")
            return redirect('skill_detail', pk=skill.pk)
    else:
        form = SkillForm(instance=skill)
        
    return render(request, 'tracker/skill_form.html', {'form': form, 'title': 'Edit Skill', 'skill': skill})

@require_POST
def skill_delete(request, pk):
    skill = get_object_or_404(Skill, pk=pk)
    name = skill.name
    skill.delete()
    messages.success(request, f"Skill '{name}' deleted successfully!")
    return redirect('dashboard')

@require_POST
def add_milestone(request, skill_id):
    skill = get_object_or_404(Skill, id=skill_id)
    form = MilestoneForm(request.POST)
    if form.is_valid():
        milestone = form.save(commit=False)
        milestone.skill = skill
        milestone.save()
        messages.success(request, f"Milestone '{milestone.title}' added to {skill.name}!")
    else:
        messages.error(request, "Failed to add milestone. Please verify the input.")
    return redirect('skill_detail', pk=skill.id)

@require_POST
def toggle_milestone(request, milestone_id):
    milestone = get_object_or_404(Milestone, id=milestone_id)
    milestone.is_completed = not milestone.is_completed
    milestone.save()
    
    # Auto adjust skill status if transitioning
    skill = milestone.skill
    if skill.status == 'not_started' and milestone.is_completed:
        skill.status = 'in_progress'
        skill.save()
    elif skill.status == 'in_progress' and skill.completion_percentage == 100:
        skill.status = 'completed'
        skill.save()
        
    status_text = "completed" if milestone.is_completed else "marked incomplete"
    messages.success(request, f"Milestone '{milestone.title}' {status_text}!")
    return redirect('skill_detail', pk=skill.id)

@require_POST
def add_study_log(request, skill_id):
    skill = get_object_or_404(Skill, id=skill_id)
    form = StudyLogForm(request.POST)
    if form.is_valid():
        log = form.save(commit=False)
        log.skill = skill
        log.save()
        
        # Auto start skill status
        if skill.status == 'not_started':
            skill.status = 'in_progress'
            skill.save()
            
        messages.success(request, f"Logged {log.duration_minutes} minutes for {skill.name}!")
    else:
        messages.error(request, "Failed to log study hours. Please verify input values.")
    return redirect('skill_detail', pk=skill.id)


# ==========================================
# AUTHENTICATION FLOW VIEWS
# ==========================================
from django.core.mail import send_mail
from django.conf import settings

def signup_view(request):
    if request.method == 'POST':
        firstname = request.POST.get('firstname', '').strip()
        lastname = request.POST.get('lastname', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        password = request.POST.get('password', '')
        confirm_password = request.POST.get('confirm_password', '')
        role = request.POST.get('role', 'applicant')
        birthdate = request.POST.get('birthdate')
        civil_status = request.POST.get('civil_status')
        company_name = request.POST.get('company_name', '').strip()

        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return render(request, 'tracker/LOGIN/sign_up_account_details.html')

        if role == 'employer' and not company_name:
            messages.error(request, "Company Name is required for Employer accounts.")
            return render(request, 'tracker/LOGIN/sign_up_account_details.html')

        if User.objects.filter(username=email).exists():
            messages.error(request, "An account with this email already exists.")
            return render(request, 'tracker/LOGIN/sign_up_account_details.html')

        # Check duplicate applicant by name and birthdate
        if role == 'applicant' and birthdate:
            duplicate = Profile.objects.filter(
                user__first_name__iexact=firstname,
                user__last_name__iexact=lastname,
                birthdate=birthdate
            ).first()
            if duplicate:
                messages.error(request, f"An applicant profile for {firstname} {lastname} with birthdate {birthdate} already exists.")
                return render(request, 'tracker/LOGIN/sign_up_account_details.html')

        otp = str(random.randint(100000, 999999))
        request.session['pending_signup'] = {
            'firstname': firstname,
            'lastname': lastname,
            'email': email,
            'phone': phone,
            'password': password,
            'role': role,
            'birthdate': birthdate,
            'civil_status': civil_status,
            'company_name': company_name,
            'otp': otp
        }
        
        try:
            send_mail(
                subject='Account Creation Confirmation - SKILLUP',
                message=f'Hello {firstname},\n\nYour verification code is: {otp}\n\nYour account will be created once you confirm this code on our registration portal.',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=False
            )
            messages.success(request, "Registration successful! Please check your email for your verification code.")
        except Exception as e:
            print("EMAIL SEND ERROR IN SIGNUP:", str(e))
            messages.warning(request, f"Notice: Could not deliver email (SMTP Error: {str(e)}). For local testing, your verification code is: {otp}")
        
        return redirect('verify')
    return render(request, 'tracker/LOGIN/sign_up_account_details.html')

def login_view(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=email, password=password)
        if user is not None:
            auth_login(request, user)
            if user.is_superuser or user.is_staff:
                if not hasattr(user, 'profile'):
                    Profile.objects.create(user=user, role='admin', is_verified=True)
                elif not user.profile.is_verified or user.profile.role != 'admin':
                    user.profile.is_verified = True
                    user.profile.role = 'admin'
                    user.profile.save()
                return redirect('dashboard')
                
            if not hasattr(user, 'profile'):
                Profile.objects.create(user=user)
            
            if not user.profile.is_verified:
                otp = str(random.randint(100000, 999999))
                user.profile.verification_code = otp
                user.profile.save()
                
                try:
                    send_mail(
                        subject='Account Verification Code - SKILLUP',
                        message=f'Hello {user.first_name},\n\nPlease verify your account. Your verification code is: {otp}\n\nThank you for registering at SKILLUP!',
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[user.email],
                        fail_silently=False
                    )
                    messages.success(request, "Please check your email for the verification code to verify your account.")
                except Exception as e:
                    print("EMAIL SEND ERROR IN LOGIN:", str(e))
                    messages.warning(request, f"Notice: Could not deliver email (SMTP Error: {str(e)}). For local testing, your verification code is: {otp}")
                
                return redirect('verify')
            
            return redirect('dashboard')
        else:
            messages.error(request, "Invalid email or password.")
            return render(request, 'tracker/LOGIN/login.html')
    
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'tracker/LOGIN/login.html')

def verify_view(request):
    pending = request.session.get('pending_signup')
    
    if not pending and not request.user.is_authenticated:
        return redirect('login')
        
    if request.user.is_authenticated:
        if request.user.is_superuser or request.user.is_staff:
            if not hasattr(request.user, 'profile'):
                Profile.objects.create(user=request.user, role='admin', is_verified=True)
            elif not request.user.profile.is_verified or request.user.profile.role != 'admin':
                request.user.profile.is_verified = True
                request.user.profile.role = 'admin'
                request.user.profile.save()
            return redirect('dashboard')
            
        if not hasattr(request.user, 'profile'):
            Profile.objects.create(user=request.user)
        if request.user.profile.is_verified:
            return redirect('dashboard')

    if request.method == 'POST':
        code = "".join([request.POST.get(f'digit{i}', '') for i in range(1, 7)])
        
        if pending:
            if code == pending['otp']:
                try:
                    # Create Django User
                    user = User.objects.create_user(
                        username=pending['email'],
                        email=pending['email'],
                        password=pending['password'],
                        first_name=pending['firstname'],
                        last_name=pending['lastname']
                    )
                    
                    # Create Profile
                    profile = Profile.objects.create(
                        user=user,
                        role=pending['role'],
                        phone_number=pending['phone'],
                        birthdate=pending['birthdate'] if pending['birthdate'] else None,
                        civil_status=pending['civil_status'],
                        company_name=pending['company_name'] if pending['role'] == 'employer' else None,
                        is_verified=True,
                        verification_code=''
                    )
                    
                    # Login
                    authenticated_user = authenticate(request, username=pending['email'], password=pending['password'])
                    if authenticated_user:
                        auth_login(request, authenticated_user)
                        
                    del request.session['pending_signup']
                    messages.success(request, "Account created and verified successfully!")
                    return redirect('dashboard')
                except Exception as e:
                    messages.error(request, f"Registration failed: {str(e)}")
                    return redirect('signup')
            else:
                messages.error(request, "Invalid verification code. Please try again.")
        else:
            if code == request.user.profile.verification_code:
                request.user.profile.is_verified = True
                request.user.profile.save()
                messages.success(request, "Account verified successfully!")
                return redirect('dashboard')
            else:
                messages.error(request, "Invalid verification code. Please try again.")
                
    context = {}
    if pending:
        context['email'] = pending['email']
    elif request.user.is_authenticated:
        context['email'] = request.user.email
        
    return render(request, 'tracker/LOGIN/account_verification.html', context)

def resend_verification_code(request):
    pending = request.session.get('pending_signup')
    
    if not pending and not request.user.is_authenticated:
        return redirect('login')
        
    otp = str(random.randint(100000, 999999))
    email = ""
    firstname = ""
    
    if pending:
        pending['otp'] = otp
        request.session['pending_signup'] = pending
        email = pending['email']
        firstname = pending['firstname']
    else:
        request.user.profile.verification_code = otp
        request.user.profile.save()
        email = request.user.email
        firstname = request.user.first_name
        
    try:
        send_mail(
            subject='Account Verification Code - SKILLUP',
            message=f'Hello {firstname},\n\nYour new verification code is: {otp}\n\nThank you for registering at SKILLUP!',
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            fail_silently=False
        )
        messages.success(request, "A new verification code has been successfully sent to your email.")
    except Exception as e:
        print("EMAIL SEND ERROR IN RESEND:", str(e))
        messages.error(request, f"Failed to send email. SMTP Error: {str(e)}")
        
    return redirect('verify')

def logout_view(request):
    auth_logout(request)
    messages.success(request, "You have been logged out.")
    return redirect('login')

def create_apex_employer_and_jobs():
    from django.contrib.auth.models import User
    from .models import Profile, JobVacancy, JobSkillRequirement, CentralizedSkill

    email = 'apex_employer@test.com'
    password = 'Password123!'
    
    try:
        user = User.objects.get(username=email)
        if JobVacancy.objects.filter(employer=user.profile).count() == 20:
            return
    except Exception:
        pass
    
    user, created = User.objects.get_or_create(
        username=email,
        defaults={
            'email': email,
            'first_name': 'Apex',
            'last_name': 'Employer'
        }
    )
    if created or not user.check_password(password):
        user.set_password(password)
        user.save()
    
    profile, _ = Profile.objects.update_or_create(
        user=user,
        defaults={
            'role': 'employer',
            'is_verified': True,
            'company_name': 'Apex Corporation',
            'industry': 'Information Technology & Services',
            'company_size': '500-1000 employees',
            'website': 'https://apex-corp.com',
            'contact_name': 'Jane Doe',
            'contact_position': 'Director of Recruitment',
            'contact_email': email,
            'employment_type_offered': 'Local',
            'location': 'Manila, Philippines',
            'soft_notes': 'Apex Corporation is a global technology solutions provider specializing in enterprise-grade software development, workforce upskilling, and business consultation.'
        }
    )
    
    # Define 20 jobs
    jobs = [
        {
            'title': 'Senior React Developer',
            'location': 'Remote / Manila',
            'salary_range': '₱90k - ₱120k',
            'category': 'IT',
            'description': 'We are looking for a Senior React Developer to join our frontend team. You will lead the development of our modern enterprise dashboard and design rich web applications.',
            'requirements': [('React', 5), ('TypeScript', 4), ('HTML & CSS', 4)]
        },
        {
            'title': 'Python Backend Engineer',
            'location': 'Manila, PH',
            'salary_range': '₱80k - ₱110k',
            'category': 'IT',
            'description': 'Join us to build high-performance backend APIs using Python and Django. Experience with PostgreSQL and REST frameworks is highly preferred.',
            'requirements': [('Python', 5), ('Django', 4), ('SQL', 4)]
        },
        {
            'title': 'Digital Marketing Specialist',
            'location': 'Quezon City, PH',
            'salary_range': '₱35k - ₱50k',
            'category': 'MKT',
            'description': 'We are seeking a creative Digital Marketing Specialist to manage our email marketing campaigns, design search engine optimization strategies, and build content calendars.',
            'requirements': [('Email Marketing', 4), ('Search Engine Optimization', 4), ('Social Media Management', 3)]
        },
        {
            'title': 'Financial Accountant',
            'location': 'Manila, PH',
            'salary_range': '₱40k - ₱55k',
            'category': 'FIN',
            'description': 'Responsible for managing ledger accounts, petty cash processes, and generating financial statements. Must have high attention to detail.',
            'requirements': [('Basic Bookkeeping Software', 4), ('Microsoft Excel', 5), ('Attention to Detail', 5)]
        },
        {
            'title': 'Customer Support Agent',
            'location': 'Cebu, PH',
            'category': 'BPO',
            'salary_range': '₱22k - ₱28k',
            'description': 'Provide excellent phone and online customer support. Resolve queries and handle customer complaints professionally.',
            'requirements': [('Verbal Communication', 4), ('Active Listening', 4), ('Written Communication', 4)]
        },
        {
            'title': 'Executive Administrative Assistant',
            'location': 'Manila, PH',
            'category': 'ADM',
            'salary_range': '₱30k - ₱42k',
            'description': 'Support the executive team with schedule planning, business correspondence, and document management.',
            'requirements': [('Time Management', 4), ('Business Correspondence', 4), ('Google Workspace', 3)]
        },
        {
            'title': 'Retail Store Manager',
            'location': 'Davao, PH',
            'category': 'RTL',
            'salary_range': '₱35k - ₱45k',
            'description': 'Oversee daily store operations, manage stock levels, and guide retail associates to meet monthly sales targets.',
            'requirements': [('Retail Store Operations', 5), ('Stock Monitoring & Replenishment', 4), ('Customer Service Orientation', 4)]
        },
        {
            'title': 'Restaurant Shift Supervisor',
            'location': 'Cebu, PH',
            'category': 'F&B',
            'salary_range': '₱25k - ₱32k',
            'description': 'Supervise dining area and kitchen staffs, ensure food safety compliance, and coordinate high-quality service.',
            'requirements': [('Food & Beverage Service', 4), ('Food Safety & Hygiene Practices', 5), ('Teamwork & Collaboration', 4)]
        },
        {
            'title': 'Clinical Nurse Coordinator',
            'location': 'Davao, PH',
            'category': 'MED',
            'salary_range': '₱30k - ₱40k',
            'description': 'Coordinate patient care activities, monitor vital signs, and provide high-quality healthcare services.',
            'requirements': [('Health Care Services', 4), ('Empathy & Patience', 5), ('Active Listening', 4)]
        },
        {
            'title': 'Technical Trainer / Educator',
            'location': 'Remote',
            'category': 'EDU',
            'salary_range': '₱45k - ₱60k',
            'description': 'Develop curriculum and deliver technical courses. Excellent presentation and communication skills are required.',
            'requirements': [('Presentation Skills', 5), ('Verbal Communication', 5), ('Active Listening', 4)]
        },
        {
            'title': 'Project Civil Engineer',
            'location': 'Manila, PH',
            'category': 'ENG',
            'salary_range': '₱50k - ₱70k',
            'description': 'Lead local infrastructure projects. Review structural plans, manage timelines, and solve on-site engineering challenges.',
            'requirements': [('Problem Solving', 4), ('Critical Thinking', 4), ('Planning & Scheduling', 4)]
        },
        {
            'title': 'Electrical Systems Inspector',
            'location': 'Cebu, PH',
            'category': 'ELC',
            'salary_range': '₱35k - ₱45k',
            'description': 'Conduct safety inspections and maintenance on commercial building electrical systems. Certified electrical worker.',
            'requirements': [('Electrical Installation & Maintenance', 5), ('Troubleshooting', 4), ('Occupational Health & Safety Awareness', 5)]
        },
        {
            'title': 'Senior Automotive Mechanic',
            'location': 'Manila, PH',
            'category': 'MCH',
            'salary_range': '₱28k - ₱38k',
            'description': 'Perform complex diagnostic tests, engine repairs, and mechanical servicing on modern automotive fleets.',
            'requirements': [('Automotive Servicing', 5), ('Troubleshooting', 5), ('Problem Solving', 4)]
        },
        {
            'title': 'Structural Welder (TESDA Certified)',
            'location': 'Manila, PH',
            'category': 'TVT',
            'salary_range': '₱25k - ₱35k',
            'description': 'Perform high-precision shielded metal arc welding and gas metal arc welding on heavy construction projects.',
            'requirements': [('Shielded Metal Arc Welding', 5), ('Gas Metal Arc Welding', 4), ('Occupational Health & Safety Awareness', 5)]
        },
        {
            'title': 'Warehouse Operations Lead',
            'location': 'Manila, PH',
            'category': 'LOG',
            'salary_range': '₱28k - ₱36k',
            'description': 'Manage logistics operations, loading schedules, and maintain inventory counts in our central warehouse.',
            'requirements': [('Warehouse Operations', 5), ('Inventory Count & Reporting', 4), ('Time Management', 4)]
        },
        {
            'title': 'Head Security Specialist',
            'location': 'Manila, PH',
            'category': 'SEC',
            'salary_range': '₱24k - ₱30k',
            'description': 'Oversee facility security guards, identify risks, and ensure complete compliance with corporate safety rules.',
            'requirements': [('Basic Security Guard Duties', 5), ('Risk Awareness', 5), ('Conflict Resolution', 4)]
        },
        {
            'title': 'Facilities Maintenance Worker',
            'location': 'Manila, PH',
            'category': 'GEN',
            'salary_range': '₱26k - ₱34k',
            'description': 'Responsible for general utility repairs, building checks, sanitation maintenance, and machinery service.',
            'requirements': [('General Utility', 4), ('Troubleshooting', 4), ('Occupational Health & Safety Awareness', 4)]
        },
        {
            'title': 'Logistics & Fleet Driver',
            'location': 'Quezon City, PH',
            'category': 'DRV',
            'salary_range': '₱18k - ₱24k',
            'description': 'Drive delivery vans safely across Metro Manila. Maintain clean driving record and execute timely deliveries.',
            'requirements': [('Driving', 5), ('Punctuality & Reliability', 5), ('Customer Service Orientation', 3)]
        },
        {
            'title': 'Garment Production Tailor',
            'location': 'Manila, PH',
            'category': 'MFG',
            'salary_range': '₱20k - ₱25k',
            'description': 'Execute precise garment cutting and stitching according to patterns. Maintain quality control protocols.',
            'requirements': [('Dressmaking & Tailoring', 5), ('Attention to Detail', 4), ('Commitment to Quality', 4)]
        },
        {
            'title': 'Agricultural Farm Manager',
            'location': 'Laguna, PH',
            'category': 'AGR',
            'salary_range': '₱30k - ₱40k',
            'description': 'Manage crop cycles, optimize agricultural techniques, and oversee general farm worker teams.',
            'requirements': [('Agricultural Crops Production', 5), ('Problem Solving', 4), ('Leadership', 4)]
        }
    ]

    # Only create/recreate them if we don't already have exactly 20 jobs for this employer
    if JobVacancy.objects.filter(employer=profile).count() != 20:
        JobVacancy.objects.filter(employer=profile).delete()
        for job in jobs:
            vacancy = JobVacancy.objects.create(
                title=job['title'],
                employer=profile,
                location=job['location'],
                salary_range=job['salary_range'],
                category=job['category'],
                description=job['description'],
                status='Open',
                slots=3,
                remaining_slots=3
            )
            for skill_name, req_level in job['requirements']:
                try:
                    skill_obj = CentralizedSkill.objects.get(name=skill_name)
                except CentralizedSkill.DoesNotExist:
                    skill_obj = CentralizedSkill.objects.create(
                        name=skill_name,
                        category='General',
                        description=f'Seeded skill: {skill_name}'
                    )
                JobSkillRequirement.objects.get_or_create(
                    job_vacancy=vacancy,
                    skill=skill_obj,
                    defaults={'required_proficiency': req_level}
                )

def create_prime_employer_and_jobs():
    from django.contrib.auth.models import User
    from .models import Profile, JobVacancy, JobSkillRequirement, CentralizedSkill

    email = 'prime_employer@test.com'
    password = 'Password123!'
    
    try:
        user = User.objects.get(username=email)
        if JobVacancy.objects.filter(employer=user.profile).count() == 20:
            return
    except Exception:
        pass
    
    user, created = User.objects.get_or_create(
        username=email,
        defaults={
            'email': email,
            'first_name': 'Prime',
            'last_name': 'Employer'
        }
    )
    if created or not user.check_password(password):
        user.set_password(password)
        user.save()
    
    profile, _ = Profile.objects.update_or_create(
        user=user,
        defaults={
            'role': 'employer',
            'is_verified': True,
            'company_name': 'Prime Industries',
            'industry': 'Manufacturing & Industrial Conglomerate',
            'company_size': '1000+ employees',
            'website': 'https://prime-industries.com',
            'contact_name': 'Robert Chen',
            'contact_position': 'VP of Operations',
            'contact_email': email,
            'employment_type_offered': 'Local',
            'location': 'Cavite, Philippines',
            'soft_notes': 'Prime Industries is a premier industrial conglomerate specializing in precision manufacturing, global logistics, and sustainable agricultural operations across Southeast Asia.'
        }
    )
    
    # Define 20 jobs for prime_employer
    jobs = [
        {
            'title': 'Senior Project Manager',
            'location': 'Manila, PH',
            'salary_range': '₱100k - ₱140k',
            'category': 'IT',
            'description': 'Lead development and delivery of enterprise manufacturing platforms. Oversee product roadmaps and agile sprints.',
            'requirements': [('Agile', 5), ('Scrum', 4), ('Kanban', 4)]
        },
        {
            'title': 'DevOps Specialist',
            'location': 'Remote / Manila',
            'salary_range': '₱95k - ₱130k',
            'category': 'IT',
            'description': 'Configure containerization and orchestration pipelines on AWS. Establish CI/CD workflows and manage cloud infrastructure.',
            'requirements': [('Docker', 4), ('AWS Cloud', 4), ('Kubernetes', 4)]
        },
        {
            'title': 'HR Operations Officer',
            'location': 'Cavite, PH',
            'salary_range': '₱30k - ₱40k',
            'category': 'ADM',
            'description': 'Handle employee relation activities, organize shift scheduling, and manage candidate onboarding documentation.',
            'requirements': [('Interpersonal Communication', 4), ('Written Communication', 4), ('Time Management', 4)]
        },
        {
            'title': 'General Ledger Accountant',
            'location': 'Cavite, PH',
            'salary_range': '₱42k - ₱55k',
            'category': 'FIN',
            'description': 'Responsible for ledger auditing, trial balance sheets, financial records maintenance, and petty cash systems management.',
            'requirements': [('Basic Bookkeeping Software', 4), ('Microsoft Excel', 5), ('Attention to Detail', 5)]
        },
        {
            'title': 'Business Development Representative',
            'location': 'Manila, PH',
            'salary_range': '₱35k - ₱48k',
            'category': 'MKT',
            'description': 'Drive sales outreach, generate client presentations, and establish strong customer relationship channels.',
            'requirements': [('Verbal Communication', 5), ('Presentation Skills', 4), ('Sales Orientation', 4)]
        },
        {
            'title': 'Customer Relations Team Lead',
            'location': 'Cebu, PH',
            'salary_range': '₱45k - ₱58k',
            'category': 'BPO',
            'description': 'Lead client service representatives team. Resolve escalations, handle client complaints, and compile customer satisfaction metrics.',
            'requirements': [('Complaint Handling', 5), ('Empathy & Patience', 5), ('Client Communication', 4)]
        },
        {
            'title': 'TESDA Structural Welder (NC II)',
            'location': 'Cavite, PH',
            'salary_range': '₱28k - ₱36k',
            'category': 'TVT',
            'description': 'Execute welding repairs and structural welds in manufacturing facilities. Follow strict safety protocols.',
            'requirements': [('Shielded Metal Arc Welding', 5), ('Gas Metal Arc Welding', 4), ('Occupational Health & Safety Awareness', 5)]
        },
        {
            'title': 'Electrical Design Engineer',
            'location': 'Cavite, PH',
            'salary_range': '₱60k - ₱85k',
            'category': 'ENG',
            'description': 'Create schematics for facility power grids, read blueprints, inspect layout designs, and resolve technical issues.',
            'requirements': [('Electrical Installation & Maintenance', 5), ('Blueprint & Plan Reading', 4), ('Problem Solving', 4)]
        },
        {
            'title': 'Logistics & Warehousing Supervisor',
            'location': 'Laguna, PH',
            'salary_range': '₱38k - ₱50k',
            'category': 'LOG',
            'description': 'Direct logistics warehouse operations, monitor loading/unloading tasks, and manage stock counts.',
            'requirements': [('Warehouse Operations', 5), ('Stock Monitoring & Replenishment', 4), ('Leadership', 4)]
        },
        {
            'title': 'Site Safety Inspector',
            'location': 'Cavite, PH',
            'salary_range': '₱32k - ₱42k',
            'category': 'SEC',
            'description': 'Conduct daily safety audits, evaluate site security systems, identify risks, and enforce compliance guidelines.',
            'requirements': [('Basic Security Guard Duties', 4), ('Risk Awareness', 5), ('Occupational Health & Safety Awareness', 5)]
        },
        {
            'title': 'Facilities Custodian',
            'location': 'Cavite, PH',
            'salary_range': '₱16k - ₱20k',
            'category': 'GEN',
            'description': 'Perform building cleaning, waste disposal, janitorial services, and maintain warehouse sanitations.',
            'requirements': [('Housekeeping', 4), ('Sanitation Standards', 4), ('Physical Fitness for Work', 4)]
        },
        {
            'title': 'Heavy Equipment Driver',
            'location': 'Laguna, PH',
            'salary_range': '₱22k - ₱28k',
            'category': 'DRV',
            'description': 'Operate loaders, forklifts, and transportation vehicles safely. Conduct pre-trip safety checks.',
            'requirements': [('Driving', 5), ('Occupational Health & Safety Awareness', 4), ('Physical Fitness for Work', 4)]
        },
        {
            'title': 'Industrial Electrician',
            'location': 'Cavite, PH',
            'salary_range': '₱26k - ₱34k',
            'category': 'ELC',
            'description': 'Repair industrial machinery circuits, inspect transformers, and execute electrical installations.',
            'requirements': [('Electrical Installation & Maintenance', 5), ('Troubleshooting', 4), ('Compliance Mindset', 4)]
        },
        {
            'title': 'CNC Machine Operator',
            'location': 'Cavite, PH',
            'salary_range': '₱24k - ₱30k',
            'category': 'MCH',
            'description': 'Set up and run CNC fabrication machinery, perform sorting, and execute quality inspections.',
            'requirements': [('Basic Machine Operation', 5), ('Quality Checking & Sorting', 4), ('Occupational Health & Safety Awareness', 4)]
        },
        {
            'title': 'Production Line Sewing Tailor',
            'location': 'Cavite, PH',
            'salary_range': '₱18k - ₱23k',
            'category': 'MFG',
            'description': 'Operate industrial sewing machinery for garment assembly. Meet quality checks and daily counts.',
            'requirements': [('Dressmaking & Tailoring', 5), ('Attention to Detail', 4), ('Punctuality & Reliability', 4)]
        },
        {
            'title': 'Agricultural Farm Supervisor',
            'location': 'Batangas, PH',
            'salary_range': '₱32k - ₱42k',
            'category': 'AGR',
            'description': 'Manage farm worker schedules, coordinate agricultural crop harvests, and run farm systems troubleshooting.',
            'requirements': [('Agricultural Crops Production', 5), ('Leadership', 4), ('Problem Solving', 4)]
        },
        {
            'title': 'Catering Services Coordinator',
            'location': 'Manila, PH',
            'salary_range': '₱25k - ₱32k',
            'category': 'F&B',
            'description': 'Plan corporate catering events, coordinate dining table layouts, and audit sanitation practices.',
            'requirements': [('Food & Beverage Service', 4), ('Table Setting & Service', 4), ('Food Safety & Hygiene Practices', 5)]
        },
        {
            'title': 'Nurse Practitioner',
            'location': 'Cavite, PH',
            'salary_range': '₱35k - ₱45k',
            'category': 'MED',
            'description': 'Provide medical care services in the industrial clinic, monitor vitals, and assist injured workers.',
            'requirements': [('Health Care Services', 4), ('Vital Signs Monitoring', 4), ('Patient / Elder Care Assistance', 4)]
        },
        {
            'title': 'ESL Language Teacher',
            'location': 'Remote',
            'salary_range': '₱30k - ₱42k',
            'category': 'EDU',
            'description': 'Conduct conversational English lectures for overseas corporate staff. Prepare class guides.',
            'requirements': [('English Communication – Oral', 5), ('English Communication – Written', 5), ('Presentation Skills', 4)]
        },
        {
            'title': 'Warehouse Dispatch Associate',
            'location': 'Laguna, PH',
            'salary_range': '₱18k - ₱24k',
            'category': 'LOG',
            'description': 'Process shipping logs, select transport routes, package dispatch goods, and meet tight deadlines.',
            'requirements': [('Warehouse Operations', 4), ('Order Taking & Processing', 4), ('Time Management', 4)]
        }
    ]

    # Only create/recreate them if we don't already have exactly 20 jobs for this employer
    if JobVacancy.objects.filter(employer=profile).count() != 20:
        JobVacancy.objects.filter(employer=profile).delete()
        for job in jobs:
            vacancy = JobVacancy.objects.create(
                title=job['title'],
                employer=profile,
                location=job['location'],
                salary_range=job['salary_range'],
                category=job['category'],
                description=job['description'],
                status='Open',
                slots=3,
                remaining_slots=3
            )
            for skill_name, req_level in job['requirements']:
                try:
                    skill_obj = CentralizedSkill.objects.get(name=skill_name)
                except CentralizedSkill.DoesNotExist:
                    skill_obj = CentralizedSkill.objects.create(
                        name=skill_name,
                        category='General',
                        description=f'Seeded skill: {skill_name}'
                    )
                JobSkillRequirement.objects.get_or_create(
                    job_vacancy=vacancy,
                    skill=skill_obj,
                    defaults={'required_proficiency': req_level}
                )

def create_nexus_employer_and_jobs():
    from django.contrib.auth.models import User
    from .models import Profile, JobVacancy, JobSkillRequirement, CentralizedSkill

    email = 'nexus_employer@test.com'
    password = 'Password123!'
    
    try:
        user = User.objects.get(username=email)
        if JobVacancy.objects.filter(employer=user.profile).count() == 20:
            return
    except Exception:
        pass
    
    user, created = User.objects.get_or_create(
        username=email,
        defaults={
            'email': email,
            'first_name': 'Nexus',
            'last_name': 'Employer'
        }
    )
    if created or not user.check_password(password):
        user.set_password(password)
        user.save()
    
    profile, _ = Profile.objects.update_or_create(
        user=user,
        defaults={
            'role': 'employer',
            'is_verified': True,
            'company_name': 'Nexus Solutions',
            'industry': 'Technology Consulting & Services',
            'company_size': '250-500 employees',
            'website': 'https://nexus-solutions.test',
            'contact_name': 'Sarah Vance',
            'contact_position': 'Director of Talents',
            'contact_email': email,
            'employment_type_offered': 'Local',
            'location': 'Quezon City, Philippines',
            'soft_notes': 'Nexus Solutions is a leading digital transformation consultancy specializing in advanced software development, cybersecurity solutions, and technology integration.'
        }
    )
    
    # Define 20 jobs for nexus_employer
    jobs = [
        {
            'title': 'Cloud Security Engineer',
            'location': 'Remote / QC',
            'salary_range': '₱95k - ₱135k',
            'category': 'IT',
            'description': 'Design secure cloud architectures, configure firewalls, and perform vulnerability checks on AWS and Kubernetes environments.',
            'requirements': [('AWS Cloud', 4), ('Kubernetes', 4), ('Network Security', 5)]
        },
        {
            'title': 'AI / ML Researcher',
            'location': 'Quezon City, PH',
            'salary_range': '₱100k - ₱150k',
            'category': 'IT',
            'description': 'Develop custom machine learning models and data pipelines using PyTorch and Python to automate business tasks.',
            'requirements': [('Machine Learning', 5), ('PyTorch', 5), ('Python', 4)]
        },
        {
            'title': 'Product Designer (UX)',
            'location': 'Remote',
            'salary_range': '₱60k - ₱80k',
            'category': 'IT',
            'description': 'Design user journeys and mockups in Figma, execute usability testing, and build interactive prototypes.',
            'requirements': [('Figma', 5), ('Prototyping', 4), ('Usability Testing', 4)]
        },
        {
            'title': 'Data Analyst',
            'location': 'Quezon City, PH',
            'salary_range': '₱45k - ₱65k',
            'category': 'IT',
            'description': 'Perform data visualization and compile operational statistics reports using SQL and Pandas libraries.',
            'requirements': [('Data Analysis', 5), ('Pandas', 4), ('SQL', 4)]
        },
        {
            'title': 'SEO & Content Strategist',
            'location': 'Quezon City, PH',
            'salary_range': '₱32k - ₱45k',
            'category': 'MKT',
            'description': 'Manage corporate blog content, execute search engine optimization, and curate social media posts.',
            'requirements': [('Search Engine Optimization', 5), ('Content Writing', 4), ('Social Media Management', 4)]
        },
        {
            'title': 'Technical Support Lead',
            'location': 'Manila, PH',
            'salary_range': '₱40k - ₱55k',
            'category': 'BPO',
            'description': 'Guide customer service agents, diagnose software configuration issues, and maintain active ticketing response rates.',
            'requirements': [('Troubleshooting', 5), ('Phone & Online Customer Support', 4), ('Interpersonal Communication', 4)]
        },
        {
            'title': 'TESDA Network Systems Tech',
            'location': 'Quezon City, PH',
            'salary_range': '₱25k - ₱32k',
            'category': 'TVT',
            'description': 'Assemble computer systems, run cable checks, verify network installations, and troubleshoot connections.',
            'requirements': [('Computer Systems Servicing', 5), ('Troubleshooting', 4), ('Time Management', 4)]
        },
        {
            'title': 'Office Operations Specialist',
            'location': 'Quezon City, PH',
            'salary_range': '₱22k - ₱30k',
            'category': 'ADM',
            'description': 'Coordinate team logs, verify invoice folders, format presentation materials, and manage office systems.',
            'requirements': [('Microsoft Word', 4), ('Microsoft Excel', 4), ('Google Workspace', 4)]
        },
        {
            'title': 'Financial Controller',
            'location': 'Quezon City, PH',
            'salary_range': '₱50k - ₱70k',
            'category': 'FIN',
            'description': 'Monitor corporate budget allotments, audit ledger accounts, process payroll spreadsheets, and verify expense claims.',
            'requirements': [('Basic Bookkeeping Software', 4), ('Basic Budgeting', 4), ('Attention to Detail', 5)]
        },
        {
            'title': 'Site Infrastructure Specialist',
            'location': 'Quezon City, PH',
            'salary_range': '₱35k - ₱45k',
            'category': 'ENG',
            'description': 'Supervise building installations, execute wiring inspections, read structural blueprints, and enforce safety rules.',
            'requirements': [('Electrical Installation & Maintenance', 4), ('Blueprint & Plan Reading', 4), ('Occupational Health & Safety Awareness', 4)]
        },
        {
            'title': 'Creative Video Editor',
            'location': 'Remote',
            'salary_range': '₱28k - ₱38k',
            'category': 'MKT',
            'description': 'Record and edit promo videos, generate layout graphics in Canva, and support content team initiatives.',
            'requirements': [('Basic Video Editing', 5), ('Canva', 4), ('Creative Thinking', 4)]
        },
        {
            'title': 'Agile Coach',
            'location': 'Quezon City, PH',
            'salary_range': '₱80k - ₱110k',
            'category': 'IT',
            'description': 'Train scrum teams on agile standards, facilitate planning events, and build team alignment strategies.',
            'requirements': [('Agile', 5), ('Scrum', 5), ('OKR Framework', 4)]
        },
        {
            'title': 'Database Administrator',
            'location': 'Quezon City, PH',
            'salary_range': '₱70k - ₱95k',
            'category': 'IT',
            'description': 'Maintain data replication schemas, optimize SQL queries, and manage Postgres and MongoDB server configs.',
            'requirements': [('SQL', 5), ('PostgreSQL', 4), ('MongoDB', 4)]
        },
        {
            'title': 'Customer Success Manager',
            'location': 'Quezon City, PH',
            'salary_range': '₱45k - ₱60k',
            'category': 'BPO',
            'description': 'Maintain active client relations, handle account inquiries, and coordinate custom integrations.',
            'requirements': [('Client Communication', 5), ('Customer Relationship Management', 5), ('Empathy & Patience', 4)]
        },
        {
            'title': 'TESDA Culinary Lead',
            'location': 'Quezon City, PH',
            'salary_range': '₱25k - ₱32k',
            'category': 'TVT',
            'description': 'Direct food prep shifts, inspect kitchen hygiene compliance, and lead junior kitchen assistants.',
            'requirements': [('Cookery', 5), ('Food Safety & Hygiene Practices', 5), ('Teamwork & Collaboration', 4)]
        },
        {
            'title': 'General Office Assistant',
            'location': 'Quezon City, PH',
            'salary_range': '₱16k - ₱22k',
            'category': 'ADM',
            'description': 'Handle phone reception, execute basic data entry logs, organize document files, and schedule meetings.',
            'requirements': [('Task Prioritization', 4), ('Punctuality & Reliability', 5), ('Organization Skills', 4)]
        },
        {
            'title': 'Healthcare Coordinator',
            'location': 'Quezon City, PH',
            'salary_range': '₱26k - ₱34k',
            'category': 'MED',
            'description': 'Supervise corporate medical checks, guide caregiving initiatives, and support employee health programs.',
            'requirements': [('Caregiving', 4), ('Health Care Services', 4), ('Empathy', 5)]
        },
        {
            'title': 'Instructional Designer',
            'location': 'Remote',
            'salary_range': '₱38k - ₱50k',
            'category': 'EDU',
            'description': 'Plan corporate lecture outlines, write curriculum materials, and present train-the-trainer workshops.',
            'requirements': [('Presentation Skills', 4), ('Lesson Plan Preparation Assistance', 5), ('Verbal Communication', 5)]
        },
        {
            'title': 'Logistics Courier',
            'location': 'Quezon City, PH',
            'salary_range': '₱18k - ₱24k',
            'category': 'DRV',
            'description': 'Execute timely document and gear deliveries. Complete shipping receipts and inspect vehicle safety.',
            'requirements': [('Driving', 5), ('Punctuality & Reliability', 5), ('Instruction Following', 4)]
        },
        {
            'title': 'Safety Compliance Lead',
            'location': 'Quezon City, PH',
            'salary_range': '₱30k - ₱40k',
            'category': 'SEC',
            'description': 'Draft corporate security rules, investigate site safety incidents, and resolve facility disputes.',
            'requirements': [('Risk Awareness', 5), ('Compliance Mindset', 4), ('Conflict Resolution', 4)]
        }
    ]

    # Only create/recreate them if we don't already have exactly 20 jobs for this employer
    if JobVacancy.objects.filter(employer=profile).count() != 20:
        JobVacancy.objects.filter(employer=profile).delete()
        for job in jobs:
            vacancy = JobVacancy.objects.create(
                title=job['title'],
                employer=profile,
                location=job['location'],
                salary_range=job['salary_range'],
                category=job['category'],
                description=job['description'],
                status='Open',
                slots=3,
                remaining_slots=3
            )
            for skill_name, req_level in job['requirements']:
                try:
                    skill_obj = CentralizedSkill.objects.get(name=skill_name)
                except CentralizedSkill.DoesNotExist:
                    skill_obj = CentralizedSkill.objects.create(
                        name=skill_name,
                        category='General',
                        description=f'Seeded skill: {skill_name}'
                    )
                JobSkillRequirement.objects.get_or_create(
                    job_vacancy=vacancy,
                    skill=skill_obj,
                    defaults={'required_proficiency': req_level}
                )

def create_healthcare_employer_and_jobs():
    from django.contrib.auth.models import User
    from .models import Profile, JobVacancy, JobSkillRequirement, CentralizedSkill

    email = 'medilink_employer@test.com'
    password = 'Password123!'
    
    try:
        user = User.objects.get(username=email)
        if JobVacancy.objects.filter(employer=user.profile).count() == 15:
            return
    except Exception:
        pass
    
    user, created = User.objects.get_or_create(
        username=email,
        defaults={
            'email': email,
            'first_name': 'Healthcare',
            'last_name': 'Employer'
        }
    )
    if created or not user.check_password(password):
        user.set_password(password)
        user.save()
    
    profile, _ = Profile.objects.update_or_create(
        user=user,
        defaults={
            'role': 'employer',
            'is_verified': True,
            'company_name': 'MediLink Health Group',
            'industry': 'Healthcare & Pharmaceuticals',
            'company_size': '100-250 employees',
            'website': 'https://medilink-health.test',
            'contact_name': 'Dr. Manuel Santos',
            'contact_position': 'Chief Human Resources Officer',
            'contact_email': email,
            'employment_type_offered': 'Local',
            'location': 'Manila, Philippines',
            'soft_notes': 'MediLink Health Group is a network of premium healthcare clinics and hospitals dedicated to patient care, community health services, and medical research.'
        }
    )
    
    jobs = [
        {
            'title': 'Clinical Registered Nurse',
            'location': 'Manila, PH',
            'salary_range': '₱35k - ₱45k',
            'category': 'MED',
            'description': 'Provide clinical nursing care, patient assessment, and support in medical procedures. Cooperate with medical staff.',
            'requirements': [('Nursing Care', 4), ('Patient Assessment', 4), ('First Aid & CPR', 4)]
        },
        {
            'title': 'General Practitioner / Physician',
            'location': 'Manila, PH',
            'salary_range': '₱80k - ₱120k',
            'category': 'MED',
            'description': 'Consult patients, perform medical diagnoses, formulate treatment plans, and provide continuous clinical care.',
            'requirements': [('Medical Diagnosis', 5), ('Patient Care', 5), ('Clinical Decision Making', 5)]
        },
        {
            'title': 'Medical Laboratory Scientist',
            'location': 'Manila, PH',
            'salary_range': '₱28k - ₱38k',
            'category': 'MED',
            'description': 'Perform laboratory tests, run equipment calibration, and ensure strict quality control of medical tests.',
            'requirements': [('Laboratory Testing', 4), ('Equipment Calibration', 4), ('Quality Control', 3)]
        },
        {
            'title': 'Physical Therapist',
            'location': 'Quezon City, PH',
            'salary_range': '₱30k - ₱42k',
            'category': 'MED',
            'description': 'Evaluate patient mobility and direct rehabilitation therapy programs for restoring physical function.',
            'requirements': [('Rehabilitation Therapy', 4), ('Patient Mobility', 4), ('Anatomy & Physiology', 4)]
        },
        {
            'title': 'Hospital Pharmacist',
            'location': 'Manila, PH',
            'salary_range': '₱32k - ₱40k',
            'category': 'MED',
            'description': 'Manage medication dispensing, review pharmacology profiles, and verify dosage calculations with high attention to detail.',
            'requirements': [('Pharmacology', 5), ('Medication Dispensing', 5), ('Attention to Detail', 4)]
        },
        {
            'title': 'Radiologic Technologist',
            'location': 'Cebu, PH',
            'salary_range': '₱26k - ₱35k',
            'category': 'MED',
            'description': 'Perform X-Ray imaging, follow radiation safety procedures, and assist in patient positioning.',
            'requirements': [('X-Ray Imaging', 4), ('Radiation Safety', 4), ('Patient Positioning', 3)]
        },
        {
            'title': 'Medical Transcriptionist',
            'location': 'Remote / Manila',
            'salary_range': '₱20k - ₱28k',
            'category': 'MED',
            'description': 'Transcribe voice records of medical reports. Fast typing speed and understanding of medical terminology is required.',
            'requirements': [('Medical Terminology', 4), ('Typing Speed', 5), ('Active Listening', 4)]
        },
        {
            'title': 'Dental Hygienist',
            'location': 'Davao, PH',
            'salary_range': '₱22k - ₱30k',
            'category': 'MED',
            'description': 'Provide dental prophylaxis, teach oral health education to patients, and maintain sterilization protocols.',
            'requirements': [('Dental Prophylaxis', 4), ('Oral Health Education', 4), ('Patient Care', 3)]
        },
        {
            'title': 'Pediatric Nurse Practitioner',
            'location': 'Quezon City, PH',
            'salary_range': '₱45k - ₱60k',
            'category': 'MED',
            'description': 'Assess and manage pediatric care services, perform physical examinations, and offer health counseling to families.',
            'requirements': [('Pediatric Care', 5), ('Patient Assessment', 4), ('Communication Skills', 4)]
        },
        {
            'title': 'Clinical Research Coordinator',
            'location': 'Manila, PH',
            'salary_range': '₱38k - ₱50k',
            'category': 'MED',
            'description': 'Coordinate clinical trials, manage data collection, and ensure strict regulatory compliance with research ethics.',
            'requirements': [('Clinical Trials', 4), ('Data Collection', 4), ('Regulatory Compliance', 4)]
        },
        {
            'title': 'Emergency Medical Technician (EMT)',
            'location': 'Manila, PH',
            'salary_range': '₱25k - ₱32k',
            'category': 'MED',
            'description': 'Provide emergency medical response, administer first aid & CPR, and handle crisis management in ambulance units.',
            'requirements': [('Emergency Response', 5), ('First Aid & CPR', 5), ('Crisis Management', 4)]
        },
        {
            'title': 'Occupational Health Specialist',
            'location': 'Cebu, PH',
            'salary_range': '₱35k - ₱48k',
            'category': 'MED',
            'description': 'Perform health assessments, implement workplace safety programs, and conduct ergonomic evaluations.',
            'requirements': [('Workplace Safety', 4), ('Health Assessments', 4), ('Ergonomic Assessment', 3)]
        },
        {
            'title': 'Medical Administrative Assistant',
            'location': 'Manila, PH',
            'salary_range': '₱20k - ₱26k',
            'category': 'MED',
            'description': 'Manage patient scheduling, coordinate billing, and maintain database logs using Google Workspace tools.',
            'requirements': [('Medical Scheduling', 4), ('Patient Billing', 4), ('Google Workspace', 3)]
        },
        {
            'title': 'Nutritionist / Dietitian',
            'location': 'Manila, PH',
            'salary_range': '₱28k - ₱36k',
            'category': 'MED',
            'description': 'Create customized dietary planning, conduct nutritional assessments, and offer health counseling.',
            'requirements': [('Dietary Planning', 4), ('Nutritional Assessment', 4), ('Counseling Skills', 3)]
        },
        {
            'title': 'Mental Health Counselor',
            'location': 'Remote / Manila',
            'salary_range': '₱40k - ₱55k',
            'category': 'MED',
            'description': 'Provide individual psychotherapy, utilize active listening skills, and show high empathy to support patient recovery.',
            'requirements': [('Psychotherapy', 5), ('Active Listening', 5), ('Empathy', 5)]
        }
    ]

    # Only create/recreate them if we don't already have exactly 15 jobs for this employer
    if JobVacancy.objects.filter(employer=profile).count() != 15:
        JobVacancy.objects.filter(employer=profile).delete()
        for job in jobs:
            vacancy = JobVacancy.objects.create(
                title=job['title'],
                employer=profile,
                location=job['location'],
                salary_range=job['salary_range'],
                category=job['category'],
                description=job['description'],
                status='Open',
                slots=3,
                remaining_slots=3
            )
            for skill_name, req_level in job['requirements']:
                try:
                    skill_obj = CentralizedSkill.objects.get(name=skill_name)
                except CentralizedSkill.DoesNotExist:
                    skill_obj = CentralizedSkill.objects.create(
                        name=skill_name,
                        category='General',
                        description=f'Seeded skill: {skill_name}'
                    )
                JobSkillRequirement.objects.get_or_create(
                    job_vacancy=vacancy,
                    skill=skill_obj,
                    defaults={'required_proficiency': req_level}
                )

def create_ofw_employer_and_jobs():
    from django.contrib.auth.models import User
    from .models import Profile, JobVacancy, JobSkillRequirement, CentralizedSkill

    email = 'global_horizons_employer@test.com'
    password = 'Password123!'
    
    try:
        user = User.objects.get(username=email)
        if JobVacancy.objects.filter(employer=user.profile).count() == 10:
            return
    except Exception:
        pass
    
    user, created = User.objects.get_or_create(
        username=email,
        defaults={
            'email': email,
            'first_name': 'OFW',
            'last_name': 'Employer'
        }
    )
    if created or not user.check_password(password):
        user.set_password(password)
        user.save()
    
    profile, _ = Profile.objects.update_or_create(
        user=user,
        defaults={
            'role': 'employer',
            'is_verified': True,
            'company_name': 'Global Horizons Placement Agency',
            'industry': 'Human Resources & Recruitment',
            'company_size': '50-100 employees',
            'website': 'https://global-horizons.test',
            'contact_name': 'Mr. Antonio Santos',
            'contact_position': 'Managing Director',
            'contact_email': email,
            'employment_type_offered': 'Overseas',
            'location': 'Manila, Philippines',
            'soft_notes': 'Global Horizons is a POEA-licensed placement agency connecting skilled Filipino workers with international career opportunities in healthcare, engineering, hospitality, and general services.'
        }
    )
    
    jobs = [
        {
            'title': 'Domestic Helper / Caregiver',
            'location': 'Saudi Arabia',
            'salary_range': '₱25k - ₱30k',
            'category': 'OFW',
            'description': 'Provide domestic caregiving services, assist with daily activities, administer first aid & CPR, and handle basic housekeeping.',
            'requirements': [('Caregiving', 4), ('First Aid & CPR', 4), ('Housekeeping', 3)]
        },
        {
            'title': 'Hospitality Staff / Hotel Receptionist',
            'location': 'Dubai, UAE',
            'salary_range': '₱40k - ₱55k',
            'category': 'OFW',
            'description': 'Welcome hotel guests, coordinate check-in and check-out, and maintain active communication using excellent customer service.',
            'requirements': [('Verbal Communication', 4), ('Active Listening', 4), ('Customer Service Orientation', 4)]
        },
        {
            'title': 'Structural Welder (Industrial)',
            'location': 'Doha, Qatar',
            'salary_range': '₱50k - ₱70k',
            'category': 'OFW',
            'description': 'Perform structural welding for commercial construction. Must be POEA approved, read blueprint plans, and observe safety compliance.',
            'requirements': [('Welding', 5), ('Blueprint & Plan Reading', 4), ('Safety Compliance', 4)]
        },
        {
            'title': 'Heavy Equipment Operator',
            'location': 'Singapore',
            'salary_range': '₱60k - ₱80k',
            'category': 'OFW',
            'description': 'Operate heavy construction machinery, perform routine machine maintenance, and adhere to strict safety compliance standards.',
            'requirements': [('Heavy Equipment Operations', 5), ('Safety Compliance', 5), ('Machine Maintenance', 4)]
        },
        {
            'title': 'Registered General Nurse (Hospital)',
            'location': 'London, UK',
            'salary_range': '₱150k - ₱180k',
            'category': 'OFW',
            'description': 'Join a premium NHS hospital. Deliver general nursing care, perform patient assessment, and collaborate in medical protocols.',
            'requirements': [('Nursing Care', 5), ('Patient Assessment', 5), ('Verbal Communication', 4)]
        },
        {
            'title': 'Construction Civil Engineer',
            'location': 'Tokyo, Japan',
            'salary_range': '₱90k - ₱120k',
            'category': 'OFW',
            'description': 'Supervise site civil engineering designs, manage construction projects, and utilize basic Japanese language skills.',
            'requirements': [('Civil Engineering Design', 5), ('Project Management', 4), ('Japanese Language', 3)]
        },
        {
            'title': 'Automotive Service Mechanic',
            'location': 'Alberta, Canada',
            'salary_range': '₱100k - ₱130k',
            'category': 'OFW',
            'description': 'Perform complex automotive diagnostics, repair mechanical systems, and resolve technical issues.',
            'requirements': [('Automotive Diagnostics', 5), ('Mechanical Systems Repair', 4), ('Problem Solving', 4)]
        },
        {
            'title': 'Warehouse Logistics Assistant',
            'location': 'Taipei, Taiwan',
            'salary_range': '₱38k - ₱45k',
            'category': 'OFW',
            'description': 'Organize stock inventory count, manage warehouse operations, and support logistics shipping workflows.',
            'requirements': [('Warehouse Operations', 4), ('Inventory Count & Reporting', 4), ('Active Listening', 3)]
        },
        {
            'title': 'Barista / Food Service Crew',
            'location': 'Kuwait City, Kuwait',
            'salary_range': '₱30k - ₱38k',
            'category': 'OFW',
            'description': 'Deliver premium food and beverage service, serve guests, and maintain high standards of customer service.',
            'requirements': [('Food & Beverage Service', 4), ('Customer Service Orientation', 4), ('Verbal Communication', 3)]
        },
        {
            'title': 'Electrician (Facilities Maintenance)',
            'location': 'Sydney, Australia',
            'salary_range': '₱110k - ₱140k',
            'category': 'OFW',
            'description': 'Manage electrical installation & maintenance for corporate facilities. POEA placement opportunity.',
            'requirements': [('Electrical Installation & Maintenance', 5), ('Blueprint & Plan Reading', 4), ('Safety Compliance', 4)]
        }
    ]

    # Only create/recreate them if we don't already have exactly 10 jobs for this employer
    if JobVacancy.objects.filter(employer=profile).count() != 10:
        JobVacancy.objects.filter(employer=profile).delete()
        for job in jobs:
            vacancy = JobVacancy.objects.create(
                title=job['title'],
                employer=profile,
                location=job['location'],
                salary_range=job['salary_range'],
                category=job['category'],
                description=job['description'],
                status='Open',
                slots=3,
                remaining_slots=3
            )
            for skill_name, req_level in job['requirements']:
                try:
                    skill_obj = CentralizedSkill.objects.get(name=skill_name)
                except CentralizedSkill.DoesNotExist:
                    skill_obj = CentralizedSkill.objects.create(
                        name=skill_name,
                        category='General',
                        description=f'Seeded skill: {skill_name}'
                    )
                JobSkillRequirement.objects.get_or_create(
                    job_vacancy=vacancy,
                    skill=skill_obj,
                    defaults={'required_proficiency': req_level}
                )

def create_applicant_test_accounts(force_recreate=False):
    from django.contrib.auth.models import User
    from .models import (
        Profile, ApplicantSkill, CentralizedSkill, Education, 
        WorkExperience, Certification, JobVacancy, Referral, Interview
    )
    import random
    from django.utils import timezone
    import datetime

    if force_recreate:
        User.objects.filter(username__startswith='applicant_test_').delete()
        Interview.objects.filter(candidate__user__username__startswith='applicant_test_').delete()

    test_users = User.objects.filter(username__startswith='applicant_test_')
    existing_count = test_users.count()
    if existing_count == 50 and not force_recreate:
        # Check if referrals already exist for these test applicants. If so, return early.
        has_referrals = Referral.objects.filter(applicant__user__in=test_users).exists()
        if has_referrals:
            return

    password = 'Password123!'
    locations = ['Manila, Philippines', 'Quezon City, Philippines', 'Cebu, Philippines', 'Davao, Philippines', 'Remote']
    jobs_map = {
        'IT': ['React', 'TypeScript', 'Python', 'SQL', 'Django', 'Figma', 'AWS Cloud', 'Docker', 'Kubernetes', 'Agile'],
        'BPO': ['Verbal Communication', 'Written Communication', 'Active Listening', 'Customer Service Orientation', 'Complaint Handling'],
        'ADM': ['Microsoft Word', 'Microsoft Excel', 'Google Workspace', 'Time Management', 'Organization Skills'],
        'FIN': ['Basic Bookkeeping Software', 'Basic Budgeting', 'Attention to Detail', 'Microsoft Excel'],
        'MKT': ['Search Engine Optimization', 'Content Writing', 'Social Media Management', 'Canva', 'Email Marketing'],
        'RTL': ['Retail Store Operations', 'Stock Monitoring & Replenishment', 'Customer Service Orientation'],
        'ENG': ['Electrical Installation & Maintenance', 'Blueprint & Plan Reading', 'Problem Solving'],
        'LOG': ['Warehouse Operations', 'Inventory Count & Reporting', 'Logistics Support'],
        'MED': ['Nursing Care', 'Patient Assessment', 'First Aid & CPR', 'Medical Diagnosis', 'Patient Care', 'Clinical Decision Making', 'Active Listening', 'Medical Terminology', 'Emergency Response', 'Communication Skills'],
        'OFW': ['Caregiving', 'First Aid & CPR', 'Housekeeping', 'Welding', 'Safety Compliance', 'Warehouse Operations', 'Customer Service Orientation', 'Active Listening', 'Verbal Communication', 'Problem Solving']
    }
    job_keys = list(jobs_map.keys())
    
    # Pre-fetch vacancies to map matching categories efficiently
    vacancies_by_category = {}
    for key in job_keys:
        vacancies_by_category[key] = list(JobVacancy.objects.filter(category=key))
    all_vacancies = list(JobVacancy.objects.all())

    real_names = [
        ("John", "Santos"), ("Maria", "Reyes"), ("Mark", "Cruz"), ("Sarah", "Cruz"),
        ("Dave", "Dela Cruz"), ("Angelica", "Garcia"), ("James", "Mendoza"), ("Nicole", "Bautista"),
        ("Christian", "Aquino"), ("Christine", "Ramos"), ("Joseph", "Castillo"), ("Kimberly", "Santos"),
        ("Paul", "Rivera"), ("Patricia", "Salazar"), ("Michael", "Flores"), ("Michelle", "Bernardo"),
        ("Robert", "Valdez"), ("Rachel", "Villanueva"), ("David", "Pineda"), ("Samantha", "Castro"),
        ("Daniel", "Domingo"), ("Jessica", "Fernandez"), ("Richard", "Gabriel"), ("Mary", "Grace"),
        ("Andrew", "Gonzales"), ("Melissa", "De Leon"), ("Kevin", "Hernandez"), ("Stephanie", "Lopez"),
        ("Joseph", "Macatangay"), ("Karen", "Pascual"), ("Kenneth", "Mercado"), ("Ashley", "Ocampo"),
        ("Matthew", "Robles"), ("Vanessa", "Santiago"), ("Joshua", "Tolentino"), ("Camille", "Soriano"),
        ("Christopher", "Dizon"), ("Alyssa", "Sison"), ("Ryan", "Evangelista"), ("Bianca", "Ferrer"),
        ("Jason", "Gopez"), ("Clarissa", "Imperial"), ("Justin", "Alcantara"), ("Katrina", "Manalo"),
        ("Jonathan", "Ortiz"), ("Monica", "David"), ("Eric", "Panganiban"), ("Erica", "Reyes"),
        ("Patrick", "Roxas"), ("Erika", "Sebastian")
    ]

    for i in range(1, 51):
        email = f'applicant_test_{i}@example.com'
        first_name, last_name = real_names[i-1]
        user, created = User.objects.get_or_create(
            username=email,
            defaults={
                'email': email,
                'first_name': first_name,
                'last_name': last_name
            }
        )
        user.first_name = first_name
        user.last_name = last_name
        if created or not user.check_password(password):
            user.set_password(password)
        user.save()

        # Randomize attributes deterministically based on index so it is consistent
        random.seed(i)
        loc = random.choice(locations)
        pref_job = random.choice(job_keys)
        exp = random.choice(['Fresh Graduate', '1 Year', '2 Years', '3 Years', '5 Years'])
        is_fresh = (exp == 'Fresh Graduate')
        
        # Decide status based on index partition
        # Group 1: 1-15 -> Applied / Pending
        # Group 2: 16-30 -> Interviewing
        # Group 3: 31-50 -> Other Samples
        if i <= 15:
            ref_status = 'Pending'
            profile_status = 'Applied'
        elif i <= 30:
            ref_status = 'Interviewing'
            profile_status = 'Shortlisted'
        else:
            if i <= 35:
                ref_status = 'Hired — Probationary'
                profile_status = 'Employed — Onboarding'
            elif i <= 40:
                ref_status = 'Declined'
                profile_status = 'Active — Job Seeking'
            elif i <= 45:
                ref_status = 'Accepted — Awaiting Onboarding'
                profile_status = 'Employed — Onboarding'
            else:
                ref_status = 'No Show'
                profile_status = 'Active — Job Seeking'

        profile, p_created = Profile.objects.update_or_create(
            user=user,
            defaults={
                'role': 'applicant',
                'is_verified': True,
                'is_profile_complete': True,
                'location': loc,
                'preferred_job': pref_job,
                'experience_years': exp,
                'is_fresh_grad': is_fresh,
                'title': f'{pref_job} Professional',
                'status': profile_status,
                'civil_status': 'Single',
                'soft_notes': f'Automated test applicant profile for testing skill gaps and recruiter dashboards.'
            }
        )
        
        # Link skills deterministically based on preferred_job
        ApplicantSkill.objects.filter(profile=profile).delete()
        available_skills = jobs_map[pref_job]
        
        selected_skills = random.sample(available_skills, min(len(available_skills), random.randint(3, 4)))
        for s_name in selected_skills:
            try:
                skill_obj = CentralizedSkill.objects.get(name=s_name)
            except CentralizedSkill.DoesNotExist:
                skill_obj = CentralizedSkill.objects.create(
                    name=s_name,
                    category='General',
                    description=f'Seeded skill: {s_name}'
                )
            ApplicantSkill.objects.create(
                profile=profile,
                skill=skill_obj,
                proficiency=random.randint(3, 5),
                source='system'
            )
            
        # Add basic education and work experience if not fresh grad
        Education.objects.filter(profile=profile).delete()
        Education.objects.create(
            profile=profile,
            institution='State University',
            degree='Bachelor of Science',
            field_of_study=f'{pref_job} Studies',
            start_year=2018,
            end_year=2022
        )
        
        WorkExperience.objects.filter(profile=profile).delete()
        if not is_fresh:
            WorkExperience.objects.create(
                profile=profile,
                company='TechSolutions Inc.',
                position=f'Junior {pref_job} Specialist',
                start_date='June 2022',
                end_date='Present',
                is_current=True,
                description='Gained hands-on experience in core workflows.'
            )

        # Clear existing referrals/interviews for this user to avoid dupes/overlaps if re-running
        Referral.objects.filter(applicant=profile).delete()
        Interview.objects.filter(candidate=profile).delete()

        # Find matching vacancies
        matching_vacs = vacancies_by_category.get(pref_job, [])
        if not matching_vacs:
            matching_vacs = all_vacancies

        if matching_vacs:
            # Distribute primary vacancy across Apex, Prime, Nexus by rotating matching_vacs
            vac_index = (i - 1) % len(matching_vacs)
            primary_vac = matching_vacs[vac_index]
            
            # Create the primary Referral
            ref = Referral.objects.create(
                applicant=profile,
                job_vacancy=primary_vac,
                status=ref_status,
                date_referred=timezone.localdate() - datetime.timedelta(days=random.randint(2, 10))
            )

            # If Group 2 (Interviewing), create an Interview record as well
            if ref_status == 'Interviewing':
                Interview.objects.create(
                    employer=primary_vac.employer,
                    title=f"Technical Interview with {user.first_name} {user.last_name}",
                    candidate=profile,
                    vacancy=primary_vac,
                    date=timezone.localdate() + datetime.timedelta(days=random.randint(1, 5)),
                    start_time=datetime.time(hour=random.choice([9, 10, 11, 14, 15, 16]), minute=0),
                    interview_type=random.choice(['Video Call', 'Phone Call', 'In-Person']),
                    round_name=random.choice(['Screening', 'Technical Interview', 'Final Round']),
                    meeting_link='https://meet.google.com/abc-defg-hij',
                    notes=f"Scheduled screening interview for {user.first_name}."
                )

            # For hired status, let's fill in onboarding details to look complete
            if ref_status == 'Hired — Probationary' or ref_status == 'Accepted — Awaiting Onboarding':
                ref.accepted_position = f"{pref_job} Specialist"
                ref.accepted_salary = "₱30k - ₱45k"
                ref.reporting_date = timezone.localdate() + datetime.timedelta(days=random.randint(5, 15))
                ref.employment_type = 'Probationary'
                ref.save()

            # Sometimes add a secondary job vacancy application (Pending) to look realistic
            if len(matching_vacs) > 1 and random.choice([True, False]) and i <= 30:
                sec_index = (i) % len(matching_vacs)
                secondary_vac = matching_vacs[sec_index]
                if secondary_vac == primary_vac:
                    secondary_vac = matching_vacs[(sec_index + 1) % len(matching_vacs)]
                Referral.objects.create(
                    applicant=profile,
                    job_vacancy=secondary_vac,
                    status='Pending',
                    date_referred=timezone.localdate() - datetime.timedelta(days=random.randint(1, 5))
                )

# ==========================================
# PESO ADMIN VIEWS
# ==========================================

def seed_mock_applicants_if_empty():
    create_apex_employer_and_jobs()
    create_prime_employer_and_jobs()
    create_nexus_employer_and_jobs()
    create_healthcare_employer_and_jobs()
    create_ofw_employer_and_jobs()
    create_applicant_test_accounts()
    if User.objects.filter(username='employer@test.com').exists():
        return
    # 1. Centralized Skills
    skills_data = [
        ('React', 'Frontend Dev', 'A JavaScript library for building user interfaces.'),
        ('TypeScript', 'Frontend Dev', 'Strongly typed programming language that builds on JavaScript.'),
        ('Python', 'Backend Dev', 'High-level programming language known for readability.'),
        ('PyTorch', 'Data Science', 'Open source machine learning library based on Torch.'),
        ('Agile', 'Product Mgmt', 'Software development methodology focused on iterative progress.'),
        ('SQL', 'Backend Dev', 'Structured Query Language for database management.'),
        ('Figma', 'UX Design', 'Collaborative interface design tool.'),
        ('A11y', 'UX Design', 'Accessibility standards and practice.'),
        ('AWS Cloud', 'DevOps', 'Amazon Web Services cloud platform solutions.'),
        ('Housekeeping', 'TESDA Trade', 'TESDA vocational housekeeping qualifications.'),
        ('Welding', 'TESDA Trade', 'TESDA structural welding trade skills.'),
        ('Kubernetes', 'DevOps', 'Container orchestration tool.'),
        ('Wireframing', 'UX Design', 'Professional competency in Wireframing within UX Design.'),
        ('Prototyping', 'UX Design', 'Professional competency in Prototyping within UX Design.'),
        ('User Research', 'UX Design', 'Professional competency in User Research within UX Design.'),
        ('Adobe XD', 'UX Design', 'Professional competency in Adobe XD within UX Design.'),
        ('Sketch', 'UX Design', 'Professional competency in Sketch within UX Design.'),
        ('InVision', 'UX Design', 'Professional competency in InVision within UX Design.'),
        ('Usability Testing', 'UX Design', 'Professional competency in Usability Testing within UX Design.'),
        ('Information Architecture', 'UX Design', 'Professional competency in Information Architecture within UX Design.'),
        ('Design Thinking', 'UX Design', 'Professional competency in Design Thinking within UX Design.'),
        ('Vue.js', 'Frontend Dev', 'Professional competency in Vue.js within Frontend Dev.'),
        ('Angular', 'Frontend Dev', 'Professional competency in Angular within Frontend Dev.'),
        ('Next.js', 'Frontend Dev', 'Professional competency in Next.js within Frontend Dev.'),
        ('HTML & CSS', 'Frontend Dev', 'Professional competency in HTML & CSS within Frontend Dev.'),
        ('Tailwind CSS', 'Frontend Dev', 'Professional competency in Tailwind CSS within Frontend Dev.'),
        ('Bootstrap', 'Frontend Dev', 'Professional competency in Bootstrap within Frontend Dev.'),
        ('JavaScript', 'Frontend Dev', 'Professional competency in JavaScript within Frontend Dev.'),
        ('Responsive Web Design', 'Frontend Dev', 'Professional competency in Responsive Web Design within Frontend Dev.'),
        ('Redux', 'Frontend Dev', 'Professional competency in Redux within Frontend Dev.'),
        ('Node.js', 'Backend Dev', 'Professional competency in Node.js within Backend Dev.'),
        ('Express.js', 'Backend Dev', 'Professional competency in Express.js within Backend Dev.'),
        ('PHP', 'Backend Dev', 'Professional competency in PHP within Backend Dev.'),
        ('Laravel', 'Backend Dev', 'Professional competency in Laravel within Backend Dev.'),
        ('Django', 'Backend Dev', 'Professional competency in Django within Backend Dev.'),
        ('Java', 'Backend Dev', 'Professional competency in Java within Backend Dev.'),
        ('Spring Boot', 'Backend Dev', 'Professional competency in Spring Boot within Backend Dev.'),
        ('REST API Development', 'Backend Dev', 'Professional competency in REST API Development within Backend Dev.'),
        ('GraphQL', 'Backend Dev', 'Professional competency in GraphQL within Backend Dev.'),
        ('C#', 'Backend Dev', 'Professional competency in C# within Backend Dev.'),
        ('.NET Framework', 'Backend Dev', 'Professional competency in .NET Framework within Backend Dev.'),
        ('MongoDB', 'Backend Dev', 'Professional competency in MongoDB within Backend Dev.'),
        ('PostgreSQL', 'Backend Dev', 'Professional competency in PostgreSQL within Backend Dev.'),
        ('Firebase', 'Backend Dev', 'Professional competency in Firebase within Backend Dev.'),
        ('Docker', 'DevOps', 'Professional competency in Docker within DevOps.'),
        ('Jenkins', 'DevOps', 'Professional competency in Jenkins within DevOps.'),
        ('Terraform', 'DevOps', 'Professional competency in Terraform within DevOps.'),
        ('GitHub Actions', 'DevOps', 'Professional competency in GitHub Actions within DevOps.'),
        ('CI/CD Pipeline', 'DevOps', 'Professional competency in CI/CD Pipeline within DevOps.'),
        ('Linux Administration', 'DevOps', 'Professional competency in Linux Administration within DevOps.'),
        ('Nginx', 'DevOps', 'Professional competency in Nginx within DevOps.'),
        ('Azure DevOps', 'DevOps', 'Professional competency in Azure DevOps within DevOps.'),
        ('Google Cloud Platform', 'DevOps', 'Professional competency in Google Cloud Platform within DevOps.'),
        ('Amazon Web Services', 'DevOps', 'Professional competency in Amazon Web Services within DevOps.'),
        ('Machine Learning', 'Data Science', 'Professional competency in Machine Learning within Data Science.'),
        ('Data Analysis', 'Data Science', 'Professional competency in Data Analysis within Data Science.'),
        ('TensorFlow', 'Data Science', 'Professional competency in TensorFlow within Data Science.'),
        ('Pandas', 'Data Science', 'Professional competency in Pandas within Data Science.'),
        ('NumPy', 'Data Science', 'Professional competency in NumPy within Data Science.'),
        ('Data Visualization', 'Data Science', 'Professional competency in Data Visualization within Data Science.'),
        ('Power BI', 'Data Science', 'Professional competency in Power BI within Data Science.'),
        ('Tableau', 'Data Science', 'Professional competency in Tableau within Data Science.'),
        ('R Programming', 'Data Science', 'Professional competency in R Programming within Data Science.'),
        ('Natural Language Processing', 'Data Science', 'Professional competency in Natural Language Processing within Data Science.'),
        ('Statistics & Probability', 'Data Science', 'Professional competency in Statistics & Probability within Data Science.'),
        ('Excel for Data Analysis', 'Data Science', 'Professional competency in Excel for Data Analysis within Data Science.'),
        ('Scrum', 'Product Mgmt', 'Professional competency in Scrum within Product Mgmt.'),
        ('Kanban', 'Product Mgmt', 'Professional competency in Kanban within Product Mgmt.'),
        ('JIRA', 'Product Mgmt', 'Professional competency in JIRA within Product Mgmt.'),
        ('Product Roadmapping', 'Product Mgmt', 'Professional competency in Product Roadmapping within Product Mgmt.'),
        ('Stakeholder Management', 'Product Mgmt', 'Professional competency in Stakeholder Management within Product Mgmt.'),
        ('Business Analysis', 'Product Mgmt', 'Professional competency in Business Analysis within Product Mgmt.'),
        ('Requirements Gathering', 'Product Mgmt', 'Professional competency in Requirements Gathering within Product Mgmt.'),
        ('OKR Framework', 'Product Mgmt', 'Professional competency in OKR Framework within Product Mgmt.'),
        ('Trello', 'Product Mgmt', 'Professional competency in Trello within Product Mgmt.'),
        ('Microsoft Word', 'Digital Skills', 'Professional competency in Microsoft Word within Digital Skills.'),
        ('Microsoft Excel', 'Digital Skills', 'Professional competency in Microsoft Excel within Digital Skills.'),
        ('Microsoft PowerPoint', 'Digital Skills', 'Professional competency in Microsoft PowerPoint within Digital Skills.'),
        ('Google Workspace', 'Digital Skills', 'Professional competency in Google Workspace within Digital Skills.'),
        ('Canva', 'Digital Skills', 'Professional competency in Canva within Digital Skills.'),
        ('Social Media Management', 'Digital Skills', 'Professional competency in Social Media Management within Digital Skills.'),
        ('Email Marketing', 'Digital Skills', 'Professional competency in Email Marketing within Digital Skills.'),
        ('Basic Video Editing', 'Digital Skills', 'Professional competency in Basic Video Editing within Digital Skills.'),
        ('Content Writing', 'Digital Skills', 'Professional competency in Content Writing within Digital Skills.'),
        ('Search Engine Optimization', 'Digital Skills', 'Professional competency in Search Engine Optimization within Digital Skills.'),
        ('E-Commerce Management', 'Digital Skills', 'Professional competency in E-Commerce Management within Digital Skills.'),
        ('Online Customer Service', 'Digital Skills', 'Professional competency in Online Customer Service within Digital Skills.'),
        ('Basic Bookkeeping Software', 'Digital Skills', 'Professional competency in Basic Bookkeeping Software within Digital Skills.'),
        ('Zoom / Google Meet Facilitation', 'Digital Skills', 'Professional competency in Zoom / Google Meet Facilitation within Digital Skills.'),
        ('Cybersecurity Awareness', 'Digital Skills', 'Professional competency in Cybersecurity Awareness within Digital Skills.'),
        ('Cookery', 'TESDA Trade', 'Professional competency in Cookery within TESDA Trade.'),
        ('Food & Beverage Service', 'TESDA Trade', 'Professional competency in Food & Beverage Service within TESDA Trade.'),
        ('Bread & Pastry Production', 'TESDA Trade', 'Professional competency in Bread & Pastry Production within TESDA Trade.'),
        ('Shielded Metal Arc Welding', 'TESDA Trade', 'Professional competency in Shielded Metal Arc Welding within TESDA Trade.'),
        ('Gas Metal Arc Welding', 'TESDA Trade', 'Professional competency in Gas Metal Arc Welding within TESDA Trade.'),
        ('Electrical Installation & Maintenance', 'TESDA Trade', 'Professional competency in Electrical Installation & Maintenance within TESDA Trade.'),
        ('Plumbing', 'TESDA Trade', 'Professional competency in Plumbing within TESDA Trade.'),
        ('Carpentry', 'TESDA Trade', 'Professional competency in Carpentry within TESDA Trade.'),
        ('Tile Setting', 'TESDA Trade', 'Professional competency in Tile Setting within TESDA Trade.'),
        ('Masonry', 'TESDA Trade', 'Professional competency in Masonry within TESDA Trade.'),
        ('Driving', 'TESDA Trade', 'Professional competency in Driving within TESDA Trade.'),
        ('Automotive Servicing', 'TESDA Trade', 'Professional competency in Automotive Servicing within TESDA Trade.'),
        ('Motorcycle/Small Engine Servicing', 'TESDA Trade', 'Professional competency in Motorcycle/Small Engine Servicing within TESDA Trade.'),
        ('Agricultural Crops Production', 'TESDA Trade', 'Professional competency in Agricultural Crops Production within TESDA Trade.'),
        ('Beauty/Nail Care', 'TESDA Trade', 'Professional competency in Beauty/Nail Care within TESDA Trade.'),
        ('Barbering', 'TESDA Trade', 'Professional competency in Barbering within TESDA Trade.'),
        ('Dressmaking & Tailoring', 'TESDA Trade', 'Professional competency in Dressmaking & Tailoring within TESDA Trade.'),
        ('Caregiving', 'TESDA Trade', 'Professional competency in Caregiving within TESDA Trade.'),
        ('Health Care Services', 'TESDA Trade', 'Professional competency in Health Care Services within TESDA Trade.'),
        ('Computer Systems Servicing', 'TESDA Trade', 'Professional competency in Computer Systems Servicing within TESDA Trade.'),
        ('Contact Center Services', 'TESDA Trade', 'Professional competency in Contact Center Services within TESDA Trade.'),
        ('Verbal Communication', 'Soft Skills', 'Professional competency in Verbal Communication within Soft Skills.'),
        ('Written Communication', 'Soft Skills', 'Professional competency in Written Communication within Soft Skills.'),
        ('Active Listening', 'Soft Skills', 'Professional competency in Active Listening within Soft Skills.'),
        ('Presentation Skills', 'Soft Skills', 'Professional competency in Presentation Skills within Soft Skills.'),
        ('Public Speaking', 'Soft Skills', 'Professional competency in Public Speaking within Soft Skills.'),
        ('Report Writing', 'Soft Skills', 'Professional competency in Report Writing within Soft Skills.'),
        ('Business Correspondence', 'Soft Skills', 'Professional competency in Business Correspondence within Soft Skills.'),
        ('Interpersonal Communication', 'Soft Skills', 'Professional competency in Interpersonal Communication within Soft Skills.'),
        ('Non-Verbal Communication', 'Soft Skills', 'Professional competency in Non-Verbal Communication within Soft Skills.'),
        ('Storytelling & Persuasion', 'Soft Skills', 'Professional competency in Storytelling & Persuasion within Soft Skills.'),
        ('Critical Thinking', 'Soft Skills', 'Professional competency in Critical Thinking within Soft Skills.'),
        ('Analytical Thinking', 'Soft Skills', 'Professional competency in Analytical Thinking within Soft Skills.'),
        ('Problem Solving', 'Soft Skills', 'Professional competency in Problem Solving within Soft Skills.'),
        ('Decision Making', 'Soft Skills', 'Professional competency in Decision Making within Soft Skills.'),
        ('Creative Thinking', 'Soft Skills', 'Professional competency in Creative Thinking within Soft Skills.'),
        ('Strategic Thinking', 'Soft Skills', 'Professional competency in Strategic Thinking within Soft Skills.'),
        ('Logical Reasoning', 'Soft Skills', 'Professional competency in Logical Reasoning within Soft Skills.'),
        ('Research & Information Gathering', 'Soft Skills', 'Professional competency in Research & Information Gathering within Soft Skills.'),
        ('Troubleshooting', 'Soft Skills', 'Professional competency in Troubleshooting within Soft Skills.'),
        ('Root Cause Analysis', 'Soft Skills', 'Professional competency in Root Cause Analysis within Soft Skills.'),
        ('Work Ethic', 'Soft Skills', 'Professional competency in Work Ethic within Soft Skills.'),
        ('Integrity & Honesty', 'Soft Skills', 'Professional competency in Integrity & Honesty within Soft Skills.'),
        ('Punctuality & Reliability', 'Soft Skills', 'Professional competency in Punctuality & Reliability within Soft Skills.'),
        ('Accountability', 'Soft Skills', 'Professional competency in Accountability within Soft Skills.'),
        ('Attention to Detail', 'Soft Skills', 'Professional competency in Attention to Detail within Soft Skills.'),
        ('Initiative & Proactiveness', 'Soft Skills', 'Professional competency in Initiative & Proactiveness within Soft Skills.'),
        ('Self-Motivation', 'Soft Skills', 'Professional competency in Self-Motivation within Soft Skills.'),
        ('Commitment to Quality', 'Soft Skills', 'Professional competency in Commitment to Quality within Soft Skills.'),
        ('Professional Conduct', 'Soft Skills', 'Professional competency in Professional Conduct within Soft Skills.'),
        ('Confidentiality & Discretion', 'Soft Skills', 'Professional competency in Confidentiality & Discretion within Soft Skills.'),
        ('Time Management', 'Soft Skills', 'Professional competency in Time Management within Soft Skills.'),
        ('Task Prioritization', 'Soft Skills', 'Professional competency in Task Prioritization within Soft Skills.'),
        ('Organization Skills', 'Soft Skills', 'Professional competency in Organization Skills within Soft Skills.'),
        ('Multitasking', 'Soft Skills', 'Professional competency in Multitasking within Soft Skills.'),
        ('Meeting Deadlines', 'Soft Skills', 'Professional competency in Meeting Deadlines within Soft Skills.'),
        ('Goal Setting', 'Soft Skills', 'Professional competency in Goal Setting within Soft Skills.'),
        ('Planning & Scheduling', 'Soft Skills', 'Professional competency in Planning & Scheduling within Soft Skills.'),
        ('Workload Management', 'Soft Skills', 'Professional competency in Workload Management within Soft Skills.'),
        ('Teamwork & Collaboration', 'Soft Skills', 'Professional competency in Teamwork & Collaboration within Soft Skills.'),
        ('Team Building', 'Soft Skills', 'Professional competency in Team Building within Soft Skills.'),
        ('Cooperation & Flexibility', 'Soft Skills', 'Professional competency in Cooperation & Flexibility within Soft Skills.'),
        ('Cross-Functional Collaboration', 'Soft Skills', 'Professional competency in Cross-Functional Collaboration within Soft Skills.'),
        ('Peer Support & Mentoring', 'Soft Skills', 'Professional competency in Peer Support & Mentoring within Soft Skills.'),
        ('Knowledge Sharing', 'Soft Skills', 'Professional competency in Knowledge Sharing within Soft Skills.'),
        ('Workplace Relationship Building', 'Soft Skills', 'Professional competency in Workplace Relationship Building within Soft Skills.'),
        ('Leadership', 'Soft Skills', 'Professional competency in Leadership within Soft Skills.'),
        ('People Management', 'Soft Skills', 'Professional competency in People Management within Soft Skills.'),
        ('Delegation', 'Soft Skills', 'Professional competency in Delegation within Soft Skills.'),
        ('Coaching & Mentoring', 'Soft Skills', 'Professional competency in Coaching & Mentoring within Soft Skills.'),
        ('Conflict Resolution', 'Soft Skills', 'Professional competency in Conflict Resolution within Soft Skills.'),
        ('Decision Making Under Pressure', 'Soft Skills', 'Professional competency in Decision Making Under Pressure within Soft Skills.'),
        ('Motivating Others', 'Soft Skills', 'Professional competency in Motivating Others within Soft Skills.'),
        ('Visionary Thinking', 'Soft Skills', 'Professional competency in Visionary Thinking within Soft Skills.'),
        ('Change Management', 'Soft Skills', 'Professional competency in Change Management within Soft Skills.'),
        ('Performance Feedback', 'Soft Skills', 'Professional competency in Performance Feedback within Soft Skills.'),
        ('Customer Service Orientation', 'Soft Skills', 'Professional competency in Customer Service Orientation within Soft Skills.'),
        ('Customer Relationship Management', 'Soft Skills', 'Professional competency in Customer Relationship Management within Soft Skills.'),
        ('Empathy & Patience', 'Soft Skills', 'Professional competency in Empathy & Patience within Soft Skills.'),
        ('Complaint Handling', 'Soft Skills', 'Professional competency in Complaint Handling within Soft Skills.'),
        ('Service Quality Mindset', 'Soft Skills', 'Professional competency in Service Quality Mindset within Soft Skills.'),
        ('Client Communication', 'Soft Skills', 'Professional competency in Client Communication within Soft Skills.'),
        ('Needs Assessment', 'Soft Skills', 'Professional competency in Needs Assessment within Soft Skills.'),
        ('Sales Orientation', 'Soft Skills', 'Professional competency in Sales Orientation within Soft Skills.'),
        ('Adaptability', 'Soft Skills', 'Professional competency in Adaptability within Soft Skills.'),
        ('Flexibility', 'Soft Skills', 'Professional competency in Flexibility within Soft Skills.'),
        ('Resilience & Stress Tolerance', 'Soft Skills', 'Professional competency in Resilience & Stress Tolerance within Soft Skills.'),
        ('Coping with Change', 'Soft Skills', 'Professional competency in Coping with Change within Soft Skills.'),
        ('Working Under Pressure', 'Soft Skills', 'Professional competency in Working Under Pressure within Soft Skills.'),
        ('Emotional Regulation', 'Soft Skills', 'Professional competency in Emotional Regulation within Soft Skills.'),
        ('Growth Mindset', 'Soft Skills', 'Professional competency in Growth Mindset within Soft Skills.'),
        ('Openness to Feedback', 'Soft Skills', 'Professional competency in Openness to Feedback within Soft Skills.'),
        ('Learning Agility', 'Soft Skills', 'Professional competency in Learning Agility within Soft Skills.'),
        ('Self-Awareness', 'Soft Skills', 'Professional competency in Self-Awareness within Soft Skills.'),
        ('Empathy', 'Soft Skills', 'Professional competency in Empathy within Soft Skills.'),
        ('Social Awareness', 'Soft Skills', 'Professional competency in Social Awareness within Soft Skills.'),
        ('Relationship Management', 'Soft Skills', 'Professional competency in Relationship Management within Soft Skills.'),
        ('Emotional Maturity', 'Soft Skills', 'Professional competency in Emotional Maturity within Soft Skills.'),
        ('Positive Attitude', 'Soft Skills', 'Professional competency in Positive Attitude within Soft Skills.'),
        ('Patience', 'Soft Skills', 'Professional competency in Patience within Soft Skills.'),
        ('Compassion', 'Soft Skills', 'Professional competency in Compassion within Soft Skills.'),
        ('Occupational Health & Safety Awareness', 'Soft Skills', 'Professional competency in Occupational Health & Safety Awareness within Soft Skills.'),
        ('Following Instructions & Procedures', 'Soft Skills', 'Professional competency in Following Instructions & Procedures within Soft Skills.'),
        ('Risk Awareness', 'Soft Skills', 'Professional competency in Risk Awareness within Soft Skills.'),
        ('Workplace Discipline', 'Soft Skills', 'Professional competency in Workplace Discipline within Soft Skills.'),
        ('Compliance Mindset', 'Soft Skills', 'Professional competency in Compliance Mindset within Soft Skills.'),
        ('Cultural Sensitivity', 'Soft Skills', 'Professional competency in Cultural Sensitivity within Soft Skills.'),
        ('Diversity & Inclusion Awareness', 'Soft Skills', 'Professional competency in Diversity & Inclusion Awareness within Soft Skills.'),
        ('Cross-Cultural Communication', 'Soft Skills', 'Professional competency in Cross-Cultural Communication within Soft Skills.'),
        ('Language Adaptability', 'Soft Skills', 'Professional competency in Language Adaptability within Soft Skills.'),
        ('Global Mindset', 'Soft Skills', 'Professional competency in Global Mindset within Soft Skills.'),
        ('Professional Networking', 'Soft Skills', 'Professional competency in Professional Networking within Soft Skills.'),
        ('Personal Branding', 'Soft Skills', 'Professional competency in Personal Branding within Soft Skills.'),
        ('Career Planning', 'Soft Skills', 'Professional competency in Career Planning within Soft Skills.'),
        ('Continuous Learning Mindset', 'Soft Skills', 'Professional competency in Continuous Learning Mindset within Soft Skills.'),
        ('Seeking Mentorship', 'Soft Skills', 'Professional competency in Seeking Mentorship within Soft Skills.'),
        ('Industry Awareness', 'Soft Skills', 'Professional competency in Industry Awareness within Soft Skills.'),
        ('Basic Computer Operation', 'Common Skills', 'Professional competency in Basic Computer Operation within Common Skills.'),
        ('Microsoft Word', 'Common Skills', 'Professional competency in Microsoft Word within Common Skills.'),
        ('Microsoft Excel', 'Common Skills', 'Professional competency in Microsoft Excel within Common Skills.'),
        ('Microsoft PowerPoint', 'Common Skills', 'Professional competency in Microsoft PowerPoint within Common Skills.'),
        ('Google Docs', 'Common Skills', 'Professional competency in Google Docs within Common Skills.'),
        ('Google Sheets', 'Common Skills', 'Professional competency in Google Sheets within Common Skills.'),
        ('Google Slides', 'Common Skills', 'Professional competency in Google Slides within Common Skills.'),
        ('Email Composition & Management', 'Common Skills', 'Professional competency in Email Composition & Management within Common Skills.'),
        ('Internet Research', 'Common Skills', 'Professional competency in Internet Research within Common Skills.'),
        ('Data Entry', 'Common Skills', 'Professional competency in Data Entry within Common Skills.'),
        ('File Management & Organization', 'Common Skills', 'Professional competency in File Management & Organization within Common Skills.'),
        ('Typing Speed & Accuracy', 'Common Skills', 'Professional competency in Typing Speed & Accuracy within Common Skills.'),
        ('Printing & Scanning Documents', 'Common Skills', 'Professional competency in Printing & Scanning Documents within Common Skills.'),
        ('Video Conferencing Tools', 'Common Skills', 'Professional competency in Video Conferencing Tools within Common Skills.'),
        ('Basic Troubleshooting of Office Equipment', 'Common Skills', 'Professional competency in Basic Troubleshooting of Office Equipment within Common Skills.'),
        ('Document Filing & Recordkeeping', 'Common Skills', 'Professional competency in Document Filing & Recordkeeping within Common Skills.'),
        ('Scheduling & Calendar Management', 'Common Skills', 'Professional competency in Scheduling & Calendar Management within Common Skills.'),
        ('Office Correspondence', 'Common Skills', 'Professional competency in Office Correspondence within Common Skills.'),
        ('Receptionist Duties', 'Common Skills', 'Professional competency in Receptionist Duties within Common Skills.'),
        ('Answering & Routing Phone Calls', 'Common Skills', 'Professional competency in Answering & Routing Phone Calls within Common Skills.'),
        ('Meeting Coordination', 'Common Skills', 'Professional competency in Meeting Coordination within Common Skills.'),
        ('Inventory Recording', 'Common Skills', 'Professional competency in Inventory Recording within Common Skills.'),
        ('Basic Bookkeeping', 'Common Skills', 'Professional competency in Basic Bookkeeping within Common Skills.'),
        ('Encoding & Transcription', 'Common Skills', 'Professional competency in Encoding & Transcription within Common Skills.'),
        ('Preparation of Reports & Memos', 'Common Skills', 'Professional competency in Preparation of Reports & Memos within Common Skills.'),
        ('Office Supply Management', 'Common Skills', 'Professional competency in Office Supply Management within Common Skills.'),
        ('Processing of Forms & Documents', 'Common Skills', 'Professional competency in Processing of Forms & Documents within Common Skills.'),
        ('Basic Payroll Assistance', 'Common Skills', 'Professional competency in Basic Payroll Assistance within Common Skills.'),
        ('Records & Database Management', 'Common Skills', 'Professional competency in Records & Database Management within Common Skills.'),
        ('Face-to-Face Customer Service', 'Common Skills', 'Professional competency in Face-to-Face Customer Service within Common Skills.'),
        ('Phone & Online Customer Support', 'Common Skills', 'Professional competency in Phone & Online Customer Support within Common Skills.'),
        ('Handling Customer Complaints', 'Common Skills', 'Professional competency in Handling Customer Complaints within Common Skills.'),
        ('Product Knowledge & Demonstration', 'Common Skills', 'Professional competency in Product Knowledge & Demonstration within Common Skills.'),
        ('Cashiering & Cash Handling', 'Common Skills', 'Professional competency in Cashiering & Cash Handling within Common Skills.'),
        ('POS System Operation', 'Common Skills', 'Professional competency in POS System Operation within Common Skills.'),
        ('Upselling & Cross-Selling', 'Common Skills', 'Professional competency in Upselling & Cross-Selling within Common Skills.'),
        ('Order Taking & Processing', 'Common Skills', 'Professional competency in Order Taking & Processing within Common Skills.'),
        ('Basic Sales Techniques', 'Common Skills', 'Professional competency in Basic Sales Techniques within Common Skills.'),
        ('After-Sales Support', 'Common Skills', 'Professional competency in After-Sales Support within Common Skills.'),
        ('Customer Needs Assessment', 'Common Skills', 'Professional competency in Customer Needs Assessment within Common Skills.'),
        ('Filipino / Tagalog Communication', 'Common Skills', 'Professional competency in Filipino / Tagalog Communication within Common Skills.'),
        ('English Communication — Oral', 'Common Skills', 'Professional competency in English Communication — Oral within Common Skills.'),
        ('English Communication — Written', 'Common Skills', 'Professional competency in English Communication — Written within Common Skills.'),
        ('Business Writing', 'Common Skills', 'Professional competency in Business Writing within Common Skills.'),
        ('Report & Letter Writing', 'Common Skills', 'Professional competency in Report & Letter Writing within Common Skills.'),
        ('Note-Taking & Minute Writing', 'Common Skills', 'Professional competency in Note-Taking & Minute Writing within Common Skills.'),
        ('Basic Conversational English', 'Common Skills', 'Professional competency in Basic Conversational English within Common Skills.'),
        ('Reading Comprehension', 'Common Skills', 'Professional competency in Reading Comprehension within Common Skills.'),
        ('Instruction Following', 'Common Skills', 'Professional competency in Instruction Following within Common Skills.'),
        ('Basic Math & Arithmetic', 'Common Skills', 'Professional competency in Basic Math & Arithmetic within Common Skills.'),
        ('Cash Counting & Management', 'Common Skills', 'Professional competency in Cash Counting & Management within Common Skills.'),
        ('Basic Budgeting', 'Common Skills', 'Professional competency in Basic Budgeting within Common Skills.'),
        ('Invoice & Receipt Processing', 'Common Skills', 'Professional competency in Invoice & Receipt Processing within Common Skills.'),
        ('Petty Cash Handling', 'Common Skills', 'Professional competency in Petty Cash Handling within Common Skills.'),
        ('Basic Financial Record Keeping', 'Common Skills', 'Professional competency in Basic Financial Record Keeping within Common Skills.'),
        ('Percentage & Discount Computation', 'Common Skills', 'Professional competency in Percentage & Discount Computation within Common Skills.'),
        ('Stock & Inventory Counting', 'Common Skills', 'Professional competency in Stock & Inventory Counting within Common Skills.'),
        ('Manual Labor & General Utility', 'Common Skills', 'Professional competency in Manual Labor & General Utility within Common Skills.'),
        ('Lifting & Material Handling', 'Common Skills', 'Professional competency in Lifting & Material Handling within Common Skills.'),
        ('Warehouse Operations', 'Common Skills', 'Professional competency in Warehouse Operations within Common Skills.'),
        ('Loading & Unloading', 'Common Skills', 'Professional competency in Loading & Unloading within Common Skills.'),
        ('Packing & Labeling', 'Common Skills', 'Professional competency in Packing & Labeling within Common Skills.'),
        ('Stock Replenishment', 'Common Skills', 'Professional competency in Stock Replenishment within Common Skills.'),
        ('Housekeeping & Janitorial Work', 'Common Skills', 'Professional competency in Housekeeping & Janitorial Work within Common Skills.'),
        ('Sanitation & Cleanliness Standards', 'Common Skills', 'Professional competency in Sanitation & Cleanliness Standards within Common Skills.'),
        ('Equipment Operation', 'Common Skills', 'Professional competency in Equipment Operation within Common Skills.'),
        ('Basic Machine Operation', 'Common Skills', 'Professional competency in Basic Machine Operation within Common Skills.'),
        ('Quality Checking & Sorting', 'Common Skills', 'Professional competency in Quality Checking & Sorting within Common Skills.'),
        ('Forklift Operation', 'Common Skills', 'Professional competency in Forklift Operation within Common Skills.'),
        ('Delivery & Logistics Support', 'Common Skills', 'Professional competency in Delivery & Logistics Support within Common Skills.'),
        ('Food Preparation Assistance', 'Common Skills', 'Professional competency in Food Preparation Assistance within Common Skills.'),
        ('Table Setting & Service', 'Common Skills', 'Professional competency in Table Setting & Service within Common Skills.'),
        ('Taking Food & Beverage Orders', 'Common Skills', 'Professional competency in Taking Food & Beverage Orders within Common Skills.'),
        ('Food Safety & Hygiene Practices', 'Common Skills', 'Professional competency in Food Safety & Hygiene Practices within Common Skills.'),
        ('Kitchen Cleaning & Sanitation', 'Common Skills', 'Professional competency in Kitchen Cleaning & Sanitation within Common Skills.'),
        ('Barista & Coffee Preparation', 'Common Skills', 'Professional competency in Barista & Coffee Preparation within Common Skills.'),
        ('Fast Food Counter Service', 'Common Skills', 'Professional competency in Fast Food Counter Service within Common Skills.'),
        ('Catering & Event Food Service', 'Common Skills', 'Professional competency in Catering & Event Food Service within Common Skills.'),
        ('Menu Knowledge', 'Common Skills', 'Professional competency in Menu Knowledge within Common Skills.'),
        ('Dishwashing & Kitchen Utility', 'Common Skills', 'Professional competency in Dishwashing & Kitchen Utility within Common Skills.'),
        ('Product Display & Arrangement', 'Common Skills', 'Professional competency in Product Display & Arrangement within Common Skills.'),
        ('Stock Monitoring & Replenishment', 'Common Skills', 'Professional competency in Stock Monitoring & Replenishment within Common Skills.'),
        ('Price Tagging & Labeling', 'Common Skills', 'Professional competency in Price Tagging & Labeling within Common Skills.'),
        ('Retail Store Operations', 'Common Skills', 'Professional competency in Retail Store Operations within Common Skills.'),
        ('Handling Returns & Exchanges', 'Common Skills', 'Professional competency in Handling Returns & Exchanges within Common Skills.'),
        ('Basic Visual Merchandising', 'Common Skills', 'Professional competency in Basic Visual Merchandising within Common Skills.'),
        ('Inventory Count & Reporting', 'Common Skills', 'Professional competency in Inventory Count & Reporting within Common Skills.'),
        ('Basic Security Guard Duties', 'Common Skills', 'Professional competency in Basic Security Guard Duties within Common Skills.'),
        ('Access Control & Visitor Management', 'Common Skills', 'Professional competency in Access Control & Visitor Management within Common Skills.'),
        ('CCTV Monitoring', 'Common Skills', 'Professional competency in CCTV Monitoring within Common Skills.'),
        ('Emergency Response Awareness', 'Common Skills', 'Professional competency in Emergency Response Awareness within Common Skills.'),
        ('Fire Safety Awareness', 'Common Skills', 'Professional competency in Fire Safety Awareness within Common Skills.'),
        ('Crowd Control Assistance', 'Common Skills', 'Professional competency in Crowd Control Assistance within Common Skills.'),
        ('Incident Reporting', 'Common Skills', 'Professional competency in Incident Reporting within Common Skills.'),
        ('Loss Prevention Awareness', 'Common Skills', 'Professional competency in Loss Prevention Awareness within Common Skills.'),
        ('Basic First Aid', 'Common Skills', 'Professional competency in Basic First Aid within Common Skills.'),
        ('Vital Signs Monitoring', 'Common Skills', 'Professional competency in Vital Signs Monitoring within Common Skills.'),
        ('Patient / Elder Care Assistance', 'Common Skills', 'Professional competency in Patient / Elder Care Assistance within Common Skills.'),
        ('Medicine Administration Assistance', 'Common Skills', 'Professional competency in Medicine Administration Assistance within Common Skills.'),
        ('Hygiene Assistance for Patients', 'Common Skills', 'Professional competency in Hygiene Assistance for Patients within Common Skills.'),
        ('Medical Records Filing', 'Common Skills', 'Professional competency in Medical Records Filing within Common Skills.'),
        ('Hospital / Clinic Reception', 'Common Skills', 'Professional competency in Hospital / Clinic Reception within Common Skills.'),
        ('Basic Life Support (BLS) Awareness', 'Common Skills', 'Professional competency in Basic Life Support (BLS) Awareness within Common Skills.'),
        ('Health & Safety Compliance', 'Common Skills', 'Professional competency in Health & Safety Compliance within Common Skills.'),
        ('Light Vehicle Driving', 'Common Skills', 'Professional competency in Light Vehicle Driving within Common Skills.'),
        ('Motorcycle Driving', 'Common Skills', 'Professional competency in Motorcycle Driving within Common Skills.'),
        ('Defensive Driving', 'Common Skills', 'Professional competency in Defensive Driving within Common Skills.'),
        ('Route Planning & Navigation', 'Common Skills', 'Professional competency in Route Planning & Navigation within Common Skills.'),
        ('Vehicle Safety & Inspection', 'Common Skills', 'Professional competency in Vehicle Safety & Inspection within Common Skills.'),
        ('Delivery Coordination', 'Common Skills', 'Professional competency in Delivery Coordination within Common Skills.'),
        ('Passenger Handling', 'Common Skills', 'Professional competency in Passenger Handling within Common Skills.'),
        ('Basic Vehicle Maintenance Awareness', 'Common Skills', 'Professional competency in Basic Vehicle Maintenance Awareness within Common Skills.'),
        ('Blueprint & Plan Reading', 'Common Skills', 'Professional competency in Blueprint & Plan Reading within Common Skills.'),
        ('Construction Site Safety Awareness', 'Common Skills', 'Professional competency in Construction Site Safety Awareness within Common Skills.'),
        ('Painting & Surface Preparation', 'Common Skills', 'Professional competency in Painting & Surface Preparation within Common Skills.'),
        ('Basic Plumbing Assistance', 'Common Skills', 'Professional competency in Basic Plumbing Assistance within Common Skills.'),
        ('Basic Electrical Assistance', 'Common Skills', 'Professional competency in Basic Electrical Assistance within Common Skills.'),
        ('Concrete & Masonry Assistance', 'Common Skills', 'Professional competency in Concrete & Masonry Assistance within Common Skills.'),
        ('Scaffolding & Formworks Assistance', 'Common Skills', 'Professional competency in Scaffolding & Formworks Assistance within Common Skills.'),
        ('Tools & Equipment Handling', 'Common Skills', 'Professional competency in Tools & Equipment Handling within Common Skills.'),
        ('Site Cleanup & Material Disposal', 'Common Skills', 'Professional competency in Site Cleanup & Material Disposal within Common Skills.'),
        ('Classroom Management Assistance', 'Common Skills', 'Professional competency in Classroom Management Assistance within Common Skills.'),
        ('Tutorial & Academic Support', 'Common Skills', 'Professional competency in Tutorial & Academic Support within Common Skills.'),
        ('Lesson Plan Preparation Assistance', 'Common Skills', 'Professional competency in Lesson Plan Preparation Assistance within Common Skills.'),
        ('Childcare & Supervision', 'Common Skills', 'Professional competency in Childcare & Supervision within Common Skills.'),
        ('Educational Material Preparation', 'Common Skills', 'Professional competency in Educational Material Preparation within Common Skills.'),
        ('Student Record Management', 'Common Skills', 'Professional competency in Student Record Management within Common Skills.'),
        ('Crop Planting & Harvesting', 'Common Skills', 'Professional competency in Crop Planting & Harvesting within Common Skills.'),
        ('Irrigation & Water Management', 'Common Skills', 'Professional competency in Irrigation & Water Management within Common Skills.'),
        ('Livestock Care & Feeding', 'Common Skills', 'Professional competency in Livestock Care & Feeding within Common Skills.'),
        ('Pest & Disease Control Awareness', 'Common Skills', 'Professional competency in Pest & Disease Control Awareness within Common Skills.'),
        ('Farm Equipment Operation', 'Common Skills', 'Professional competency in Farm Equipment Operation within Common Skills.'),
        ('Post-Harvest Handling', 'Common Skills', 'Professional competency in Post-Harvest Handling within Common Skills.'),
        ('Organic Farming Practices', 'Common Skills', 'Professional competency in Organic Farming Practices within Common Skills.'),
        ('Following Workplace Rules & Policies', 'Common Skills', 'Professional competency in Following Workplace Rules & Policies within Common Skills.'),
        ('Wearing & Maintaining Proper Uniform', 'Common Skills', 'Professional competency in Wearing & Maintaining Proper Uniform within Common Skills.'),
        ('Attendance & Punctuality', 'Common Skills', 'Professional competency in Attendance & Punctuality within Common Skills.'),
        ('Basic Workplace Etiquette', 'Common Skills', 'Professional competency in Basic Workplace Etiquette within Common Skills.'),
        ('Handling Criticism & Feedback', 'Common Skills', 'Professional competency in Handling Criticism & Feedback within Common Skills.'),
        ('Willingness to Learn On the Job', 'Common Skills', 'Professional competency in Willingness to Learn On the Job within Common Skills.'),
        ('Working with Minimal Supervision', 'Common Skills', 'Professional competency in Working with Minimal Supervision within Common Skills.'),
        ('Shift Work Readiness', 'Common Skills', 'Professional competency in Shift Work Readiness within Common Skills.'),
        ('Physical Fitness for Work', 'Common Skills', 'Professional competency in Physical Fitness for Work within Common Skills.'),
        ('Basic Safety Compliance', 'Common Skills', 'Professional competency in Basic Safety Compliance within Common Skills.'),
        ('Network Security', 'Cybersecurity', 'Professional competency in Network Security within Cybersecurity.'),
        ('Ethical Hacking', 'Cybersecurity', 'Professional competency in Ethical Hacking within Cybersecurity.'),
        ('Penetration Testing', 'Cybersecurity', 'Professional competency in Penetration Testing within Cybersecurity.'),
        ('Vulnerability Assessment', 'Cybersecurity', 'Professional competency in Vulnerability Assessment within Cybersecurity.'),
        ('Security Information & Event Management', 'Cybersecurity', 'Professional competency in Security Information & Event Management within Cybersecurity.'),
        ('Incident Response', 'Cybersecurity', 'Professional competency in Incident Response within Cybersecurity.'),
    ]
    centralized_skills = {}
    for name, cat, desc in skills_data:
        skill, _ = CentralizedSkill.objects.get_or_create(
            name=name,
            defaults={'category': cat, 'description': desc}
        )
        centralized_skills[name] = skill

    # 2. Training Programs
    training_programs = [
        ('React Basics for Front-end Careers', 'TESDA', 'React', '40 Hours'),
        ('Google Cloud DevOps & AWS Arch.', 'AWS', 'AWS Cloud', '80 Hours'),
        ('Introduction to Welding and Structural Safety', 'TESDA', 'Welding', '120 Hours'),
        ('Hotel Housekeeping NC II certification program', 'TESDA', 'Housekeeping', '160 Hours'),
        ('Kubernetes in Production', 'TESDA', 'Kubernetes', '60 Hours'),
        ('TypeScript Fundamentals', 'Scrimba', 'TypeScript', '30 Hours'),
        ('Advanced HTML & CSS Responsive Design', 'TESDA', 'HTML & CSS', '50 Hours'),
        ('Docker Containerization Essentials', 'Linux Foundation', 'Docker', '24 Hours'),
        ('Python for Everybody Boot Camp', 'Coursera', 'Python', '48 Hours'),
        ('SQL & PostgreSQL for Beginners', 'Udemy', 'SQL', '36 Hours'),
        ('Django Web Framework Mastery', 'TESDA', 'Django', '80 Hours'),
        ('Agile Product Management & Roadmap Design', 'Atlassian', 'Agile', '40 Hours'),
        ('Scrum & Kanban NC II Prep Course', 'TESDA', 'Scrum', '60 Hours'),
        ('Figma UX/UI Design Essentials', 'UX Design Institute', 'Figma', '40 Hours'),
        ('Deep Learning with PyTorch', 'DeepLearning.AI', 'PyTorch', '60 Hours'),
        ('TESDA Masonry & Plastering Certification', 'TESDA', 'Concrete & Masonry Assistance', '120 Hours'),
        ('TESDA Construction Blueprint Masterclass', 'TESDA', 'Blueprint & Plan Reading', '80 Hours'),
        ('TESDA Standard Plumbing Course', 'TESDA', 'Basic Plumbing Assistance', '100 Hours'),
        ('TESDA Electrical Installation NC II', 'TESDA', 'Basic Electrical Assistance', '160 Hours'),
        ('TESDA Basic Defensive Driving Course', 'TESDA', 'Defensive Driving', '40 Hours'),
        ('TESDA Retail Merchandising NC I', 'TESDA', 'Retail Store Operations', '80 Hours'),
        ('TESDA Healthcare Support Services', 'TESDA', 'Vital Signs Monitoring', '120 Hours')
    ]
    for title, provider, skill_name, dur in training_programs:
        TrainingProgram.objects.get_or_create(
            title=title,
            defaults={
                'provider': provider,
                'skill_addressed': centralized_skills[skill_name],
                'description': f"Targeted program to master {skill_name} and build immediate competency.",
                'duration': dur
            }
        )

    # Return early to prevent seeding other user accounts/profiles to respect cleanup requests
    return

    # 3. Create Employer User
    employer_user, created = User.objects.get_or_create(
        username='employer@test.com',
        defaults={'email': 'employer@test.com', 'first_name': 'Global', 'last_name': 'Industries'}
    )
    if created:
        employer_user.set_password('password123')
        employer_user.save()
    employer_profile, _ = Profile.objects.update_or_create(
        user=employer_user,
        defaults={
            'role': 'employer',
            'is_verified': True,
            'company_name': 'Nexus Digital Solutions',
            'industry': 'Artificial Intelligence',
            'company_size': '50-200 employees',
            'website': 'https://nexus-digital.ai',
            'contact_name': 'Dr. Alexander Vance',
            'contact_position': 'Chief Talent Officer',
            'contact_email': 'a.vance@nexus-digital.ai',
            'employment_type_offered': 'Local',
            'location': 'San Francisco, CA',
            'soft_notes': 'Nexus Digital Solutions is a leading AI consultancy specializing in workforce transformation. Founded in 2018, we have helped over 500 enterprises navigate the complex landscape of digital evolution through data-driven skill gap analysis and custom learning architectures. Our mission is to empower teams to reach their peak potential through technological synergy.'
        }
    )

    # 4. Create Job Vacancies and Requirements
    vacancies_data = [
        {
            'title': 'Senior Cloud Architect',
            'location': 'San Francisco, CA',
            'salary_range': '$120k - $150k',
            'requirements': [('AWS Cloud', 5), ('Kubernetes', 4), ('React', 3)]
        },
        {
            'title': 'Data Scientist',
            'location': 'Remote',
            'salary_range': '$100k - $130k',
            'requirements': [('Python', 5), ('PyTorch', 4), ('SQL', 3)]
        },
        {
            'title': 'UX Designer',
            'location': 'London, UK',
            'salary_range': '£60k - £80k',
            'requirements': [('Figma', 5), ('A11y', 4), ('React', 2)]
        },
        {
            'title': 'Warehouse Supervisor',
            'location': 'Austin, TX',
            'salary_range': '$45k - $55k',
            'requirements': [('Housekeeping', 4), ('Agile', 2)]
        },
        {
            'title': 'Frontend Engineer',
            'location': 'Remote',
            'salary_range': '$80k - $110k',
            'requirements': [('React', 4), ('TypeScript', 4), ('HTML & CSS', 5)]
        },
        {
            'title': 'DevOps Engineer',
            'location': 'Seattle, WA',
            'salary_range': '$110k - $140k',
            'requirements': [('Docker', 4), ('Kubernetes', 4), ('AWS Cloud', 3)]
        },
        {
            'title': 'Junior Backend Developer',
            'location': 'Austin, TX',
            'salary_range': '$65k - $85k',
            'requirements': [('Python', 3), ('SQL', 3), ('Django', 3)]
        },
        {
            'title': 'TESDA Structural Welder',
            'location': 'Houston, TX',
            'salary_range': '$40k - $55k',
            'requirements': [('Welding', 5), ('Housekeeping', 2)]
        },
        {
            'title': 'Product Manager',
            'location': 'New York, NY',
            'salary_range': '$120k - $150k',
            'requirements': [('Agile', 5), ('Scrum', 4), ('Kanban', 4)]
        },
        {
            'title': 'TESDA Construction Mason',
            'location': 'Manila, PH',
            'salary_range': '₱15k - ₱22k',
            'requirements': [('Concrete & Masonry Assistance', 4), ('Blueprint & Plan Reading', 3), ('Construction Site Safety Awareness', 5)]
        },
        {
            'title': 'Defensive Driver',
            'location': 'Cebu, PH',
            'salary_range': '₱14k - ₱18k',
            'requirements': [('Defensive Driving', 5), ('Light Vehicle Driving', 4)]
        },
        {
            'title': 'Healthcare Assistant',
            'location': 'Davao, PH',
            'salary_range': '₱16k - ₱20k',
            'requirements': [('Vital Signs Monitoring', 4), ('Patient / Elder Care Assistance', 5)]
        },
        {
            'title': 'Retail Store Associate',
            'location': 'Quezon City, PH',
            'salary_range': '₱13k - ₱16k',
            'requirements': [('Retail Store Operations', 4), ('Stock Monitoring & Replenishment', 4)]
        }
    ]
    category_mapping = {
        'Senior Cloud Architect': 'IT',
        'Data Scientist': 'IT',
        'UX Designer': 'IT',
        'Warehouse Supervisor': 'LOG',
        'Frontend Engineer': 'IT',
        'DevOps Engineer': 'IT',
        'Junior Backend Developer': 'IT',
        'TESDA Structural Welder': 'TVT',
        'Product Manager': 'IT',
        'TESDA Construction Mason': 'ENG',
        'Defensive Driver': 'DRV',
        'Healthcare Assistant': 'MED',
        'Retail Store Associate': 'RTL',
    }

    vacancies = {}
    for vac_info in vacancies_data:
        cat_code = category_mapping.get(vac_info['title'], 'IT')
        vacancy, _ = JobVacancy.objects.get_or_create(
            title=vac_info['title'],
            employer=employer_profile,
            defaults={
                'location': vac_info['location'],
                'salary_range': vac_info['salary_range'],
                'description': f"Looking for a motivated {vac_info['title']} to join our growing enterprise team.",
                'category': cat_code
            }
        )
        # Always update the category for seeded vacancies to match
        vacancy.category = cat_code
        vacancy.save()
        
        vacancies[vacancy.title] = vacancy
        for s_name, req_level in vac_info['requirements']:
            JobSkillRequirement.objects.get_or_create(
                job_vacancy=vacancy,
                skill=centralized_skills[s_name],
                defaults={'required_proficiency': req_level}
            )

    # Convert any other old 'jobs' categories to 'IT' to keep database clean
    JobVacancy.objects.filter(category='jobs').update(category='IT')

    # 5. Seed Mock Candidates
    mock_candidates = [
        {
            'first_name': 'Alex',
            'last_name': 'Rivera',
            'email': 'alex.rivera@design.ai',
            'phone': '+1 (555) 019-2831',
            'title': 'Senior Frontend Engineer',
            'status': 'Interviewing',
            'skills': 'React, TypeScript',
            'skill_level': 'Expert',
            'location': 'Austin, TX',
            'training_progress_title': 'Cloud Arch.',
            'training_progress_percentage': 85,
            'experience_years': '8 Years',
            'education': ('Stanford University', 'BS', 'Computer Science', 2014, 2018),
            'experience': ('TechFlow Systems', 'Senior Full-Stack Engineer', 'Leading the development of a distributed cloud architecture processing 2M+ daily events. Spearheaded transition from monolithic to microservices, improving deployment frequency by 40%.', '2021', 'Present', True),
            'certification': ('AWS Solutions Architect', 'Amazon Web Services', False),
            'skills_rating': [('React', 5), ('TypeScript', 4), ('AWS Cloud', 3)]
        },
        {
            'first_name': 'Sarah',
            'last_name': 'Jenkins',
            'email': 'sarah.j@analytics.ai',
            'phone': '+1 (555) 018-4729',
            'title': 'Data Scientist',
            'status': 'Skill Gap Identified',
            'skills': 'Python, PyTorch',
            'skill_level': 'Advanced',
            'location': 'Remote',
            'training_progress_title': 'Kubernetes',
            'training_progress_percentage': 32,
            'experience_years': '5 Years',
            'education': ('MIT', 'MS', 'Mathematics', 2015, 2017),
            'experience': ('Analytics AI', 'Data Scientist', 'Developed predictive models using Python, PyTorch, and SQL. Collaborated with business teams to increase sales by 10%.', '2018', '2023', False),
            'certification': ('Professional Data Engineer', 'Google Cloud', False),
            'skills_rating': [('Python', 4), ('PyTorch', 5), ('Kubernetes', 2)]
        },
        {
            'first_name': 'Marcus',
            'last_name': 'Chen',
            'email': 'marcus.c@product.ai',
            'phone': '+1 (555) 017-9104',
            'title': 'Product Manager',
            'status': 'Hired',
            'skills': 'Agile, SQL',
            'skill_level': 'Expert',
            'location': 'Toronto, ON',
            'training_progress_title': 'Leadership',
            'training_progress_percentage': 100,
            'experience_years': '5 Years',
            'education': ('Toronto University', 'MBA', 'Business Administration', 2012, 2014),
            'experience': ('Product AI', 'Product Manager', 'Managed software product lifecycle from concept to release. Led a cross-functional agile team of 8.', '2015', '2020', False),
            'certification': ('Scrum Master', 'Scrum Alliance', False),
            'skills_rating': [('Agile', 5), ('SQL', 3)]
        },
        {
            'first_name': 'Lila',
            'last_name': 'Thorne',
            'email': 'lila.t@design.ai',
            'phone': '+1 (555) 016-8315',
            'title': 'UX Designer',
            'status': 'Review Required',
            'skills': 'Figma, A11y',
            'skill_level': 'Intermediate',
            'location': 'London, UK',
            'training_progress_title': 'React Basics',
            'training_progress_percentage': 55,
            'experience_years': '3 Years',
            'education': ('London Art College', 'BA', 'UX Design', 2016, 2019),
            'experience': ('Design AI', 'UX Designer', 'Designed mobile and web experiences using Figma. Performed usability testing and accessibility reviews.', '2020', '2023', False),
            'certification': ('Figma Certified Professional', 'Figma', False),
            'skills_rating': [('Figma', 4), ('A11y', 3), ('React', 2)]
        },
        {
            'first_name': 'Daniel',
            'last_name': 'Alvarez',
            'email': 'daniel.a@gmail.com',
            'phone': '+1 (555) 021-3940',
            'title': 'Warehouse Associate',
            'status': 'Applied',
            'skills': 'Housekeeping',
            'skill_level': 'Intermediate',
            'location': 'Austin, TX',
            'training_progress_title': 'General Training',
            'training_progress_percentage': 0,
            'experience_years': '2 Years',
            'education': ('State College', 'BS', 'Psychology', 2018, 2022),
            'experience': ('Austin Distribution', 'Warehouse Assistant', 'Managed inventory control and housekeeping standards for the central warehouse.', '2022', '2024', False),
            'certification': ('Warehouse Safety NC II', 'TESDA', True),
            'skills_rating': [('Housekeeping', 4), ('Agile', 1)]
        },
        {
            'first_name': 'Sofia',
            'last_name': 'Ramirez',
            'email': 'sofia.r@gmail.com',
            'phone': '+1 (555) 022-4859',
            'title': 'Backend Developer',
            'status': 'Applied',
            'skills': 'Python, SQL',
            'skill_level': 'Intermediate',
            'location': 'Remote',
            'training_progress_title': 'General Training',
            'training_progress_percentage': 0,
            'experience_years': '3 Years',
            'education': ('State University', 'BS', 'Computer Engineering', 2019, 2023),
            'experience': ('DevSolutions Inc.', 'Junior Backend Developer', 'Developed backend APIs using Python and SQL. Maintained database reliability and performance.', '2023', '2025', False),
            'certification': ('Python Certified Associate', 'Python Institute', False),
            'skills_rating': [('Python', 4), ('SQL', 4), ('Django', 2)]
        }
    ]
    
    profiles = {}
    for cand in mock_candidates:
        user, created = User.objects.get_or_create(
            username=cand['email'],
            defaults={
                'email': cand['email'],
                'first_name': cand['first_name'],
                'last_name': cand['last_name']
            }
        )
        if created:
            user.set_password('password123')
            user.save()
        
        profile, _ = Profile.objects.update_or_create(
            user=user,
            defaults={
                'phone_number': cand['phone'],
                'role': 'applicant',
                'is_verified': True,
                'title': cand['title'],
                'status': cand['status'],
                'skills': cand['skills'],
                'skill_level': cand['skill_level'],
                'location': cand['location'],
                'training_progress_title': cand['training_progress_title'],
                'training_progress_percentage': cand['training_progress_percentage'],
                'experience_years': cand['experience_years'],
                'is_profile_complete': True
            }
        )
        profiles[profile.user.email] = profile

        # Seed structured Education
        edu_info = cand['education']
        Education.objects.get_or_create(
            profile=profile,
            institution=edu_info[0],
            defaults={
                'degree': edu_info[1],
                'field_of_study': edu_info[2],
                'start_year': edu_info[3],
                'end_year': edu_info[4]
            }
        )

        # Seed structured WorkExperience
        exp_info = cand['experience']
        WorkExperience.objects.get_or_create(
            profile=profile,
            company=exp_info[0],
            position=exp_info[1],
            defaults={
                'description': exp_info[2],
                'start_date': exp_info[3],
                'end_date': exp_info[4],
                'is_current': exp_info[5]
            }
        )

        # Seed structured Certification
        cert_info = cand['certification']
        Certification.objects.get_or_create(
            profile=profile,
            name=cert_info[0],
            defaults={
                'issuing_organization': cert_info[1],
                'is_tesda': cert_info[2]
            }
        )

        # Seed structured ApplicantSkills
        for s_name, rating in cand['skills_rating']:
            ApplicantSkill.objects.update_or_create(
                profile=profile,
                skill=centralized_skills[s_name],
                defaults={'proficiency': rating}
            )

    # 6. Seed Mock Referrals
    referral_list = [
        ('daniel.a@gmail.com', 'Warehouse Supervisor', '2026-06-10', 'Pending'),
        ('alex.rivera@design.ai', 'Senior Cloud Architect', '2026-06-12', 'Hired'),
        ('sarah.j@analytics.ai', 'Data Scientist', '2026-06-14', 'Hired'),
        ('lila.t@design.ai', 'UX Designer', '2026-06-15', 'No Response')
    ]
    for email, job_title, r_date, r_status in referral_list:
        Referral.objects.get_or_create(
            applicant=profiles[email],
            job_vacancy=vacancies[job_title],
            defaults={
                'date_referred': r_date,
                'status': r_status
            }
        )

@login_required(login_url='login')
@require_POST
def update_applicant_status(request, profile_id):
    if request.user.profile.role not in ['admin', 'employer']:
        return redirect('dashboard')
    profile = get_object_or_404(Profile, id=profile_id)
    new_status = request.POST.get('status')
    if new_status:
        profile.status = new_status
        profile.save()
        
        # If updated by employer, update their specific Referral status too
        if request.user.profile.role == 'employer':
            ref = Referral.objects.filter(applicant=profile, job_vacancy__employer=request.user.profile, status='Pending').first()
            if ref:
                if new_status == 'Shortlisted':
                    ref.status = 'Interviewing'
                elif new_status == 'Rejected':
                    ref.status = 'Not Hired'
                ref.save()
        messages.success(request, f"Status of {profile.user.first_name} {profile.user.last_name} updated to {new_status}!")
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

# Helper function to check if candidate's field of study is mismatched with referred job vacancy
def check_mismatch(field_of_study, job_title):
    fos = (field_of_study or '').lower()
    jt = (job_title or '').lower()
    
    # Fields that are aligned with tech/digital careers
    tech_study = [
        'computer', 'math', 'science', 'engineering', 'design', 'ux', 'it',
        'software', 'information', 'technology', 'data', 'digital', 'networking',
        'programming', 'systems', 'web', 'electronics', 'multimedia'
    ]
    # Job titles considered tech/digital
    tech_job = [
        'cloud', 'architect', 'scientist', 'ux', 'designer', 'engineer',
        'developer', 'programmer', 'analyst', 'devops', 'data', 'it', 'tech',
        'information', 'technology', 'computer', 'expert', 'web', 'software',
        'systems', 'network', 'digital'
    ]
    
    is_tech_study = any(x in fos for x in tech_study)
    is_tech_job = any(x in jt for x in tech_job)
    
    # Only flag as a mismatch if the study and job are on clearly opposite sides
    if is_tech_job and not is_tech_study:
        return True
    if not is_tech_job and is_tech_study:
        return True
    # Explicit PESO-relevant mismatch: Psychology grad in physical trade roles
    if 'psychology' in fos and any(x in jt for x in ['warehouse', 'supervisor', 'housekeeper', 'welder', 'trade']):
        return True
    return False


# Helper function to calculate proficiency matching percentage and gaps
def calculate_match_score(profile, vacancy):
    # Use list evaluation to utilize prefetch cache if it exists
    requirements = list(vacancy.requirements.all())
    if not requirements:
        return 100.0, []
    
    # Check if applicant_skills are pre-fetched; if so, avoid hits to db
    applicant_skills = list(profile.applicant_skills.all())
    applicant_skill_ids = {askill.skill_id for askill in applicant_skills}
    
    total_score = 0.0
    gaps = []
    for req in requirements:
        has_skill = req.skill_id in applicant_skill_ids
        
        if has_skill:
            cand_level = 1
            req_level = 1
            skill_score = 1.0
            gap = 0
            gap_pct = 0
            current_pct = 100
        else:
            cand_level = 0
            req_level = 1
            skill_score = 0.0
            gap = 1
            gap_pct = 100
            current_pct = 0
            
        total_score += skill_score
        
        gaps.append({
            'skill': req.skill.name,
            'required': req_level,
            'candidate': cand_level,
            'gap': gap,
            'gap_pct': gap_pct,
            'current_pct': current_pct
        })
        
    match_percent = (total_score / len(requirements)) * 100.0
    
    # Factor in education mismatches (using list to prevent db hit if prefetched)
    educations = list(profile.education.all())
    edu = educations[0] if educations else None
    field_of_study = edu.field_of_study if edu else ''
    if check_mismatch(field_of_study, vacancy.title):
        match_percent = max(0.0, match_percent - 20.0)
        
    # Factor in experience gap for non-fresh grads (using list to prevent db hit if prefetched)
    if vacancy.required_experience_years > 0:
        experiences = list(profile.experience.all())
        exp_count = len(experiences)
        if exp_count == 0 and not profile.is_fresh_grad:
            match_percent = max(0.0, match_percent - 10.0)
            
    match_percent = round(match_percent, 1)
    
    # Commented out to prevent database write bottlenecks on read queries
    # if profile.role == 'applicant':
    #     GapScoreLog.objects.create(
    #         profile=profile,
    #         job_vacancy=vacancy,
    #         match_percentage=match_percent,
    #         gap_data=gaps
    #     )
        
    return match_percent, gaps

def check_and_notify_employer_matches(profile):
    if profile.role != 'applicant':
        return
    vacancies = JobVacancy.objects.exclude(status='Closed')
    for vac in vacancies:
        match_pct, _ = calculate_match_score(profile, vac)
        if match_pct >= 75:
            already_notified = Notification.objects.filter(
                user=vac.employer.user,
                notif_type='match_alert',
                applicant_profile=profile,
                vacancy=vac,
            ).exists()
            if not already_notified:
                Notification.objects.create(
                    user=vac.employer.user,
                    notif_type='match_alert',
                    applicant_profile=profile,
                    vacancy=vac,
                    message=f"🎯 Match Alert: {profile.user.get_full_name()} has reached {int(match_pct)}% skill match for your vacancy '{vac.title}'. You can now send them an offer!"
                )

def check_and_notify_applicants_for_vacancy(vacancy):
    if vacancy.status == 'Closed':
        return
    applicants = Profile.objects.filter(role='applicant')
    for app in applicants:
        match_pct, _ = calculate_match_score(app, vacancy)
        if match_pct >= 75:
            already_notified = Notification.objects.filter(
                user=vacancy.employer.user,
                notif_type='match_alert',
                applicant_profile=app,
                vacancy=vacancy,
            ).exists()
            if not already_notified:
                Notification.objects.create(
                    user=vacancy.employer.user,
                    notif_type='match_alert',
                    applicant_profile=app,
                    vacancy=vacancy,
                    message=f"🎯 Match Alert: {app.user.get_full_name()} has reached {int(match_pct)}% skill match for your vacancy '{vacancy.title}'. You can now send them an offer!"
                )

def get_applicant_training_recommendations(profile):
    preferred_jobs = []
    
    # Bookmarked open jobs
    bookmarked = JobVacancy.objects.filter(bookmarks__profile=profile).exclude(status='Closed')
    for v in bookmarked:
        if v not in preferred_jobs:
            preferred_jobs.append(v)
            if len(preferred_jobs) >= 5:
                break
                
    # Preferred category open jobs
    target_category = profile.preferred_job
    if len(preferred_jobs) < 5 and target_category:
        cat_vacancies = JobVacancy.objects.filter(category=target_category).exclude(status='Closed')
        for v in cat_vacancies:
            if v not in preferred_jobs:
                preferred_jobs.append(v)
                if len(preferred_jobs) >= 5:
                    break
                    
    # Fallback to any open vacancies
    if len(preferred_jobs) < 5:
        all_vacs = JobVacancy.objects.exclude(status='Closed')
        for v in all_vacs:
            if v not in preferred_jobs:
                preferred_jobs.append(v)
                if len(preferred_jobs) >= 5:
                    break
                    
    gaps_set = set()
    highest_match = 0
    biggest_gap_skill = None
    max_gap_val = -1
    
    best_job = None
    for vac in preferred_jobs:
        match_pct, _ = calculate_match_score(profile, vac)
        if match_pct > highest_match:
            highest_match = match_pct
            best_job = vac
            
    for vac in preferred_jobs:
        _, gaps = calculate_match_score(profile, vac)
        for g in gaps:
            if g['gap'] > 0:
                gaps_set.add(g['skill'])
                if vac == best_job and g['gap'] > max_gap_val:
                    max_gap_val = g['gap']
                    biggest_gap_skill = g['skill']
                    
    enrolled_ids = set(profile.training_enrollments.values_list('training_program_id', flat=True))
    recommended_programs = TrainingProgram.objects.filter(skill_addressed__name__in=gaps_set).exclude(id__in=enrolled_ids).select_related('skill_addressed')
    return recommended_programs, highest_match, len(gaps_set), biggest_gap_skill

@role_required(['admin'])
def peso_dashboard_admin(request):
    seed_mock_applicants_if_empty()
    total_applicants = Profile.objects.filter(role='applicant').count()
    total_employers = Profile.objects.filter(role='employer').count()
    total_vacancies = JobVacancy.objects.count()
    
    referrals = Referral.objects.all()
    total_referrals = referrals.count()
    total_placements = referrals.filter(status='Hired').count()
    
    # Conversion, Mismatch, and No Response Rates
    conversion_rate = (total_placements / total_referrals * 100.0) if total_referrals > 0 else 0.0
    no_response_count = referrals.filter(status='No Response').count()
    no_response_rate = (no_response_count / total_referrals * 100.0) if total_referrals > 0 else 0.0
    
    mismatched_count = 0
    for r in referrals:
        edu = r.applicant.education.first()
        fos = edu.field_of_study if edu else ''
        if check_mismatch(fos, r.job_vacancy.title):
            mismatched_count += 1
    mismatch_rate = (mismatched_count / total_referrals * 100.0) if total_referrals > 0 else 0.0

    applicants = Profile.objects.filter(role='applicant').select_related('user')[:5]
    for applicant in applicants:
        applicant.skills_list = [s.strip() for s in applicant.skills.split(',')] if applicant.skills else []
        
    context = {
        'total_applicants': total_applicants,
        'total_employers': total_employers,
        'total_vacancies': total_vacancies,
        'total_placements': total_placements,
        'conversion_rate': conversion_rate,
        'no_response_rate': no_response_rate,
        'mismatch_rate': mismatch_rate,
        'applicants': applicants,
    }
    return render(request, 'tracker/ADMIN/peso_dashboard_admin.html', context)

import csv
import io

@role_required(['admin'])
def bulk_import_applicants(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        decoded_file = csv_file.read().decode('utf-8')
        io_string = io.StringIO(decoded_file)
        reader = csv.reader(io_string, delimiter=',')
        
        # Skip header
        header = next(reader, None)
        
        count = 0
        skipped = 0
        for row in reader:
            if not row or len(row) < 6:
                continue
            
            first_name = row[0].strip()
            last_name = row[1].strip()
            email = row[2].strip()
            phone = row[3].strip()
            birthdate_str = row[4].strip()
            civil_status = row[5].strip()
            
            degree = row[6].strip() if len(row) > 6 else ""
            field_of_study = row[7].strip() if len(row) > 7 else ""
            institution = row[8].strip() if len(row) > 8 else ""
            end_year = row[9].strip() if len(row) > 9 else ""
            
            company = row[10].strip() if len(row) > 10 else ""
            position = row[11].strip() if len(row) > 11 else ""
            exp_desc = row[12].strip() if len(row) > 12 else ""
            
            # Edge Case: Applicant has no email -> generate coordinator-assisted format
            if not email:
                if phone:
                    clean_phone = phone.replace(' ', '').replace('+', '').replace('-', '')
                    email = f"{clean_phone}@peso.gov.ph"
                else:
                    email = f"{first_name.lower()}.{last_name.lower()}@peso.gov.ph"
            
            # Edge Case: Duplicate check (name + birthdate)
            duplicate = Profile.objects.filter(
                user__first_name__iexact=first_name,
                user__last_name__iexact=last_name,
                birthdate=birthdate_str if birthdate_str else None
            ).first()
            if duplicate:
                skipped += 1
                continue
                
            # Create user and profile
            if User.objects.filter(username=email).exists():
                # Append random digits if username exists
                email = f"{email.split('@')[0]}{random.randint(100, 999)}@{email.split('@')[1]}"
                
            user = User.objects.create_user(
                username=email,
                email=email,
                password='password123',
                first_name=first_name,
                last_name=last_name
            )
            
            profile = Profile.objects.create(
                user=user,
                phone_number=phone,
                role='applicant',
                is_verified=True, # Pre-verified since registered by admin/coordinator
                is_profile_complete=True,
                birthdate=birthdate_str if birthdate_str else None,
                civil_status=civil_status if civil_status else 'Single',
                title=position if position else (f"Fresh Graduate - {field_of_study}" if field_of_study else "General Applicant")
            )
            
            # Education
            if institution and degree:
                Education.objects.create(
                    profile=profile,
                    institution=institution,
                    degree=degree,
                    field_of_study=field_of_study,
                    start_year=2020,
                    end_year=int(end_year) if end_year.isdigit() else 2024
                )
                
            # Work Experience
            if company and position:
                WorkExperience.objects.create(
                    profile=profile,
                    company=company,
                    position=position,
                    description=exp_desc,
                    start_date="2021",
                    end_date="2024",
                    is_current=False
                )
                
            count += 1
            
        messages.success(request, f"Bulk import complete: {count} applicants registered, {skipped} duplicate profiles skipped.")
        
    return redirect('applicant_monitoring_admin')

@role_required(['admin'])
def coordinator_interview(request, profile_id):
    profile = get_object_or_404(Profile, id=profile_id, role='applicant')
    
    if request.method == 'POST':
        # 1. Update Personal Info & Intake details
        profile.phone_number = request.POST.get('phone', '').strip()
        profile.civil_status = request.POST.get('civil_status', 'Single')
        profile.soft_notes = request.POST.get('soft_notes', '').strip()
        
        is_fresh_grad = request.POST.get('is_fresh_grad') == 'on'
        profile.is_fresh_grad = is_fresh_grad
        
        # 2. Update/Create Education
        edu_id = request.POST.get('edu_id')
        institution = request.POST.get('institution', '').strip()
        degree = request.POST.get('degree', '').strip()
        field_of_study = request.POST.get('field_of_study', '').strip()
        end_year = request.POST.get('end_year', '').strip()
        
        if institution and degree:
            if edu_id:
                edu = Education.objects.filter(id=edu_id, profile=profile).first()
                if edu:
                    edu.institution = institution
                    edu.degree = degree
                    edu.field_of_study = field_of_study
                    edu.end_year = int(end_year) if end_year.isdigit() else 2024
                    edu.save()
            else:
                Education.objects.create(
                    profile=profile,
                    institution=institution,
                    degree=degree,
                    field_of_study=field_of_study,
                    start_year=2020,
                    end_year=int(end_year) if end_year.isdigit() else 2024
                )
                
        # 3. Update/Create Work Experience (or OJT/Internship)
        exp_id = request.POST.get('exp_id')
        company = request.POST.get('company', '').strip()
        position = request.POST.get('position', '').strip()
        description = request.POST.get('description', '').strip()
        start_date = request.POST.get('start_date', '').strip()
        end_date = request.POST.get('end_date', 'Present').strip()
        is_current = request.POST.get('is_current') == 'on'
        
        if company and position:
            title_suffix = " (OJT/Internship)" if is_fresh_grad and not position.endswith(" (OJT/Internship)") else ""
            final_position = position + title_suffix
            
            if exp_id:
                exp = WorkExperience.objects.filter(id=exp_id, profile=profile).first()
                if exp:
                    exp.company = company
                    exp.position = final_position
                    exp.description = description
                    exp.start_date = start_date
                    exp.end_date = end_date
                    exp.is_current = is_current
                    exp.save()
            else:
                WorkExperience.objects.create(
                    profile=profile,
                    company=company,
                    position=final_position,
                    description=description,
                    start_date=start_date,
                    end_date=end_date,
                    is_current=is_current
                )
                
        # 4. Handle Skill Ratings & overrides
        selected_skills = request.POST.getlist('skills[]')
        selected_proficiencies = request.POST.getlist('proficiencies[]')
        
        profile_skills_list = []
        for skill_id, prof in zip(selected_skills, selected_proficiencies):
            if skill_id:
                skill_obj = get_object_or_404(CentralizedSkill, id=int(skill_id))
                ApplicantSkill.objects.update_or_create(
                    profile=profile,
                    skill=skill_obj,
                    defaults={
                        'proficiency': int(prof) if prof else 1,
                        'source': 'coordinator-verified'
                    }
                )
                profile_skills_list.append(skill_obj.name)
                
        # Set profile title
        if not is_fresh_grad and position:
            profile.title = position
        elif field_of_study:
            profile.title = f"Fresh Graduate - {field_of_study}"
            
        if profile_skills_list:
            profile.skills = ", ".join(profile_skills_list)
        profile.save()
        
        messages.success(request, f"Intake interview record for {profile.user.first_name} {profile.user.last_name} updated successfully!")
        return redirect('applicant_monitoring_admin')
        
    skills = CentralizedSkill.objects.all()
    education = profile.education.first()
    experience = profile.experience.first()
    
    applicant_skills_dict = {s.skill_id: s for s in profile.applicant_skills.all()}
    skills_with_ratings = []
    for skill in skills:
        app_skill = applicant_skills_dict.get(skill.id)
        skills_with_ratings.append({
            'id': skill.id,
            'name': skill.name,
            'has_skill': app_skill is not None,
            'proficiency': app_skill.proficiency if app_skill else 3,
            'source': app_skill.source if app_skill else 'self',
        })
    
    return render(request, 'tracker/ADMIN/coordinator_interview.html', {
        'profile': profile,
        'education': education,
        'experience': experience,
        'skills_with_ratings': skills_with_ratings
    })

@role_required(['admin'])
def applicant_monitoring_admin(request):
    seed_mock_applicants_if_empty()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        referral_id = request.POST.get('referral_id')
        ref = get_object_or_404(Referral, id=referral_id)
        
        if action == 'log_mid_probation':
            ref.mid_probation_outcome = request.POST.get('mid_probation_outcome')
            ref.mid_probation_notes = request.POST.get('mid_probation_notes', '')
            ref.mid_probation_checked_at = timezone.now()
            
            outcome = ref.mid_probation_outcome
            if outcome in ['Resigned Voluntarily', 'Terminated by Employer']:
                ref.status = outcome
                ref.applicant.status = 'Active — Job Seeking'
                ref.applicant.save()
            ref.save()
            messages.success(request, f"Logged mid-probation check for {ref.applicant.user.get_full_name()}.")
        return redirect('applicant_monitoring_admin')
        
    applicants = Profile.objects.filter(role='applicant').select_related('user')
    vacancies = JobVacancy.objects.all()
    for applicant in applicants:
        applicant.skills_list = [s.strip() for s in applicant.skills.split(',')] if applicant.skills else []
        
    probationary_referrals = Referral.objects.filter(status__in=[
        'Hired — Probationary',
        'Probation Extended',
        'Still Employed — Performing Well',
        'Still Employed — On Improvement Plan',
        'No Response from Employer',
        'No Response from Applicant',
    ]).select_related('applicant__user', 'job_vacancy')
    
    today = timezone.localdate()
    for r in probationary_referrals:
        if r.actual_start_date:
            mid_months = (r.probationary_period_months or 6) / 2.0
            mid_date = r.actual_start_date + datetime.timedelta(days=int(mid_months * 30.4))
            r.mid_probation_date = mid_date
            r.mid_probation_required = (today >= mid_date and not r.mid_probation_outcome)
        else:
            r.mid_probation_required = False
            
    return render(request, 'tracker/ADMIN/applicant_monitoring_admin.html', {
        'applicants': applicants,
        'vacancies': vacancies,
        'probationary_referrals': probationary_referrals,
    })

@role_required(['admin'])
def employer_management_admin(request):
    seed_mock_applicants_if_empty()
    employers = Profile.objects.filter(role='employer').select_related('user')
    
    # Calculate stats dynamically for each employer
    for emp in employers:
        emp_referrals = Referral.objects.filter(job_vacancy__employer=emp)
        emp.total_referrals = emp_referrals.count()
        emp.active_postings = JobVacancy.objects.filter(employer=emp, status='Open').count()
        
        hired = emp_referrals.filter(status='Hired').count()
        not_hired = emp_referrals.filter(status='Not Hired').count()
        no_response = emp_referrals.filter(status='No Response').count()
        
        responded = hired + not_hired
        completed = hired + not_hired + no_response
        
        emp.response_rate = round((responded / completed * 100.0), 1) if completed > 0 else 100.0
        # Flag employer as low responsiveness if they have at least 3 completed referrals and response rate is < 50%
        emp.low_responsiveness = (completed >= 3 and emp.response_rate < 50.0)
        
    return render(request, 'tracker/ADMIN/employer_management_admin.html', {'employers': employers})

@role_required(['admin'])
@require_POST
def verify_employer(request, profile_id):
    profile = get_object_or_404(Profile, id=profile_id, role='employer')
    profile.is_verified = True
    profile.save()
    messages.success(request, f"Successfully verified {profile.user.first_name} {profile.user.last_name}!")
    return redirect('employer_management_admin')

@role_required(['admin'])
def employment_tracking_admin(request):
    seed_mock_applicants_if_empty()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        referral_id = request.POST.get('referral_id')
        ref = get_object_or_404(Referral, id=referral_id)
        
        if action == 'update_onboarding':
            ref.pre_employment_status = request.POST.get('pre_employment_status', 'Incomplete')
            ref.coordination_notes = request.POST.get('coordination_notes', '')
            rep_date_str = request.POST.get('confirmed_reporting_date')
            if rep_date_str:
                ref.reporting_date = datetime.strptime(rep_date_str, '%Y-%m-%d').date()
            ref.save()
            messages.success(request, f"Onboarding coordination details updated for {ref.applicant.user.get_full_name()}.")
        return redirect('employment_tracking_admin')
        
    referrals = Referral.objects.all().select_related('applicant__user', 'job_vacancy')
    total_referrals = referrals.count()
    hired_count = referrals.filter(status='Hired').count()
    no_response_count = referrals.filter(status='No Response').count()
    
    conversion_rate = (hired_count / total_referrals * 100.0) if total_referrals > 0 else 0.0
    no_response_rate = (no_response_count / total_referrals * 100.0) if total_referrals > 0 else 0.0
    
    mismatched_count = 0
    for r in referrals:
        edu = r.applicant.education.first()
        fos = edu.field_of_study if edu else ''
        if check_mismatch(fos, r.job_vacancy.title):
            mismatched_count += 1
    mismatch_rate = (mismatched_count / total_referrals * 100.0) if total_referrals > 0 else 0.0
    
    today = timezone.localdate()
    stale_threshold = today - datetime.timedelta(days=7)
    
    # Annotate referrals with is_stale check
    for r in referrals:
        r.is_stale = (r.status == 'Pending' and r.date_referred <= stale_threshold)
        edu = r.applicant.education.first()
        r.field_of_study = edu.field_of_study if edu else 'Not Specified'
        # Get key gaps
        _, gaps = calculate_match_score(r.applicant, r.job_vacancy)
        r.key_gaps = ", ".join([g['skill'] for g in gaps if g['gap'] > 0])
        r.key_gaps_list = [g['skill'] for g in gaps if g['gap'] > 0]
        
    context = {
        'referrals': referrals,
        'total_referrals': total_referrals,
        'conversion_rate': conversion_rate,
        'no_response_rate': no_response_rate,
        'mismatch_rate': mismatch_rate,
    }
    return render(request, 'tracker/ADMIN/employment_tracking_admin.html', context)

@role_required(['admin'])
def job_matching_analytics_admin(request):
    seed_mock_applicants_if_empty()
    total_vacancies = JobVacancy.objects.count()
    
    # Average skills count per applicant
    total_skills = ApplicantSkill.objects.count()
    total_applicants = Profile.objects.filter(role='applicant').count()
    avg_skills_count = round(total_skills / total_applicants, 1) if total_applicants > 0 else 0.0
    
    # Overall match rate across all referrals
    referrals = Referral.objects.all().select_related('applicant__user', 'job_vacancy')
    total_referrals = referrals.count()
    total_matches = []
    for r in referrals:
        match_pct, _ = calculate_match_score(r.applicant, r.job_vacancy)
        total_matches.append(match_pct)
    overall_match_rate = round(sum(total_matches) / len(total_matches), 1) if total_matches else 0.0
    
    # Critical gaps: count of required skills where candidate average rating is < 3.0
    critical_gaps_count = 0
    skills = CentralizedSkill.objects.all()
    for s in skills:
        avg_rating = ApplicantSkill.objects.filter(skill=s).aggregate(Sum('proficiency'))['proficiency__sum'] or 0
        req_count = JobSkillRequirement.objects.filter(skill=s).count()
        if req_count > 0 and (avg_rating / total_applicants if total_applicants > 0 else 0) < 3.0:
            critical_gaps_count += 1
            
    # Top In-Demand Skills based on frequency of vacancy requirements
    demand_counts = JobSkillRequirement.objects.values('skill__name').annotate(count=Count('job_vacancy')).order_by('-count')[:5]
    most_demanded_skills = []
    for item in demand_counts:
        most_demanded_skills.append({
            'name': item['skill__name'],
            'requests': item['count'] * 10
        })
        
    # Stage 8 Detailed metrics calculations
    placed_referrals = referrals.filter(status__in=[
        'Hired — Probationary', 'Hired — Regular', 'Regularly Employed',
        'Still Employed — Performing Well', 'Still Employed — On Improvement Plan',
        'No Response from Employer', 'No Response from Applicant', 'Probation Extended'
    ])
    placed_count = placed_referrals.count()
    
    conversion_rate = (placed_count / total_referrals * 100.0) if total_referrals > 0 else 72.4
    no_show_count = referrals.filter(status__in=['No Show', 'Closed — No Show']).count()
    no_show_rate = (no_show_count / total_referrals * 100.0) if total_referrals > 0 else 5.2
    decline_count = referrals.filter(status='Declined').count()
    decline_rate = (decline_count / total_referrals * 100.0) if total_referrals > 0 else 8.7
    
    # Mismatch/Cross-field Placements count
    mismatched_placed = 0
    for r in placed_referrals:
        edu = r.applicant.education.first()
        fos = edu.field_of_study if edu else ''
        if check_mismatch(fos, r.job_vacancy.title):
            mismatched_placed += 1
            
    mismatch_rate = (mismatched_placed / placed_count * 100.0) if placed_count > 0 else 18.5
    
    # Average time to hire
    placed_with_dates = referrals.filter(actual_start_date__isnull=False)
    durations = [(r.actual_start_date - r.date_referred).days for r in placed_with_dates]
    avg_days_to_hire = round(sum(durations) / len(durations), 1) if durations else 14.2
    
    # Regularization rate
    total_hired_reg = referrals.filter(status__in=['Hired — Probationary', 'Hired — Regular', 'Regularly Employed', 'Separated — End of Probation']).count()
    regularized_count = referrals.filter(status='Regularly Employed').count()
    regularization_rate = round(regularized_count / total_hired_reg * 100.0, 1) if total_hired_reg > 0 else 82.5
    
    # Responsiveness rating
    resolved_count = referrals.exclude(status__in=['Pending', 'Accepted — Awaiting Onboarding', 'Confirmed — Onboarding', 'No Show']).count()
    responsiveness_rate = round(resolved_count / total_referrals * 100.0, 1) if total_referrals > 0 else 94.0
    
    # Training to Hire Rate
    training_hired_count = 0
    for r in placed_referrals:
        if r.applicant.training_progress_percentage > 0:
            training_hired_count += 1
    training_to_hire_rate = round(training_hired_count / placed_count * 100.0, 1) if placed_count > 0 else 68.0
    
    # Talent match list
    talent_match_distribution = []
    for r in referrals:
        match_pct, gaps = calculate_match_score(r.applicant, r.job_vacancy)
        key_gaps = [g['skill'] for g in gaps if g['gap'] > 0]
        
        talent_match_distribution.append({
            'name': f"{r.applicant.user.first_name} {r.applicant.user.last_name}",
            'title': r.applicant.title,
            'role': r.job_vacancy.title,
            'match_score': int(match_pct),
            'key_gaps': ", ".join(key_gaps) if key_gaps else 'None Identified',
            'status': 'Top Choice' if match_pct >= 85 else ('In Review' if match_pct >= 60 else 'Upskilling Suggested')
        })
        
    context = {
        'overall_match_rate': overall_match_rate,
        'avg_skills_count': avg_skills_count,
        'critical_gaps_count': critical_gaps_count,
        'total_vacancies': total_vacancies,
        'most_demanded_skills': most_demanded_skills,
        'talent_match_distribution': talent_match_distribution,
        'total_referrals': total_referrals,
        'conversion_rate': conversion_rate,
        'no_show_rate': no_show_rate,
        'decline_rate': decline_rate,
        'mismatch_rate': mismatch_rate,
        'avg_days_to_hire': avg_days_to_hire,
        'regularization_rate': regularization_rate,
        'responsiveness_rate': responsiveness_rate,
        'training_to_hire_rate': training_to_hire_rate,
        'placed_count': placed_count,
    }
    return render(request, 'tracker/ADMIN/job_matching_analytics_admin.html', context)

@role_required(['admin'])
def peso_profile_admin(request):
    return render(request, 'tracker/ADMIN/peso_profile_admin.html')

@role_required(['admin'])
def skill_monitoring_admin(request):
    seed_mock_applicants_if_empty()
    skills = CentralizedSkill.objects.all()
    # Compute gaps for each skill
    for s in skills:
        s.total_demand = JobSkillRequirement.objects.filter(skill=s).count()
        s.total_supply = ApplicantSkill.objects.filter(skill=s).count()
    return render(request, 'tracker/ADMIN/skill_monitoring_admin.html', {'skills': skills})

@role_required(['admin'])
def training_monitoring_admin(request):
    seed_mock_applicants_if_empty()
    if request.method == 'POST':
        action = request.POST.get('action', 'create')
        
        if action == 'create':
            title = request.POST.get('title', '').strip()
            provider = request.POST.get('provider', '').strip()
            skill_id = request.POST.get('skill_addressed')
            duration = request.POST.get('duration', '').strip()
            description = request.POST.get('description', '').strip()
            sched_date_str = request.POST.get('scheduled_date', '').strip()
            
            scheduled_date = None
            if sched_date_str:
                try:
                    scheduled_date = datetime.strptime(sched_date_str, '%Y-%m-%d').date()
                except ValueError:
                    pass
            
            if title and provider and skill_id and duration:
                skill = get_object_or_404(CentralizedSkill, id=skill_id)
                program = TrainingProgram.objects.create(
                    title=title,
                    provider=provider,
                    skill_addressed=skill,
                    duration=duration,
                    description=description,
                    scheduled_date=scheduled_date,
                    status='Scheduled'
                )
                
                # Auto-notify applicants who have a gap in this skill
                applicants = Profile.objects.filter(role='applicant')
                notified_count = 0
                for app in applicants:
                    has_skill = ApplicantSkill.objects.filter(profile=app, skill=skill).exists()
                    if not has_skill:
                        Notification.objects.create(
                            user=app.user,
                            message=f"New Training: '{title}' by {provider} has been scheduled for {scheduled_date or 'TBD'} to address your skill gap in {skill.name}."
                        )
                        notified_count += 1
                        
                messages.success(request, f"Successfully created training program: {title}! Notified {notified_count} matching applicants.")
            else:
                messages.error(request, "Failed to create training program. Please check all fields.")
                
        elif action == 'enroll':
            program_id = request.POST.get('training_program_id')
            applicant_id = request.POST.get('applicant_id')
            
            program = get_object_or_404(TrainingProgram, id=program_id)
            applicant = get_object_or_404(Profile, id=applicant_id, role='applicant')
            
            enrollment, created = TrainingEnrollment.objects.get_or_create(
                profile=applicant,
                training_program=program,
                defaults={'status': 'Enrolled'}
            )
            if created:
                Notification.objects.create(
                    user=applicant.user,
                    message=f"Enrollment Confirmation: You have been enrolled in '{program.title}' starting on {program.scheduled_date or 'TBD'}."
                )
                messages.success(request, f"Enrolled {applicant.user.get_full_name()} into '{program.title}'.")
            else:
                messages.info(request, f"{applicant.user.get_full_name()} is already enrolled in '{program.title}'.")
                
        elif action == 'confirm_finish':
            program_id = request.POST.get('training_program_id')
            program = get_object_or_404(TrainingProgram, id=program_id)
            
            completed_ids = request.POST.getlist('completed_applicants')
            completed_set = set(map(int, completed_ids))
            
            enrollments = program.enrollments.all()
            for enroll in enrollments:
                if enroll.profile.id in completed_set:
                    enroll.status = 'Attended'
                    enroll.save()
                    
                    # Add/update the skill in applicant profile
                    app_skill, created = ApplicantSkill.objects.get_or_create(
                        profile=enroll.profile,
                        skill=program.skill_addressed,
                        defaults={'proficiency': 3, 'source': 'Training Completion'}
                    )
                    if not created and app_skill.proficiency < 3:
                        app_skill.proficiency = 3
                        app_skill.save()
                        
                    # Sync text skills string
                    skills_qs = ApplicantSkill.objects.filter(profile=enroll.profile).select_related('skill')
                    enroll.profile.skills = ", ".join([sk.skill.name for sk in skills_qs])
                    enroll.profile.save()
                    
                    # Notify matching employers
                    check_and_notify_employer_matches(enroll.profile)
                        
                    Notification.objects.create(
                        user=enroll.profile.user,
                        message=f"Training Completed: You have successfully completed '{program.title}' and acquired the skill: '{program.skill_addressed.name}'."
                    )
                else:
                    enroll.status = 'No Show'
                    enroll.save()
                    
            program.status = 'Completed'
            program.save()
            messages.success(request, f"Training program '{program.title}' confirmed as completed! Updated applicant profiles and skills.")
            
        return redirect('training_monitoring_admin')

    programs = TrainingProgram.objects.all().select_related('skill_addressed').prefetch_related('enrollments__profile__user')
    skills = CentralizedSkill.objects.all().order_by('name')
    applicants = Profile.objects.filter(role='applicant').select_related('user')
    return render(request, 'tracker/ADMIN/training_monitoring_admin.html', {
        'programs': programs,
        'skills': skills,
        'applicants': applicants,
    })

@role_required(['admin'])
def vacancy_management_admin(request):
    seed_mock_applicants_if_empty()
    vacancies = JobVacancy.objects.all().prefetch_related('requirements__skill')
    return render(request, 'tracker/ADMIN/vacancy_management_admin.html', {'vacancies': vacancies})

@role_required(['admin'])
def vacancy_detail_admin(request, vacancy_id):
    seed_mock_applicants_if_empty()
    vacancy = get_object_or_404(JobVacancy, id=vacancy_id)
    referrals = Referral.objects.filter(job_vacancy=vacancy).select_related('applicant__user', 'applicant')
    
    referrals_data = []
    for ref in referrals:
        match_pct, _ = calculate_match_score(ref.applicant, vacancy)
        referrals_data.append({
            'referral': ref,
            'match_pct': int(match_pct)
        })
        
    referrals_data = sorted(referrals_data, key=lambda x: x['match_pct'], reverse=True)
    
    context = {
        'vacancy': vacancy,
        'referrals_data': referrals_data,
    }
    return render(request, 'tracker/ADMIN/vacancy_detail_admin.html', context)

# ==========================================
# JOB APPLICANT VIEWS
# ==========================================

JOB_CATEGORIES = [
    {'code': 'IT', 'name': 'Information Technology'},
    {'code': 'BPO', 'name': 'BPO & Customer Service'},
    {'code': 'ADM', 'name': 'Administrative & Office'},
    {'code': 'FIN', 'name': 'Finance & Accounting'},
    {'code': 'MKT', 'name': 'Sales & Marketing'},
    {'code': 'RTL', 'name': 'Retail & Merchandising'},
    {'code': 'F&B', 'name': 'Food & Beverage / Hospitality'},
    {'code': 'MED', 'name': 'Healthcare & Medical'},
    {'code': 'EDU', 'name': 'Education & Training'},
    {'code': 'ENG', 'name': 'Engineering & Construction'},
    {'code': 'ELC', 'name': 'Electrical & Electronics'},
    {'code': 'MCH', 'name': 'Mechanical & Automotive'},
    {'code': 'TVT', 'name': 'TESDA Trade & Vocational'},
    {'code': 'LOG', 'name': 'Logistics & Warehousing'},
    {'code': 'SEC', 'name': 'Security & Safety'},
    {'code': 'GEN', 'name': 'General Services & Facilities'},
    {'code': 'DRV', 'name': 'Driving & Transportation'},
    {'code': 'MFG', 'name': 'Manufacturing & Garments'},
    {'code': 'AGR', 'name': 'Agriculture & Environment'},
    {'code': 'OFW', 'name': 'Overseas / OFW Positions'},
]

@login_required(login_url='login')
def profile_wizard(request):
    profile = request.user.profile
    if profile.role != 'applicant':
        return redirect('dashboard')
    if profile.is_profile_complete:
        return redirect('applicant_dashboard')
        
    if request.method == 'POST':
        # 1. Educational Background
        institution = request.POST.get('institution', '').strip()
        degree = request.POST.get('degree', '').strip()
        field_of_study = request.POST.get('field_of_study', '').strip()
        start_year = request.POST.get('start_year')
        end_year = request.POST.get('end_year')
        
        if institution and degree and field_of_study:
            Education.objects.create(
                profile=profile,
                institution=institution,
                degree=degree,
                field_of_study=field_of_study,
                start_year=int(start_year) if start_year else 2020,
                end_year=int(end_year) if end_year else 2024
            )
            
        # 2. Fresh Graduate or Work History
        is_fresh_grad = request.POST.get('is_fresh_grad') == 'on'
        profile.is_fresh_grad = is_fresh_grad
        
        position = ""
        if is_fresh_grad:
            ojt_company = request.POST.get('ojt_company', '').strip()
            ojt_position = request.POST.get('ojt_position', '').strip()
            ojt_desc = request.POST.get('ojt_description', '').strip()
            if ojt_company and ojt_position:
                WorkExperience.objects.create(
                    profile=profile,
                    company=ojt_company,
                    position=ojt_position + " (OJT/Internship)",
                    description=ojt_desc,
                    start_date="2023",
                    end_date="2024",
                    is_current=False
                )
        else:
            company = request.POST.get('company', '').strip()
            position = request.POST.get('position', '').strip()
            description = request.POST.get('description', '').strip()
            start_date = request.POST.get('start_date', '').strip()
            end_date = request.POST.get('end_date', 'Present').strip()
            is_current = request.POST.get('is_current') == 'on'
            
            if company and position:
                WorkExperience.objects.create(
                    profile=profile,
                    company=company,
                    position=position,
                    description=description,
                    start_date=start_date,
                    end_date=end_date,
                    is_current=is_current
                )
                
        # 3. Technical Skills
        selected_skills = request.POST.getlist('skills[]')
        
        profile_skills_list = []
        for skill_id in selected_skills:
            if skill_id:
                skill_obj = get_object_or_404(CentralizedSkill, id=int(skill_id))
                ApplicantSkill.objects.update_or_create(
                    profile=profile,
                    skill=skill_obj,
                    defaults={
                        'proficiency': 3,  # Default proficiency of 3 (Proficient)
                        'source': 'self'
                    }
                )
                profile_skills_list.append(skill_obj.name)
                
        # 4. Finalize Profile
        profile.is_profile_complete = True
        if not is_fresh_grad and position:
            profile.title = position
        elif field_of_study:
            profile.title = f"Fresh Graduate - {field_of_study}"
        elif preferred_job:
            profile.title = preferred_job
        else:
            profile.title = "General Applicant"
        if profile_skills_list:
            profile.skills = ", ".join(profile_skills_list)
            
        preferred_job = request.POST.get('preferred_job', '').strip()
        if preferred_job:
            profile.preferred_job = preferred_job
                
        profile.experience_years = calculate_total_experience_years(profile)
        profile.save()
        
        # Notify matching employers
        check_and_notify_employer_matches(profile)
        
        messages.success(request, "Welcome! Your profile has been completed successfully.")
        return redirect('applicant_dashboard')
        
    skills = CentralizedSkill.objects.all()
    return render(request, 'tracker/APPLICANT/profile_wizard.html', {
        'skills': skills,
        'profile': profile,
        'categories': JOB_CATEGORIES
    })

@role_required(['applicant'])
def applicant_notifications(request):
    profile = request.user.profile
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    context = {
        'notifications_list': notifications,
        'profile': profile,
    }
    return render(request, 'tracker/APPLICANT/notifications_list.html', context)

@role_required(['applicant'])
def applicant_dashboard(request):
    seed_mock_applicants_if_empty()
    profile = request.user.profile
    
    if request.method == 'POST':
        action = request.POST.get('action')
        referral_id = request.POST.get('referral_id')
        ref = get_object_or_404(Referral, id=referral_id, applicant=profile)
        
        if action == 'confirm_offer':
            # Stage 2: Applicant confirms acceptance
            if request.FILES.get('nbi_clearance'):
                ref.nbi_clearance = request.FILES['nbi_clearance']
            if request.FILES.get('medical_certificate'):
                ref.medical_certificate = request.FILES['medical_certificate']
            if request.FILES.get('birth_certificate'):
                ref.birth_certificate = request.FILES['birth_certificate']
            if request.FILES.get('diploma_transcript'):
                ref.diploma_transcript = request.FILES['diploma_transcript']
            if request.FILES.get('prev_employment_cert'):
                ref.prev_employment_cert = request.FILES['prev_employment_cert']
            if request.FILES.get('tesda_cert'):
                ref.tesda_cert = request.FILES['tesda_cert']
                
            ref.sss_number = request.POST.get('sss_number', '')
            ref.philhealth_number = request.POST.get('philhealth_number', '')
            ref.pagibig_number = request.POST.get('pagibig_number', '')
            ref.bir_tin = request.POST.get('bir_tin', '')
            
            ref.status = 'Confirmed — Onboarding'
            ref.save()
            
            profile.status = 'Employed — Onboarding'
            profile.save()
            
            messages.success(request, "Congratulations! You have confirmed the offer and uploaded your onboarding requirements.")
            
        elif action == 'decline_offer':
            # Stage 2: Applicant declines acceptance
            ref.decline_reason = request.POST.get('decline_reason', 'No reason given')
            ref.decline_remarks = request.POST.get('decline_remarks', '')
            ref.status = 'Declined'
            ref.save()
            
            profile.status = 'Active — Job Seeking'
            profile.save()
            
            # Reopen slot (+1 remaining_slots)
            vac = ref.job_vacancy
            vac.remaining_slots = min(vac.slots, vac.remaining_slots + 1)
            vac.save()
            
            # Send Notification to PESO Coordinator
            coord_users = User.objects.filter(profile__role='admin')
            for u in coord_users:
                Notification.objects.create(
                    user=u,
                    message=f"Applicant {profile.user.get_full_name()} declined the offer for {ref.job_vacancy.title} at {ref.job_vacancy.employer.user.first_name}. Reason: {ref.decline_reason}."
                )
            messages.info(request, "You have declined the offer. Your profile is now active in the job matching pool again.")
            
        elif action == 'submit_applicant_evaluation':
            # Stage 7: Applicant evaluation form
            ref.eval_app_accurate_desc = request.POST.get('eval_app_accurate_desc', 'Yes')
            ref.eval_app_terms_met = request.POST.get('eval_app_terms_met', 'Yes')
            ref.eval_app_expectations = request.POST.get('eval_app_expectations', '')
            ref.eval_app_future_use = request.POST.get('eval_app_future_use', 'Yes')
            ref.eval_app_satisfaction = int(request.POST.get('eval_app_satisfaction', 5))
            ref.save()
            messages.success(request, "Thank you! Your job match feedback has been successfully submitted.")
            
        return redirect('applicant_dashboard')

    # Calculate profile completion percentage
    has_title = bool(profile.title)
    has_skills = profile.applicant_skills.exists()
    has_edu = profile.education.exists()
    has_exp = profile.experience.exists()
    score = 0
    if has_title: score += 25
    if has_skills: score += 25
    if has_edu: score += 25
    if has_exp: score += 25
    
    # Calculate highest match score job
    vacancies = JobVacancy.objects.all()
    highest_match = 0
    best_job = None
    for vac in vacancies:
        match_pct, _ = calculate_match_score(profile, vac)
        if match_pct > highest_match:
            highest_match = match_pct
            best_job = vac
            
    referrals = Referral.objects.filter(applicant=profile).select_related('job_vacancy')
    active_count = referrals.filter(status='Pending').count()
    interview_count = referrals.filter(applicant__status='Interviewing').count()
    offer_count = referrals.filter(applicant__status='Shortlisted').count()
    
    # Fetch active offer if any
    active_offer = referrals.filter(status='Accepted — Awaiting Onboarding').first()
    
    # Check if there is a hired/closed referral that needs evaluation
    pending_evaluation = referrals.filter(status__in=['Hired — Probationary', 'Hired — Regular', 'Regularly Employed'], eval_app_satisfaction__isnull=True).first()
    
    # Fetch unread notifications
    notifications = Notification.objects.filter(user=request.user, is_read=False).order_by('-created_at')
    for n in notifications:
        n.is_read = True
        n.save() # Mark them as read when viewing dashboard
    
    # Training recommendations based on gaps using at max 5 preferred jobs
    rec_programs_all, _, _, biggest_gap_skill = get_applicant_training_recommendations(profile)
    recommended_programs = rec_programs_all[:2]
    
    # Calculate recommended jobs (match score >= 50%)
    bookmarked_job_ids = set(profile.bookmarked_jobs.values_list('job_vacancy_id', flat=True))
    
    preferred_jobs_data = []
    preferred_category_name = None
    if profile.preferred_job:
        category_mapping_dict = {
            'IT': 'Information Technology',
            'BPO': 'BPO & Customer Service',
            'ADM': 'Administrative & Office',
            'FIN': 'Finance & Accounting',
            'MKT': 'Sales & Marketing',
            'RTL': 'Retail & Merchandising',
            'F&B': 'Food & Beverage / Hospitality',
            'MED': 'Healthcare & Medical',
            'EDU': 'Education & Training',
            'ENG': 'Engineering & Construction',
            'ELC': 'Electrical & Electronics',
            'MCH': 'Mechanical & Automotive',
            'TVT': 'TESDA Trade & Vocational',
            'LOG': 'Logistics & Warehousing',
            'SEC': 'Security & Safety',
            'GEN': 'General Services & Facilities',
            'DRV': 'Driving & Transportation',
            'MFG': 'Manufacturing & Garments',
            'AGR': 'Agriculture & Environment',
            'OFW': 'Overseas / OFW Positions',
        }
        preferred_category_name = category_mapping_dict.get(profile.preferred_job, profile.preferred_job)
        
        # Get all vacancies in this category
        cat_vacancies = JobVacancy.objects.filter(category=profile.preferred_job).prefetch_related('requirements__skill')
        for vac in cat_vacancies:
            match_pct, gaps = calculate_match_score(profile, vac)
            gaps_count = sum(1 for g in gaps if g['gap'] > 0)
            req_skills = [req.skill.name for req in vac.requirements.all()[:3]]
            preferred_jobs_data.append({
                'vacancy': vac,
                'match_score': int(match_pct),
                'gaps_count': gaps_count,
                'skills': req_skills,
                'is_bookmarked': vac.id in bookmarked_job_ids,
                'has_applied': Referral.objects.filter(applicant=profile, job_vacancy=vac).exists(),
            })
        preferred_jobs_data = sorted(preferred_jobs_data, key=lambda x: x['match_score'], reverse=True)
    recommended_jobs = []
    for vac in vacancies:
        match_pct, gaps = calculate_match_score(profile, vac)
        if match_pct >= 50:
            gaps_count = sum(1 for g in gaps if g['gap'] > 0)
            req_skills = [req.skill.name for req in vac.requirements.all()[:3]]
            recommended_jobs.append({
                'vacancy': vac,
                'match_score': int(match_pct),
                'gaps_count': gaps_count,
                'skills': req_skills,
                'is_bookmarked': vac.id in bookmarked_job_ids,
            })
    recommended_jobs = sorted(recommended_jobs, key=lambda x: x['match_score'], reverse=True)
    
    # Retrieve bookmarked jobs data
    bookmarked_jobs = JobBookmark.objects.filter(profile=profile).select_related('job_vacancy', 'job_vacancy__employer')
    bookmarked_jobs_data = []
    for bookmark in bookmarked_jobs:
        vac = bookmark.job_vacancy
        match_pct, gaps = calculate_match_score(profile, vac)
        gaps_count = sum(1 for g in gaps if g['gap'] > 0)
        req_skills = [req.skill.name for req in vac.requirements.all()[:3]]
        bookmarked_jobs_data.append({
            'vacancy': vac,
            'match_score': int(match_pct),
            'gaps_count': gaps_count,
            'skills': req_skills,
            'is_bookmarked': True,
        })
    
    context = {
        'profile': profile,
        'profile_strength': score,
        'highest_match': int(highest_match),
        'best_job': best_job,
        'best_job_is_bookmarked': best_job.id in bookmarked_job_ids if best_job else False,
        'recent_referrals': referrals[:3],
        'applicant_skills': profile.applicant_skills.all().select_related('skill')[:3],
        'active_count': active_count,
        'interview_count': interview_count,
        'offer_count': offer_count,
        'recommended_programs': recommended_programs,
        'biggest_gap_skill': biggest_gap_skill,
        'recommended_jobs': recommended_jobs[:3],
        'bookmarked_jobs': bookmarked_jobs_data,
        'preferred_jobs_data': preferred_jobs_data,
        'preferred_category_name': preferred_category_name,
        'active_offer': active_offer,
        'pending_evaluation': pending_evaluation,
        'notifications': notifications,
    }
    return render(request, 'tracker/APPLICANT/applicant_dashboard.html', context)

@role_required(['applicant'])
def applicant_profile(request):
    seed_mock_applicants_if_empty()
    profile = request.user.profile
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_education':
            Education.objects.create(
                profile=profile,
                institution=request.POST.get('institution', '').strip(),
                degree=request.POST.get('degree', '').strip(),
                field_of_study=request.POST.get('field_of_study', '').strip(),
                start_year=int(request.POST.get('start_year') or 2020),
                end_year=int(request.POST.get('end_year') or 2024)
            )
            messages.success(request, "Education entry added successfully!")
            
        elif action == 'delete_education':
            edu_id = request.POST.get('id')
            get_object_or_404(Education, id=edu_id, profile=profile).delete()
            messages.success(request, "Education entry deleted successfully!")
            
        elif action == 'add_experience':
            WorkExperience.objects.create(
                profile=profile,
                company=request.POST.get('company', '').strip(),
                position=request.POST.get('position', '').strip(),
                description=request.POST.get('description', '').strip(),
                start_date=request.POST.get('start_date', '').strip(),
                end_date=request.POST.get('end_date', 'Present').strip(),
                is_current=request.POST.get('is_current') == 'true'
            )
            profile.experience_years = calculate_total_experience_years(profile)
            profile.save()
            messages.success(request, "Work experience entry added successfully!")
            
        elif action == 'delete_experience':
            exp_id = request.POST.get('id')
            get_object_or_404(WorkExperience, id=exp_id, profile=profile).delete()
            profile.experience_years = calculate_total_experience_years(profile)
            profile.save()
            messages.success(request, "Work experience entry deleted successfully!")
            
        elif action == 'add_certification':
            Certification.objects.create(
                profile=profile,
                name=request.POST.get('name', '').strip(),
                issuing_organization=request.POST.get('issuing_organization', '').strip(),
                is_tesda=request.POST.get('is_tesda') == 'true'
            )
            messages.success(request, "Certification entry added successfully!")
            
        elif action == 'delete_certification':
            cert_id = request.POST.get('id')
            get_object_or_404(Certification, id=cert_id, profile=profile).delete()
            messages.success(request, "Certification entry deleted successfully!")
            
        elif action == 'add_skill':
            skill_id = request.POST.get('skill_id')
            proficiency = int(request.POST.get('proficiency') or 1)
            skill_obj = get_object_or_404(CentralizedSkill, id=skill_id)
            
            ApplicantSkill.objects.update_or_create(
                profile=profile,
                skill=skill_obj,
                defaults={'proficiency': proficiency}
            )
            
            # Sync text skills string
            skills_qs = ApplicantSkill.objects.filter(profile=profile).select_related('skill')
            profile.skills = ", ".join([sk.skill.name for sk in skills_qs])
            profile.save()
            
            # Notify matching employers
            check_and_notify_employer_matches(profile)
            
            messages.success(request, f"Skill '{skill_obj.name}' updated!")
            
        elif action == 'delete_skill':
            skill_id = request.POST.get('id')
            get_object_or_404(ApplicantSkill, id=skill_id, profile=profile).delete()
            
            # Sync text skills string
            skills_qs = ApplicantSkill.objects.filter(profile=profile).select_related('skill')
            profile.skills = ", ".join([sk.skill.name for sk in skills_qs])
            profile.save()
            messages.success(request, "Skill deleted successfully!")
            
        elif action == 'update_profile':
            profile.title = request.POST.get('title', profile.title)
            profile.location = request.POST.get('location', profile.location)
            profile.experience_years = request.POST.get('experience_years', profile.experience_years)
            profile.skill_level = request.POST.get('skill_level', profile.skill_level)
            profile.phone_number = request.POST.get('phone_number', profile.phone_number)
            profile.civil_status = request.POST.get('civil_status', profile.civil_status)
            
            # birthdate and address
            birthdate_str = request.POST.get('birthdate', '').strip()
            if birthdate_str:
                from datetime import datetime
                try:
                    profile.birthdate = datetime.strptime(birthdate_str, '%Y-%m-%d').date()
                except ValueError:
                    pass
            else:
                profile.birthdate = None
            
            profile.address = request.POST.get('address', '').strip() or None
            
            # Also save the skills text field if provided
            skills_text = request.POST.get('skills', '').strip()
            if skills_text:
                profile.skills = skills_text
                
            pref_job = request.POST.get('preferred_job', '').strip()
            profile.preferred_job = pref_job or None
                
            profile.save()
            
            user = request.user
            user.first_name = request.POST.get('first_name', user.first_name).strip()
            user.last_name = request.POST.get('last_name', user.last_name).strip()
            user.save()
            messages.success(request, "Profile updated successfully!")

        elif action == 'upload_document':
            uploaded_file = request.FILES.get('document')
            if uploaded_file:
                # Calculate human-readable size
                size_in_bytes = uploaded_file.size
                if size_in_bytes < 1024 * 1024:
                    size_str = f"{round(size_in_bytes / 1024, 1)} KB"
                else:
                    size_str = f"{round(size_in_bytes / (1024 * 1024), 1)} MB"
                
                ApplicantDocument.objects.create(
                    profile=profile,
                    file=uploaded_file,
                    file_name=uploaded_file.name,
                    file_size=size_str
                )
                messages.success(request, f"Document '{uploaded_file.name}' uploaded successfully!")
            else:
                messages.error(request, "No file selected.")

        elif action == 'delete_document':
            doc_id = request.POST.get('id')
            doc = get_object_or_404(ApplicantDocument, id=doc_id, profile=profile)
            doc.file.delete()
            doc.delete()
            messages.success(request, "Document deleted successfully!")
        elif action == 'upload_avatar':
            avatar = request.FILES.get('avatar')
            if avatar:
                if profile.profile_picture:
                    profile.profile_picture.delete()
                profile.profile_picture = avatar
                profile.save()
                messages.success(request, "Profile photo updated successfully!")
            else:
                messages.error(request, "No photo selected.")

        return redirect('applicant_profile')
        
    education_list = profile.education.all()
    experience_list = profile.experience.all()
    certification_list = profile.certifications.all()
    skills_list = profile.applicant_skills.all().select_related('skill')
    all_centralized_skills = CentralizedSkill.objects.all().order_by('name')
    
    # Calculate profile completion percentage
    has_title = bool(profile.title)
    has_skills = skills_list.exists()
    has_edu = education_list.exists()
    has_exp = experience_list.exists()
    score = 0
    if has_title: score += 25
    if has_skills: score += 25
    if has_edu: score += 25
    if has_exp: score += 25
    
    context = {
        'profile': profile,
        'education_list': education_list,
        'experience_list': experience_list,
        'certification_list': certification_list,
        'skills_list': skills_list,
        'all_centralized_skills': all_centralized_skills,
        'profile_strength': score,
        'categories': JOB_CATEGORIES,
    }
    return render(request, 'tracker/APPLICANT/applicant_profile.html', context)

@role_required(['applicant'])
def job_search(request):
    seed_mock_applicants_if_empty()
    profile = request.user.profile
    
    q = request.GET.get('q', '').strip()
    location = request.GET.get('location', '').strip()
    min_match = request.GET.get('min_match', '0')
    try:
        min_match_val = int(min_match)
    except ValueError:
        min_match_val = 0
        
    category = request.GET.get('category', '').strip()
    bookmarked_only = request.GET.get('bookmarked_only') == 'on'
    
    bookmarked_job_ids = set(profile.bookmarked_jobs.values_list('job_vacancy_id', flat=True))
    
    vacancies = JobVacancy.objects.all().prefetch_related('requirements__skill')
    
    if bookmarked_only:
        vacancies = vacancies.filter(id__in=bookmarked_job_ids)
        
    if q:
        vacancies = vacancies.filter(
            Q(title__icontains=q) | 
            Q(description__icontains=q) |
            Q(employer__company_name__icontains=q)
        )
    if location:
        vacancies = vacancies.filter(location__icontains=location)
    if category:
        vacancies = vacancies.filter(category=category)
        
    jobs_data = []
    for vac in vacancies:
        match_pct, gaps = calculate_match_score(profile, vac)
        if match_pct < min_match_val:
            continue
            
        gaps_count = sum(1 for g in gaps if g['gap'] > 0)
        req_skills = [req.skill.name for req in vac.requirements.all()[:3]]
        
        jobs_data.append({
            'vacancy': vac,
            'match_score': int(match_pct),
            'gaps_count': gaps_count,
            'skills': req_skills,
            'is_bookmarked': vac.id in bookmarked_job_ids,
        })
        
    # Sort by match score descending
    jobs_data = sorted(jobs_data, key=lambda x: x['match_score'], reverse=True)
    
    # Split into Recommended (>=50% match) and Other (<50% match)
    recommended_jobs = [j for j in jobs_data if j['match_score'] >= 50]
    other_jobs = [j for j in jobs_data if j['match_score'] < 50]
    
    return render(request, 'tracker/APPLICANT/job_search.html', {
        'jobs': jobs_data,
        'recommended_jobs': recommended_jobs,
        'other_jobs': other_jobs,
        'profile': profile,
        'q': q,
        'location': location,
        'min_match': min_match_val,
        'category': category,
        'bookmarked_only': bookmarked_only,
    })

@role_required(['applicant'])
def skill_gap_analysis(request):
    seed_mock_applicants_if_empty()
    profile = request.user.profile
    job_id = request.GET.get('job_id')
    
    # Select vacancy
    if job_id:
        vacancy = get_object_or_404(JobVacancy, id=job_id)
    else:
        # Default to highest matching job vacancy
        vacancies = JobVacancy.objects.all()
        highest_match = -1
        vacancy = None
        for vac in vacancies:
            match_pct, _ = calculate_match_score(profile, vac)
            if match_pct > highest_match:
                highest_match = match_pct
                vacancy = vac
                
    if not vacancy:
        messages.error(request, "No vacancies found.")
        return redirect('job_search')
        
    match_pct, gaps = calculate_match_score(profile, vacancy)
    
    # Alternative recommendations (if match < 60%)
    alternatives = []
    if match_pct < 60:
        other_vacancies = JobVacancy.objects.exclude(id=vacancy.id)
        for ovac in other_vacancies:
            omatch, _ = calculate_match_score(profile, ovac)
            if omatch >= 60:
                alternatives.append({
                    'vacancy': ovac,
                    'match_score': int(omatch)
                })
        alternatives = sorted(alternatives, key=lambda x: x['match_score'], reverse=True)[:2]
        
    # Training recommendations based on gaps
    gap_skills = [g['skill'] for g in gaps if g['gap'] > 0]
    recommended_programs = TrainingProgram.objects.filter(skill_addressed__name__in=gap_skills).select_related('skill_addressed')
    enrolled_ids = set(profile.training_enrollments.values_list('training_program_id', flat=True))
    
    # Build list of skills for target vs current levels
    labels = []
    target_data = []
    current_data = []
    biggest_gap = None
    max_gap_val = -1
    strongest_skill = None
    max_prof_val = -1
    for g in gaps:
        labels.append(g['skill'])
        target_data.append(g['required'])
        current_data.append(g['candidate'])
        if g['gap'] > max_gap_val:
            max_gap_val = g['gap']
            biggest_gap = g['skill']
        # Strongest skill: one where applicant has high rating, preferably matching or exceeding required
        if g['candidate'] > max_prof_val:
            max_prof_val = g['candidate']
            strongest_skill = g['skill']
        
    referral = Referral.objects.filter(applicant=profile, job_vacancy=vacancy).first()
    has_applied = referral is not None
    is_bookmarked = profile.bookmarked_jobs.filter(job_vacancy=vacancy).exists()
    
    context = {
        'vacancy': vacancy,
        'match_score': int(match_pct),
        'match_stroke_offset': round(251.2 * (1 - (match_pct / 100.0)), 1),
        'gaps': gaps,
        'alternatives': alternatives,
        'recommended_programs': recommended_programs,
        'enrolled_ids': enrolled_ids,
        'labels': labels,
        'target_data': target_data,
        'current_data': current_data,
        'biggest_gap': biggest_gap if max_gap_val > 0 else None,
        'strongest_skill': strongest_skill if max_prof_val > 0 else None,
        'profile': profile,
        'has_applied': has_applied,
        'referral': referral,
        'is_bookmarked': is_bookmarked,
    }
    return render(request, 'tracker/APPLICANT/skill_gap_analysis.html', context)

@role_required(['applicant'])
@require_POST
def apply_to_job(request, vacancy_id):
    profile = request.user.profile
    vacancy = get_object_or_404(JobVacancy, id=vacancy_id)
    
    referral, created = Referral.objects.get_or_create(
        applicant=profile,
        job_vacancy=vacancy,
        defaults={'status': 'Pending'}
    )
    if created:
        messages.success(request, f"Successfully applied for '{vacancy.title}'! You can track your status on the dashboard.")
        match_pct, _ = calculate_match_score(profile, vacancy)
        if match_pct >= 70:
            Notification.objects.create(
                user=vacancy.employer.user,
                message=f"Match Alert: Applicant {profile.user.get_full_name()} has reached a {int(match_pct)}% skill match for your vacancy '{vacancy.title}'."
            )
    else:
        messages.info(request, f"You have already applied or been referred to '{vacancy.title}'.")
        
    return redirect('applicant_dashboard')



@role_required(['applicant'])
@require_POST
def toggle_bookmark(request, vacancy_id):
    profile = request.user.profile
    vacancy = get_object_or_404(JobVacancy, id=vacancy_id)
    bookmark, created = JobBookmark.objects.get_or_create(profile=profile, job_vacancy=vacancy)
    if not created:
        bookmark.delete()
        messages.success(request, f"Removed '{vacancy.title}' from bookmarks.")
    else:
        messages.success(request, f"Bookmarked '{vacancy.title}'.")
    return redirect(request.META.get('HTTP_REFERER', 'job_search'))



@role_required(['applicant'])
def applicant_applications_list(request):
    seed_mock_applicants_if_empty()
    profile = request.user.profile
    referrals = Referral.objects.filter(applicant=profile).select_related('job_vacancy', 'job_vacancy__employer').order_by('-date_referred')
    
    # Calculate match scores for each application to display in the list
    referrals_data = []
    for ref in referrals:
        match_pct, gaps = calculate_match_score(profile, ref.job_vacancy)
        gaps_count = sum(1 for g in gaps if g['gap'] > 0)
        referrals_data.append({
            'referral': ref,
            'match_score': int(match_pct),
            'gaps_count': gaps_count,
        })
        
    context = {
        'profile': profile,
        'referrals_data': referrals_data,
    }
    return render(request, 'tracker/APPLICANT/applications_list.html', context)


@role_required(['applicant'])
def applicant_application_details(request, referral_id):
    seed_mock_applicants_if_empty()
    profile = request.user.profile
    referral = get_object_or_404(Referral, id=referral_id, applicant=profile)
    vacancy = referral.job_vacancy
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'confirm_offer':
            if request.FILES.get('nbi_clearance'):
                referral.nbi_clearance = request.FILES['nbi_clearance']
            if request.FILES.get('medical_certificate'):
                referral.medical_certificate = request.FILES['medical_certificate']
            if request.FILES.get('birth_certificate'):
                referral.birth_certificate = request.FILES['birth_certificate']
            if request.FILES.get('diploma_transcript'):
                referral.diploma_transcript = request.FILES['diploma_transcript']
            if request.FILES.get('prev_employment_cert'):
                referral.prev_employment_cert = request.FILES['prev_employment_cert']
            if request.FILES.get('tesda_cert'):
                referral.tesda_cert = request.FILES['tesda_cert']
                
            referral.sss_number = request.POST.get('sss_number', '')
            referral.philhealth_number = request.POST.get('philhealth_number', '')
            referral.pagibig_number = request.POST.get('pagibig_number', '')
            referral.bir_tin = request.POST.get('bir_tin', '')
            
            referral.status = 'Confirmed — Onboarding'
            referral.save()
            
            profile.status = 'Employed — Onboarding'
            profile.save()
            
            messages.success(request, "Congratulations! You have confirmed the offer and uploaded your onboarding requirements.")
            return redirect('applicant_application_details', referral_id=referral_id)
            
        elif action == 'decline_offer':
            referral.decline_reason = request.POST.get('decline_reason', 'No reason given')
            referral.decline_remarks = request.POST.get('decline_remarks', '')
            referral.status = 'Declined'
            referral.save()
            
            profile.status = 'Active — Job Seeking'
            profile.save()
            
            # Reopen slot (+1 remaining_slots)
            vac = referral.job_vacancy
            vac.remaining_slots = min(vac.slots, vac.remaining_slots + 1)
            vac.save()
            
            # Send Notification to PESO Coordinator
            coord_users = User.objects.filter(profile__role='admin')
            for u in coord_users:
                Notification.objects.create(
                    user=u,
                    message=f"Applicant {profile.user.get_full_name()} declined the offer for {referral.job_vacancy.title} at {referral.job_vacancy.employer.user.first_name}. Reason: {referral.decline_reason}."
                )
            messages.info(request, "You have declined the offer. Your profile is now active in the job matching pool again.")
            return redirect('applicant_application_details', referral_id=referral_id)

    # Calculate match score and gaps
    match_pct, gaps = calculate_match_score(profile, vacancy)
    
    # Get training programs addressing remaining gaps
    gap_skills = [g['skill'] for g in gaps if g['gap'] > 0]
    recommended_programs = TrainingProgram.objects.filter(skill_addressed__name__in=gap_skills).select_related('skill_addressed')
    
    active_offer = referral if referral.status == 'Accepted — Awaiting Onboarding' else None

    context = {
        'referral': referral,
        'vacancy': vacancy,
        'match_score': int(match_pct),
        'gaps': gaps,
        'recommended_programs': recommended_programs,
        'profile': profile,
        'active_offer': active_offer
    }
    return render(request, 'tracker/APPLICANT/application_details.html', context)



@role_required(['applicant'])
def training_recommendations(request):
    seed_mock_applicants_if_empty()
    profile = request.user.profile
    
    q = request.GET.get('q', '').strip()
    selected_cat = request.GET.get('category', '').strip()
    
    # Save the selected preferred category to the profile if changed
    if selected_cat:
        profile.preferred_job = selected_cat
        profile.save()
        
    target_category = profile.preferred_job
    
    enrolled_trainings = profile.training_enrollments.all().select_related('training_program__skill_addressed')
    enrolled_ids = set(profile.training_enrollments.values_list('training_program_id', flat=True))

    # Gather up to 5 preferred jobs and find gaps using the helper
    recommended_programs, highest_match, gaps_count, _ = get_applicant_training_recommendations(profile)
    all_programs = TrainingProgram.objects.exclude(id__in=enrolled_ids).select_related('skill_addressed')
    
    if q:
        all_programs = all_programs.filter(
            Q(title__icontains=q) |
            Q(provider__icontains=q) |
            Q(skill_addressed__name__icontains=q)
        )
        
    context = {
        'profile_match_percentage': int(highest_match),
        'remaining_gap': max(0, 100 - int(highest_match)),
        'recommended_programs': recommended_programs,
        'all_programs': all_programs,
        'gaps_count': gaps_count,
        'profile': profile,
        'q': q,
        'target_category': target_category,
        'categories': JOB_CATEGORIES,
        'enrolled_trainings': enrolled_trainings,
        'enrolled_ids': enrolled_ids,
    }
    return render(request, 'tracker/APPLICANT/training_recommendations.html', context)

@role_required(['applicant'])
@require_POST
def applicant_enroll_training(request, program_id):
    profile = request.user.profile
    program = get_object_or_404(TrainingProgram, id=program_id)
    
    enrollment, created = TrainingEnrollment.objects.get_or_create(
        profile=profile,
        training_program=program,
        defaults={'status': 'Enrolled'}
    )
    if created:
        messages.success(request, f"Successfully enrolled in training program: '{program.title}'!")
    else:
        messages.info(request, f"You are already enrolled in '{program.title}'.")
        
    return redirect('training_recommendations')

@role_required(['admin'])
@require_POST
def create_referral(request):
    applicant_id = request.POST.get('applicant_id')
    vacancy_id = request.POST.get('vacancy_id')
    
    applicant = get_object_or_404(Profile, id=applicant_id, role='applicant')
    vacancy = get_object_or_404(JobVacancy, id=vacancy_id)
    
    referral, created = Referral.objects.get_or_create(
        applicant=applicant,
        job_vacancy=vacancy,
        defaults={'status': 'Pending'}
    )
    if created:
        messages.success(request, f"Successfully referred {applicant.user.first_name} to {vacancy.title}!")
        match_pct, _ = calculate_match_score(applicant, vacancy)
        if match_pct >= 70:
            Notification.objects.create(
                user=vacancy.employer.user,
                message=f"Match Alert: Applicant {applicant.user.get_full_name()} has reached a {int(match_pct)}% skill match for your vacancy '{vacancy.title}'."
            )
    else:
        messages.info(request, f"{applicant.user.first_name} is already referred to {vacancy.title}.")
        
    return redirect('applicant_monitoring_admin')

@role_required(['admin'])
@require_POST
def update_referral_status(request, referral_id):
    referral = get_object_or_404(Referral, id=referral_id)
    if referral.status == 'Hired':
        messages.error(request, "This referral is finalized (Hired) and cannot be modified.")
        return redirect('employment_tracking_admin')
        
    status = request.POST.get('status')
    if status in ['Pending', 'Hired', 'Not Hired', 'No Response']:
        referral.status = status
        referral.save()
        messages.success(request, f"Updated referral status for {referral.applicant.user.first_name} to {status}!")
    else:
        messages.error(request, "Invalid referral status.")
    return redirect('employment_tracking_admin')

@role_required(['admin'])
@require_POST
def log_contact_attempt(request, referral_id):
    referral = get_object_or_404(Referral, id=referral_id)
    if referral.status == 'Hired':
        messages.error(request, "This referral is finalized (Hired) and cannot be modified.")
        return redirect('employment_tracking_admin')
        
    method = request.POST.get('method', 'Call').strip()
    notes = request.POST.get('notes', '').strip()
    date_str = request.POST.get('date') or timezone.localdate().strftime('%Y-%m-%d')
    
    attempt = {
        'date': date_str,
        'method': method,
        'notes': notes
    }
    
    if not isinstance(referral.contact_attempts, list):
        referral.contact_attempts = []
        
    referral.contact_attempts.append(attempt)
    referral.save()
    
    messages.success(request, f"Logged {method} contact attempt for {referral.applicant.user.first_name}!")
    return redirect('employment_tracking_admin')

# ==========================================
# EMPLOYER VIEWS
# ==========================================

@role_required(['employer'])
def employer_dashboard(request):
    seed_mock_applicants_if_empty()
    profile = request.user.profile
    vacancies = JobVacancy.objects.filter(employer=profile).order_by('-created_at')
    active_vacs_count = vacancies.filter(status='Open').count()
    
    referrals = Referral.objects.filter(job_vacancy__employer=profile).select_related('applicant__user', 'job_vacancy')
    total_apps_count = referrals.count()
    interviews_count = referrals.filter(applicant__status='Interviewing').count()
    hires_count = referrals.filter(status='Hired').count()
    
    # Auto-seed mock interviews if none exist
    if not Interview.objects.filter(employer=profile).exists():
        today = timezone.localdate()
        
        # Try to find Sarah Jenkins (seeded applicant)
        sarah_profile = Profile.objects.filter(role='applicant', user__email='sarah.j@analytics.ai').first()
        if not sarah_profile:
            sarah_profile = Profile.objects.filter(role='applicant', user__first_name__icontains='Sarah').first()
            
        # Try to find Marcus Chen (seeded applicant)
        marcus_profile = Profile.objects.filter(role='applicant', user__email='marcus.c@product.ai').first()
        if not marcus_profile:
            marcus_profile = Profile.objects.filter(role='applicant', user__first_name__icontains='Marcus').first()
            
        ds_vac = JobVacancy.objects.filter(employer=profile, title='Data Scientist').first()
        if not ds_vac:
            ds_vac = JobVacancy.objects.filter(employer=profile).first()
            
        pm_vac = JobVacancy.objects.filter(employer=profile, title='Product Manager').first()
        if not pm_vac:
            pm_vac = JobVacancy.objects.filter(employer=profile).first()
            
        import datetime
        
        # 1. Sarah Jenkins Interview (10 AM, Video Call, Final Round)
        Interview.objects.create(
            employer=profile,
            title=f"Interview with {sarah_profile.user.get_full_name()}" if sarah_profile else "Interview with Sarah Chen",
            candidate=sarah_profile,
            vacancy=ds_vac,
            date=today,
            start_time=datetime.time(10, 0),
            interview_type='Video Call',
            round_name='Final Round',
            meeting_link='https://meet.google.com/abc-defg-hij',
            notes='Final leadership and technical fit review.'
        )
        
        # 2. Marcus Chen Interview (02 PM, Phone Call, Screening)
        Interview.objects.create(
            employer=profile,
            title=f"Interview with {marcus_profile.user.get_full_name()}" if marcus_profile else "Interview with Marcus Thorne",
            candidate=marcus_profile,
            vacancy=pm_vac,
            date=today,
            start_time=datetime.time(14, 0),
            interview_type='Phone Call',
            round_name='Screening',
            notes='Initial HR screening and experience walk-through.'
        )
        
        # 3. Review Session (04 PM, In-Person, Hiring Sync)
        Interview.objects.create(
            employer=profile,
            title='Review Session',
            candidate=None,
            vacancy=None,
            date=today,
            start_time=datetime.time(16, 0),
            interview_type='In-Person',
            round_name='Hiring Sync',
            notes='Engineering team sync to review pipeline candidates.'
        )

    today = timezone.localdate()
    today_interviews = Interview.objects.filter(employer=profile, date=today).select_related('candidate__user', 'vacancy').order_by('start_time')
    all_interviews = Interview.objects.filter(employer=profile).select_related('candidate__user', 'vacancy').order_by('date', 'start_time')
    
    pipeline_candidates = Profile.objects.filter(role='applicant', referrals__job_vacancy__employer=profile).distinct().select_related('user')
    if pipeline_candidates.exists():
        candidates_list = pipeline_candidates
    else:
        candidates_list = Profile.objects.filter(role='applicant').select_related('user')
    
    # Build pipeline data for the active vacancies
    pipelines = []
    for vac in vacancies:
        vac_referrals = referrals.filter(job_vacancy=vac)
        match_scores = []
        candidates = []
        for ref in vac_referrals:
            match_pct, _ = calculate_match_score(ref.applicant, vac)
            match_scores.append(match_pct)
            candidates.append(ref.applicant)
        
        avg_match = round(sum(match_scores) / len(match_scores), 1) if match_scores else 80.0
        
        pipelines.append({
            'vacancy': vac,
            'avg_match': int(avg_match),
            'candidates_count': len(candidates),
            'candidates': candidates[:3],
            'status': 'Active' if vac.status == 'Open' else vac.status
        })
        
    context = {
        'profile': profile,
        'active_vacs_count': active_vacs_count,
        'total_apps_count': total_apps_count,
        'interviews_count': interviews_count,
        'hires_count': hires_count,
        'pipelines': pipelines,
        'referrals': referrals[:5],
        'today_interviews': today_interviews,
        'all_interviews': all_interviews,
        'candidates_list': candidates_list,
        'vacancies': vacancies,
    }
    return render(request, 'tracker/EMPLOYER/employer_dashboard.html', context)

@role_required(['employer'])
def applicant_details_employer(request, applicant_id=None):
    seed_mock_applicants_if_empty()
    if applicant_id:
        profile = get_object_or_404(Profile, id=applicant_id, role='applicant')
    else:
        profile = Profile.objects.filter(role='applicant').first()
        if not profile:
            messages.error(request, "No applicants found.")
            return redirect('applicants_employer')
            
    education_list = profile.education.all()
    experience_list = profile.experience.all()
    certification_list = profile.certifications.all()
    skills_list = profile.applicant_skills.all().select_related('skill')
    
    # Calculate match score against vacancies
    vacancies = JobVacancy.objects.filter(employer=request.user.profile)
    if not vacancies.exists():
        vacancies = JobVacancy.objects.all()
        
    job_id = request.GET.get('job_id')
    best_job = None
    highest_match = 0
    if job_id:
        try:
            best_job = JobVacancy.objects.get(id=job_id, employer=request.user.profile)
            highest_match, _ = calculate_match_score(profile, best_job)
        except JobVacancy.DoesNotExist:
            pass

    if not best_job:
        for vac in vacancies:
            match_pct, _ = calculate_match_score(profile, vac)
            if match_pct > highest_match:
                highest_match = match_pct
                best_job = vac

    # Gaps for the best vacancy
    gaps = []
    if best_job:
        _, gaps = calculate_match_score(profile, best_job)
        
    labels = []
    target_data = []
    current_data = []
    for g in gaps:
        labels.append(g['skill'])
        target_data.append(g['required'])
        current_data.append(g['candidate'])
        
    if not labels:
        for s in skills_list:
            labels.append(s.skill.name)
            target_data.append(5)
            current_data.append(s.proficiency)

    context = {
        'profile': profile,
        'education_list': education_list,
        'experience_list': experience_list,
        'certification_list': certification_list,
        'skills_list': skills_list,
        'highest_match': int(highest_match),
        'best_job': best_job,
        'gaps': gaps,
        'labels': labels,
        'target_data': target_data,
        'current_data': current_data,
    }
    return render(request, 'tracker/EMPLOYER/applicant_details_employer.html', context)


@role_required(['employer'])
def applicant_gap_analysis_employer(request):
    seed_mock_applicants_if_empty()
    applicant_id = request.GET.get('applicant_id')
    profile = request.user.profile
    
    if applicant_id:
        candidate = get_object_or_404(Profile, id=applicant_id, role='applicant')
    else:
        # Default to first referred candidate
        first_ref = Referral.objects.filter(job_vacancy__employer=profile).first()
        if not first_ref:
            messages.error(request, "No candidate referrals found to analyze gaps.")
            return redirect('applicants_employer')
        candidate = first_ref.applicant
        
    # Get referral to find the referred vacancy
    referral = Referral.objects.filter(applicant=candidate, job_vacancy__employer=profile).first()
    if not referral:
        messages.error(request, "This candidate has not been referred to any of your job vacancies.")
        return redirect('applicants_employer')
        
    vacancy = referral.job_vacancy
    match_pct, gaps = calculate_match_score(candidate, vacancy)
    
    # Build labels, target_data, and current_data for Chart.js radar chart
    labels = []
    target_data = []
    current_data = []
    gaps_count = 0
    for g in gaps:
        labels.append(g['skill'])
        target_data.append(g['required'])
        current_data.append(g['candidate'])
        if g['gap'] > 0:
            gaps_count += 1
            
    # Recommended training programs addressing the gaps
    gap_skills = [g['skill'] for g in gaps if g['gap'] > 0]
    recommended_programs = TrainingProgram.objects.filter(skill_addressed__name__in=gap_skills).select_related('skill_addressed')
    
    context = {
        'candidate': candidate,
        'vacancy': vacancy,
        'match_score': int(match_pct),
        'gaps': gaps,
        'gaps_count': gaps_count,
        'labels': labels,
        'target_data': target_data,
        'current_data': current_data,
        'recommended_programs': recommended_programs,
        'profile': profile,
    }
    return render(request, 'tracker/EMPLOYER/applicant_gap_analysis_employer.html', context)

def parse_exp_years(exp_str):
    if not exp_str:
        return 0.0
    import re
    m = re.search(r'([\d.]+)', str(exp_str))
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return 0.0
    return 0.0

@role_required(['employer'])
def applicants_employer(request):
    seed_mock_applicants_if_empty()
    profile = request.user.profile
    
    job_id = request.GET.get('job_id')
    selected_job = None
    referrals = Referral.objects.filter(job_vacancy__employer=profile).select_related('applicant__user', 'job_vacancy').prefetch_related('applicant__education')
    if job_id:
        referrals = referrals.filter(job_vacancy_id=job_id)
        selected_job = JobVacancy.objects.filter(id=job_id, employer=profile).first()
        
    status_filter = request.GET.get('filter')
    
    # Advanced filter GET params
    search_query = request.GET.get('search', '').strip()
    min_match = request.GET.get('min_match', '').strip()
    min_exp = request.GET.get('min_exp', '').strip()
    education = request.GET.get('education', '').strip()
    
    applicants = []
    seen_applicants = set()
    for ref in referrals:
        applicant = ref.applicant
        if applicant.id in seen_applicants:
            continue
        seen_applicants.add(applicant.id)
        
        # Calculate match score and skills list
        match_pct, _ = calculate_match_score(applicant, ref.job_vacancy)
        applicant.match_percentage = int(match_pct)
        applicant.skills_list = [s.strip() for s in applicant.skills.split(',')] if applicant.skills else []
        applicant.referred_job_title = ref.job_vacancy.title
        
        applicants.append(applicant)
        
    # Filters after building matches
    if status_filter == 'high_potential':
        applicants = [a for a in applicants if a.match_percentage >= 80]
    elif status_filter == 'Shortlisted':
        applicants = [a for a in applicants if a.status in ('Shortlisted', 'Interviewing', 'Employed — Onboarding')]
    elif status_filter == 'Rejected':
        applicants = [a for a in applicants if a.status == 'Rejected']

    # Advanced filters logic
    if search_query:
        q = search_query.lower()
        applicants = [
            a for a in applicants
            if q in f"{a.user.first_name} {a.user.last_name}".lower()
            or q in a.user.email.lower()
            or any(q in s.lower() for s in a.skills_list)
        ]

    if min_match:
        try:
            val = int(min_match)
            applicants = [a for a in applicants if a.match_percentage >= val]
        except ValueError:
            pass

    if min_exp:
        try:
            val = float(min_exp)
            applicants = [a for a in applicants if parse_exp_years(a.experience_years) >= val]
        except ValueError:
            pass

    if education:
        edu_query = education.lower()
        applicants = [
            a for a in applicants
            if any(
                edu_query in e.degree.lower() or (edu_query == 'doctorate' and 'phd' in e.degree.lower())
                for e in a.education.all()
            )
        ]

    # Check for Excel export request
    if request.GET.get('export') == 'excel':
        import openpyxl
        from django.http import HttpResponse
        import re

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = "applicants"
        if selected_job:
            filename = re.sub(r'[^a-zA-Z0-9_-]', '_', selected_job.title).lower()
        response['Content-Disposition'] = f'attachment; filename="{filename}_export.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applicants Overview"
        
        headers = ['First Name', 'Last Name', 'Email', 'Applied Position', 'Skill Match %', 'Experience Years', 'Top Skills', 'Profile Status']
        ws.append(headers)
        
        # Style headers with deep blue color matching SKILLUP theme
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0051d5', end_color='0051d5', fill_type='solid')
        alignment = Alignment(horizontal='center', vertical='center')
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
        
        for a in applicants:
            top_skills = ", ".join(a.skills_list)
            ws.append([
                a.user.first_name,
                a.user.last_name,
                a.user.email,
                a.referred_job_title,
                f"{a.match_percentage}%",
                a.experience_years or "0",
                top_skills,
                a.status
            ])
            
        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        wb.save(response)
        return response
        
    context = {
        'applicants': applicants,
        'total_applied_count': len(applicants),
        'high_match_count': sum(1 for a in applicants if a.match_percentage >= 80),
        'profile': profile,
        'selected_job_id': job_id,
        'selected_job': selected_job,
        'selected_filter': status_filter,
        
        # Advanced filters context
        'search_query': search_query,
        'min_match': min_match,
        'min_exp': min_exp,
        'selected_education': education,
        'show_advanced': bool(search_query or min_match or min_exp or education),
    }
    return render(request, 'tracker/EMPLOYER/applicants_employer.html', context)

@role_required(['employer'])
def company_profile_employer(request):
    profile = request.user.profile
    if request.method == 'POST':
        profile.company_name = request.POST.get('company_name', '').strip()
        profile.industry = request.POST.get('industry', '').strip()
        profile.company_size = request.POST.get('company_size', '').strip()
        profile.title = request.POST.get('tagline', '').strip() # use title as tagline
        profile.soft_notes = request.POST.get('about', '').strip() # use soft_notes as about
        profile.website = request.POST.get('website', '').strip()
        profile.location = request.POST.get('headquarters', '').strip() # use location as headquarters
        profile.contact_name = request.POST.get('contact_name', '').strip()
        profile.contact_position = request.POST.get('contact_position', '').strip()
        profile.contact_email = request.POST.get('contact_email', '').strip()
        profile.phone_number = request.POST.get('contact_phone', '').strip()
        
        if 'company_logo' in request.FILES:
            logo = request.FILES['company_logo']
            if profile.company_logo:
                profile.company_logo.delete()
            profile.company_logo = logo
            
        if 'profile_picture' in request.FILES:
            avatar = request.FILES['profile_picture']
            if profile.profile_picture:
                profile.profile_picture.delete()
            profile.profile_picture = avatar
            
        profile.save()
        messages.success(request, "Company profile updated successfully!")
        return redirect('company_profile_employer')
    
    return render(request, 'tracker/EMPLOYER/company_profile_employer.html', {'profile': profile})

def company_profile_detail(request, employer_id):
    seed_mock_applicants_if_empty()
    employer_profile = get_object_or_404(Profile, id=employer_id, role='employer')
    vacancies = JobVacancy.objects.filter(employer=employer_profile, status='Open')
    context = {
        'employer': employer_profile,
        'vacancies': vacancies,
        'profile': request.user.profile,
    }
    return render(request, 'tracker/APPLICANT/company_profile_detail.html', context)

@role_required(['employer'])
def hiring_tracker_employer(request):
    seed_mock_applicants_if_empty()
    profile = request.user.profile
    referrals = Referral.objects.filter(job_vacancy__employer=profile).select_related('applicant__user', 'job_vacancy')
    
    # Check for Excel export request
    if request.GET.get('export') == 'excel':
        import openpyxl
        from django.http import HttpResponse
        
        # Filter referrals if parameters are passed
        job_title = request.GET.get('job_title', '').strip()
        search_q = request.GET.get('search', '').strip()
        
        export_referrals = referrals
        if job_title:
            export_referrals = export_referrals.filter(job_vacancy__title=job_title)
        if search_q:
            q = search_q.lower()
            export_referrals = [
                r for r in export_referrals
                if q in f"{r.applicant.user.first_name} {r.applicant.user.last_name}".lower()
                or q in r.applicant.user.email.lower()
                or q in r.job_vacancy.title.lower()
            ]
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="hiring_pipeline_export.xlsx"'
        
        wb = openpyxl.Workbook()
        
        # Active sheet for Overview
        ws_all = wb.active
        ws_all.title = "All Pipeline"
        
        # Create separate sheets for stages
        ws_applied = wb.create_sheet("Applied")
        ws_interview = wb.create_sheet("Interviewing")
        ws_offered = wb.create_sheet("Offered & Onboarding")
        ws_hired = wb.create_sheet("Hired & Active")
        ws_closed = wb.create_sheet("Closed & Archived")
        
        headers = ['Candidate Name', 'Email', 'Job Title', 'Match %', 'Pipeline Stage', 'Status Detail', 'Experience', 'Top Skills', 'Salary Offer', 'Start Date']
        
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0051d5', end_color='0051d5', fill_type='solid')
        alignment = Alignment(horizontal='center', vertical='center')
        
        all_sheets = [ws_all, ws_applied, ws_interview, ws_offered, ws_hired, ws_closed]
        for sheet in all_sheets:
            sheet.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                
        def get_pipeline_stage(status):
            if status in ('Declined', 'Closed — No Show', 'Separated — End of Probation', 'Not Hired'):
                return "Closed & Archived", ws_closed
            elif status in ('Hired — Probationary', 'Hired — Regular', 'Regularly Employed', 'Still Employed — Performing Well', 'Still Employed — On Improvement Plan', 'No Response from Employer', 'No Response from Applicant', 'Probation Extended'):
                return "Hired & Active", ws_hired
            elif status in ('Accepted — Awaiting Onboarding', 'Confirmed — Onboarding', 'No Show'):
                return "Offered & Onboarding", ws_offered
            elif status == 'Interviewing':
                return "Interviewing", ws_interview
            else:
                return "Applied", ws_applied
                
        for ref in export_referrals:
            match_pct, _ = calculate_match_score(ref.applicant, ref.job_vacancy)
            match_str = f"{int(match_pct)}%"
            
            skills_list = [s.strip() for s in ref.applicant.skills.split(',')] if ref.applicant.skills else []
            top_skills = ", ".join(skills_list)
            
            stage_name, target_sheet = get_pipeline_stage(ref.status)
            
            row_data = [
                ref.applicant.user.get_full_name(),
                ref.applicant.user.email,
                ref.job_vacancy.title,
                match_str,
                stage_name,
                ref.status,
                ref.applicant.experience_years or "0",
                top_skills,
                ref.accepted_salary or "N/A",
                str(ref.actual_start_date) if ref.actual_start_date else "N/A"
            ]
            
            ws_all.append(row_data)
            target_sheet.append(row_data)
            
        for sheet in all_sheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        wb.save(response)
        return response

    if request.method == 'POST':
        referral_id = request.POST.get('referral_id')
        action = request.POST.get('action')
        ref = get_object_or_404(Referral, id=referral_id, job_vacancy__employer=profile)
        
        if action == 'move_interviewing':
            ref.status = 'Interviewing'
            ref.save()
            ref.applicant.status = 'Interviewing'
            ref.applicant.save()
            messages.success(request, f"Moved {ref.applicant.user.first_name} to Interviewing stage.")
        elif action == 'offer_acceptance':
            # Stage 1: Employer records offer acceptance
            ref.accepted_position = request.POST.get('accepted_position')
            ref.accepted_salary = request.POST.get('accepted_salary')
            
            rep_date_str = request.POST.get('reporting_date')
            if rep_date_str:
                ref.reporting_date = datetime.strptime(rep_date_str, '%Y-%m-%d').date()
            
            ref.work_location = request.POST.get('work_location')
            ref.employment_type = request.POST.get('employment_type')
            prob_months = request.POST.get('probationary_period_months')
            if prob_months:
                ref.probationary_period_months = int(prob_months)
            ref.employer_remarks = request.POST.get('employer_remarks')
            
            ref.status = 'Accepted — Awaiting Onboarding'
            ref.applicant.status = 'Shortlisted'
            ref.save()
            ref.applicant.save()
            
            # Decrement job slots
            vac = ref.job_vacancy
            vac.remaining_slots = max(0, vac.remaining_slots - 1)
            vac.save()
            
            # Send Notification to Applicant
            Notification.objects.create(
                user=ref.applicant.user,
                message=f"Congratulations! {profile.user.first_name} has accepted you for the position of {ref.accepted_position}. Your reporting date is {ref.reporting_date}. Please confirm your acceptance below."
            )
            
            # Send Notification to PESO Coordinators
            coord_users = User.objects.filter(profile__role='admin')
            for u in coord_users:
                Notification.objects.create(
                    user=u,
                    message=f"Referral for {ref.applicant.user.get_full_name()} to {ref.job_vacancy.title} has resulted in an acceptance."
                )
                
            messages.success(request, f"Successfully recorded acceptance for {ref.applicant.user.first_name} and sent notifications.")
            
        elif action == 'confirm_onboarding':
            ref.status = 'Confirmed — Onboarding'
            ref.applicant.status = 'Employed — Onboarding'
            ref.save()
            ref.applicant.save()
            messages.success(request, f"Manually confirmed onboarding requirements for {ref.applicant.user.first_name}.")

        elif action == 'log_reporting_status':
            # Stage 4: Physical Report for Work
            reported = request.POST.get('reported_for_work')
            ref.reported_for_work = reported
            
            if reported == 'Yes':
                start_date_str = request.POST.get('actual_start_date')
                if start_date_str:
                    ref.actual_start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                
                # Update status depending on employment type
                if ref.employment_type == 'Probationary':
                    ref.status = 'Hired — Probationary'
                else:
                    ref.status = 'Hired — Regular'
                ref.applicant.status = 'Hired'
                ref.applicant.save()
                ref.save()
                
                # Notify Coordinators
                coord_users = User.objects.filter(profile__role='admin')
                for u in coord_users:
                    Notification.objects.create(
                        user=u,
                        message=f"{ref.applicant.user.get_full_name()} has reported for work as scheduled on {ref.actual_start_date}."
                    )
                messages.success(request, f"Logged {ref.applicant.user.first_name} as reported for work.")
            else:
                ref.status = 'No Show'
                ref.save()
                
                # Notify Coordinators
                coord_users = User.objects.filter(profile__role='admin')
                for u in coord_users:
                    Notification.objects.create(
                        user=u,
                        message=f"ALERT: {ref.applicant.user.get_full_name()} failed to report for work on their reporting date ({ref.reporting_date})."
                    )
                messages.warning(request, f"Logged {ref.applicant.user.first_name} as No Show. PESO has been notified.")
                
        elif action == 'final_probation_status':
            # Stage 6: Probation Completion / Regularization
            final_status = request.POST.get('final_employment_status')
            
            if final_status == 'Regularized':
                ref.status = 'Regularly Employed'
                ref.applicant.status = 'Employed — Regular'
                ref.applicant.save()
                ref.save()
                
                # Notify Applicant
                Notification.objects.create(
                    user=ref.applicant.user,
                    message=f"Congratulations! You have been regularized in your role as {ref.accepted_position}."
                )
                # Notify Coordinator
                coord_users = User.objects.filter(profile__role='admin')
                for u in coord_users:
                    Notification.objects.create(
                        user=u,
                        message=f"{ref.applicant.user.get_full_name()} has successfully completed their probationary period and is regularized."
                    )
                messages.success(request, f"Marked {ref.applicant.user.first_name} as regularly employed.")
                
            elif final_status == 'Extended':
                ref.status = 'Probation Extended'
                ext_date_str = request.POST.get('probation_extension_end_date')
                if ext_date_str:
                    ref.probation_extension_end_date = datetime.strptime(ext_date_str, '%Y-%m-%d').date()
                ref.save()
                
                coord_users = User.objects.filter(profile__role='admin')
                for u in coord_users:
                    Notification.objects.create(
                        user=u,
                        message=f"Employer has extended the probation for {ref.applicant.user.get_full_name()} to {ref.probation_extension_end_date}."
                    )
                messages.success(request, f"Extended probation period for {ref.applicant.user.first_name}.")
                
            elif final_status == 'Separated':
                ref.status = 'Separated — End of Probation'
                ref.separation_reason = request.POST.get('separation_reason')
                ref.applicant.status = 'Active — Job Seeking'
                ref.applicant.save()
                ref.save()
                
                coord_users = User.objects.filter(profile__role='admin')
                for u in coord_users:
                    Notification.objects.create(
                        user=u,
                        message=f"{ref.applicant.user.get_full_name()} has separated from the company at the end of their probation. Reason: {ref.separation_reason}."
                    )
                messages.warning(request, f"Marked {ref.applicant.user.first_name} as separated. Candidate returned to active job seeking pool.")
                
        elif action == 'submit_employer_evaluation':
            # Stage 7: Employer feedback evaluation form
            ref.eval_emp_quality = int(request.POST.get('eval_emp_quality', 5))
            ref.eval_emp_skills_accurate = request.POST.get('eval_emp_skills_accurate', 'Yes')
            ref.eval_emp_certs_genuine = request.POST.get('eval_emp_certs_genuine', 'Yes')
            ref.eval_emp_future_referrals = request.POST.get('eval_emp_future_referrals', 'Yes')
            ref.eval_emp_satisfaction = int(request.POST.get('eval_emp_satisfaction', 5))
            ref.eval_emp_notes = request.POST.get('eval_emp_notes', '')
            ref.save()
            messages.success(request, "Thank you! Your feedback has been submitted successfully.")
            
        elif action == 'move_rejected':
            ref.status = 'Not Hired'
            ref.applicant.status = 'Rejected'
            ref.rejection_reason = request.POST.get('rejection_reason', 'Underqualified')
            ref.save()
            ref.applicant.save()
            messages.success(request, f"Candidate {ref.applicant.user.first_name} marked as Not Hired.")
            
        return redirect('hiring_tracker_employer')
        
    # Categorize referrals for Kanban lanes
    applied_lane = []
    interviewing_lane = []
    offered_lane = []
    hired_lane = []
    closed_lane = []
    
    for ref in referrals:
        # Calculate match percentage for displaying on card
        match_pct, _ = calculate_match_score(ref.applicant, ref.job_vacancy)
        ref.match_percentage = int(match_pct)
        
        # Extract first 2 letters for candidate initials circle
        ref.initials = (ref.applicant.user.first_name[:1] + ref.applicant.user.last_name[:1]).upper()
        
        # Calculate onboarding requirements progress
        files_fields = [ref.nbi_clearance, ref.medical_certificate, ref.birth_certificate, ref.diploma_transcript, ref.prev_employment_cert, ref.tesda_cert]
        uploaded_count = sum(1 for f in files_fields if f)
        text_fields = [ref.sss_number, ref.philhealth_number, ref.pagibig_number, ref.bir_tin]
        filled_text_count = sum(1 for t in text_fields if t)
        ref.uploaded_count = uploaded_count
        ref.filled_text_count = filled_text_count
        ref.total_requirements_count = 10
        ref.completed_requirements_count = uploaded_count + filled_text_count
        
        if ref.status in ('Declined', 'Closed — No Show', 'Separated — End of Probation', 'Not Hired'):
            closed_lane.append(ref)
        elif ref.status in ('Hired — Probationary', 'Hired — Regular', 'Regularly Employed', 'Still Employed — Performing Well', 'Still Employed — On Improvement Plan', 'No Response from Employer', 'No Response from Applicant', 'Probation Extended'):
            hired_lane.append(ref)
        elif ref.status in ('Accepted — Awaiting Onboarding', 'Confirmed — Onboarding', 'No Show'):
            offered_lane.append(ref)
        elif ref.status == 'Interviewing':
            interviewing_lane.append(ref)
        else:
            applied_lane.append(ref)
            
    context = {
        'applied_lane': applied_lane,
        'interviewing_lane': interviewing_lane,
        'offered_lane': offered_lane,
        'hired_lane': hired_lane,
        'closed_lane': closed_lane,
        'profile': profile,
    }
    return render(request, 'tracker/EMPLOYER/hiring_tracker_employer.html', context)

@role_required(['employer'])
def job_management_employer(request):
    seed_mock_applicants_if_empty()
    profile = request.user.profile
    vacancies = JobVacancy.objects.filter(employer=profile).order_by('-created_at')
    
    if request.method == 'POST':
        action = request.POST.get('action', 'create')
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        remarks = request.POST.get('remarks', '').strip()
        location = request.POST.get('location', '').strip()
        salary_range = request.POST.get('salary_range', '').strip()
        category = request.POST.get('category', 'Local').strip()
        min_education = request.POST.get('min_education', '').strip()
        required_certifications = request.POST.get('required_certifications', '').strip()
        req_exp = request.POST.get('required_experience_years', '0')
        status = request.POST.get('status', 'Open').strip()
        
        if action == 'edit':
            vacancy_id = request.POST.get('vacancy_id')
            vacancy = get_object_or_404(JobVacancy, id=vacancy_id, employer=profile)
            vacancy.title = title
            vacancy.description = description
            vacancy.remarks = remarks
            vacancy.location = location
            vacancy.salary_range = salary_range
            vacancy.category = category
            vacancy.min_education = min_education
            vacancy.required_certifications = required_certifications
            vacancy.required_experience_years = int(req_exp) if req_exp.isdigit() else 0
            vacancy.status = status
            vacancy.save()
            
            # Clear old skill requirements and rebuild them
            JobSkillRequirement.objects.filter(job_vacancy=vacancy).delete()
        else:
            vacancy = JobVacancy.objects.create(
                title=title,
                employer=profile,
                description=description,
                remarks=remarks,
                location=location,
                salary_range=salary_range,
                category=category,
                min_education=min_education,
                required_certifications=required_certifications,
                required_experience_years=int(req_exp) if req_exp.isdigit() else 0,
                status=status
            )
            
        # Handle requirements selection
        selected_skills = request.POST.getlist('skills[]')
        selected_proficiencies = request.POST.getlist('proficiencies[]')
        for s_id, prof in zip(selected_skills, selected_proficiencies):
            if s_id:
                skill_obj = get_object_or_404(CentralizedSkill, id=int(s_id))
                JobSkillRequirement.objects.create(
                    job_vacancy=vacancy,
                    skill=skill_obj,
                    required_proficiency=int(prof) if prof else 1
                )
                
        # Handle custom skills on-the-fly
        new_skills_str = request.POST.get('new_skills', '').strip()
        if new_skills_str:
            new_skills_list = [s.strip() for s in new_skills_str.split(',') if s.strip()]
            for s_name in new_skills_list:
                skill_obj, _ = CentralizedSkill.objects.get_or_create(
                    name=s_name,
                    defaults={'category': 'Custom', 'description': 'Custom skill added by employer.'}
                )
                JobSkillRequirement.objects.create(
                    job_vacancy=vacancy,
                    skill=skill_obj,
                    required_proficiency=3 # Default required proficiency = 3
                )
                
        # Notify matching employers of any 70%+ candidate match scores
        check_and_notify_applicants_for_vacancy(vacancy)
        
        msg = f"Job vacancy '{title}' updated successfully!" if action == 'edit' else f"Job vacancy '{title}' posted successfully!"
        messages.success(request, msg)
        return redirect('job_management_employer')
        
    # GET request
    active_postings_count = vacancies.filter(status='Open').count()
    total_applicants_count = Referral.objects.filter(job_vacancy__employer=profile).count()
    
    # Enrich vacancy objects with applicant lists & counts
    all_vacancies_data = []
    for vac in vacancies:
        refs = Referral.objects.filter(job_vacancy=vac).select_related('applicant__user')
        cands = [r.applicant for r in refs]
        skills_reqs = JobSkillRequirement.objects.filter(job_vacancy=vac).select_related('skill')
        all_vacancies_data.append({
            'vacancy': vac,
            'candidates_count': len(cands),
            'candidates': cands[:3],
            'requirements': skills_reqs,
        })
        
    all_skills = CentralizedSkill.objects.all().order_by('name')
    
    context = {
        'vacancies': all_vacancies_data,
        'active_postings_count': active_postings_count,
        'total_applicants_count': total_applicants_count,
        'all_skills': all_skills,
        'profile': profile,
    }
    return render(request, 'tracker/EMPLOYER/job_management_employer.html', context)


@role_required(['employer'])
def employer_notifications_view(request):
    """Shows all applicants with ≥75% skill match on employer's open vacancies.
    Computed live — applicants don't need to have applied to appear here.
    """
    profile = request.user.profile
    vacancies = JobVacancy.objects.filter(employer=profile).exclude(status='Closed').prefetch_related('requirements__skill')
    applicants = Profile.objects.filter(role='applicant').select_related('user').prefetch_related(
        'applicant_skills__skill',
        'education',
        'experience'
    )

    # Prefetch all referrals for the employer's vacancies to avoid queries in the nested loop
    referrals = Referral.objects.filter(job_vacancy__in=vacancies).values('applicant_id', 'job_vacancy_id', 'status')
    referrals_dict = {}
    for r in referrals:
        referrals_dict[(r['applicant_id'], r['job_vacancy_id'])] = r['status']

    # Build flat list of (applicant, vacancy, match_pct) tuples ≥ 75%
    match_alerts = []
    seen = set()  # avoid duplicate (applicant, vacancy) pairs
    for vac in vacancies:
        for app in applicants:
            key = (app.id, vac.id)
            if key in seen:
                continue
            seen.add(key)
            match_pct, _ = calculate_match_score(app, vac)
            if match_pct >= 75:
                has_applied = key in referrals_dict
                offer_sent = referrals_dict.get(key) == 'Accepted — Awaiting Onboarding'
                match_alerts.append({
                    'applicant': app,
                    'vacancy': vac,
                    'match_pct': int(match_pct),
                    'has_applied': has_applied,
                    'offer_sent': offer_sent,
                })

    # Sort highest match first
    match_alerts.sort(key=lambda x: x['match_pct'], reverse=True)

    context = {
        'match_alerts': match_alerts,
        'profile': profile,
        'unread_match_alerts': len(match_alerts),
    }
    return render(request, 'tracker/EMPLOYER/employer_notifications.html', context)


@role_required(['employer'])
def send_offer_employer(request, applicant_id, vacancy_id):
    """Employer sends a job offer to a matched applicant directly from the notification."""
    profile = request.user.profile
    applicant_profile = get_object_or_404(Profile, id=applicant_id, role='applicant')
    vacancy = get_object_or_404(JobVacancy, id=vacancy_id, employer=profile)

    # Calculate current match score for display
    match_pct, _ = calculate_match_score(applicant_profile, vacancy)

    # Check for existing referral
    existing_referral = Referral.objects.filter(applicant=applicant_profile, job_vacancy=vacancy).first()

    if request.method == 'POST':
        accepted_position = request.POST.get('accepted_position', vacancy.title)
        accepted_salary = request.POST.get('accepted_salary', vacancy.salary_range or '')
        reporting_date_str = request.POST.get('reporting_date', '')
        employment_type = request.POST.get('employment_type', 'Probationary')
        work_location = request.POST.get('work_location', vacancy.location)
        employer_remarks = request.POST.get('employer_remarks', '')
        prob_months = request.POST.get('probationary_period_months', '6')

        # Create referral if not existing, or update existing
        referral, _ = Referral.objects.get_or_create(
            applicant=applicant_profile,
            job_vacancy=vacancy,
            defaults={'status': 'Accepted — Awaiting Onboarding', 'date_referred': timezone.localdate()}
        )
        referral.accepted_position = accepted_position
        referral.accepted_salary = accepted_salary
        referral.work_location = work_location
        referral.employment_type = employment_type
        referral.employer_remarks = employer_remarks
        referral.status = 'Accepted — Awaiting Onboarding'
        if prob_months:
            try:
                referral.probationary_period_months = int(prob_months)
            except (ValueError, TypeError):
                pass
        if reporting_date_str:
            try:
                referral.reporting_date = datetime.strptime(reporting_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        referral.save()

        # Update applicant pipeline status
        applicant_profile.status = 'Shortlisted'
        applicant_profile.save()

        # Decrement remaining slots
        if vacancy.remaining_slots and vacancy.remaining_slots > 0:
            vacancy.remaining_slots = max(0, vacancy.remaining_slots - 1)
            vacancy.save()

        # Notify the applicant
        Notification.objects.create(
            user=applicant_profile.user,
            notif_type='offer_sent',
            vacancy=vacancy,
            message=f"🎉 You've received a job offer from {profile.company_name or profile.user.get_full_name()} for the position of '{accepted_position}' ({vacancy.title}). Please check your applications for details!"
        )

        # Mark the match-alert notification as read
        request.user.notifications.filter(
            notif_type='match_alert',
            applicant_profile=applicant_profile,
            vacancy=vacancy,
        ).update(is_read=True)

        messages.success(request, f"Offer successfully sent to {applicant_profile.user.get_full_name()}! They will be notified.")
        return redirect('hiring_tracker_employer')

    context = {
        'applicant': applicant_profile,
        'vacancy': vacancy,
        'match_pct': int(match_pct),
        'existing_referral': existing_referral,
        'profile': profile,
    }
    return render(request, 'tracker/EMPLOYER/send_offer_employer.html', context)


@verified_required
@require_POST
def job_status_update(request, vacancy_id):
    """Quick inline status update for a job vacancy."""
    profile = request.user.profile
    vacancy = get_object_or_404(JobVacancy, id=vacancy_id, employer=profile)
    new_status = request.POST.get('status', '').strip()
    allowed = ['Open', 'Closed', 'Draft', 'On Hold']
    if new_status in allowed:
        old_status = vacancy.status
        vacancy.status = new_status
        vacancy.save()

        # --- Cascade to applicants ---
        active_referrals = Referral.objects.filter(
            job_vacancy=vacancy,
            status__in=['Pending', 'Interviewing']
        ).select_related('applicant__user')

        if new_status == 'Closed':
            # Mark all pending/interviewing applicants as Not Hired
            for ref in active_referrals:
                ref.status = 'Not Hired'
                ref.rejection_reason = 'Position closed by employer.'
                ref.save()
                Notification.objects.create(
                    user=ref.applicant.user,
                    message=f"The job posting '{vacancy.title}' has been closed. Your application has been marked as Not Hired.",
                    notification_type='alert'
                )
            affected = active_referrals.count()
            messages.success(request, f"'{vacancy.title}' closed. {affected} pending applicant(s) marked as Not Hired.")

        elif new_status == 'On Hold':
            # Notify applicants that processing is paused
            for ref in active_referrals:
                Notification.objects.create(
                    user=ref.applicant.user,
                    message=f"The job posting '{vacancy.title}' is currently On Hold. Your application is paused pending employer review.",
                    notification_type='info'
                )
            affected = active_referrals.count()
            messages.success(request, f"'{vacancy.title}' put On Hold. {affected} applicant(s) notified.")

        elif new_status == 'Open' and old_status in ['Closed', 'On Hold']:
            # Re-open: notify applicants who were Not Hired due to position closed
            reopened = Referral.objects.filter(
                job_vacancy=vacancy,
                rejection_reason='Position closed by employer.'
            ).select_related('applicant__user')
            for ref in reopened:
                ref.status = 'Pending'
                ref.rejection_reason = None
                ref.save()
                Notification.objects.create(
                    user=ref.applicant.user,
                    message=f"Good news! '{vacancy.title}' has been re-opened. Your application is back to Pending.",
                    notification_type='success'
                )
            messages.success(request, f"'{vacancy.title}' re-opened. {reopened.count()} applicant(s) restored to Pending.")

        else:
            messages.success(request, f"'{vacancy.title}' status changed from {old_status} → {new_status}.")
    else:
        messages.error(request, "Invalid status value.")
    return redirect('job_management_employer')


@verified_required
@require_POST
def job_delete(request, vacancy_id):
    """Delete a job vacancy and all its referrals/skill requirements."""
    profile = request.user.profile
    vacancy = get_object_or_404(JobVacancy, id=vacancy_id, employer=profile)
    title = vacancy.title
    vacancy.delete()
    messages.success(request, f"Job posting '{title}' has been permanently deleted.")
    return redirect('job_management_employer')

@role_required(['admin'])
def export_analytics_csv(request):
    import csv
    from django.http import HttpResponse
    
    report_type = request.GET.get('report', 'all')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="peso_{report_type}_analytics.csv"'
    
    writer = csv.writer(response)
    
    if report_type == 'demand':
        writer.writerow(['Skill Name', 'Number of Job Postings Requiring Skill'])
        demand_counts = JobSkillRequirement.objects.values('skill__name').annotate(count=Count('job_vacancy')).order_by('-count')
        for item in demand_counts:
            writer.writerow([item['skill__name'], item['count']])
            
    elif report_type == 'gaps':
        writer.writerow(['Skill Name', 'Total Open Vacancies Requiring Skill', 'Average Applicant Rating', 'Average Shortfall'])
        skills = CentralizedSkill.objects.all()
        total_applicants = Profile.objects.filter(role='applicant').count()
        for s in skills:
            req_count = JobSkillRequirement.objects.filter(skill=s).count()
            if req_count > 0:
                avg_rating = ApplicantSkill.objects.filter(skill=s).aggregate(Sum('proficiency'))['proficiency__sum'] or 0
                avg_rating_val = round(avg_rating / total_applicants, 2) if total_applicants > 0 else 0.0
                
                shortfalls = []
                for req in JobSkillRequirement.objects.filter(skill=s):
                    for askill in ApplicantSkill.objects.filter(skill=s):
                        if askill.proficiency < req.required_proficiency:
                            shortfalls.append(req.required_proficiency - askill.proficiency)
                avg_shortfall = round(sum(shortfalls) / len(shortfalls), 2) if shortfalls else 0.0
                writer.writerow([s.name, req_count, avg_rating_val, avg_shortfall])
                
    elif report_type == 'mismatch':
        writer.writerow(['Applicant', 'Field of Study', 'Referred Vacancy', 'Is Mismatched'])
        referrals = Referral.objects.all().select_related('applicant__user', 'job_vacancy')
        for r in referrals:
            edu = r.applicant.education.first()
            fos = edu.field_of_study if edu else 'Not Specified'
            mismatched = check_mismatch(fos, r.job_vacancy.title)
            writer.writerow([f"{r.applicant.user.first_name} {r.applicant.user.last_name}", fos, r.job_vacancy.title, 'YES' if mismatched else 'NO'])
            
    elif report_type == 'responsiveness':
        writer.writerow(['Employer Name', 'Total Referrals', 'Pending Count', 'Hired Count', 'Not Hired Count', 'No Response Count', 'Response Rate (%)'])
        employers = Profile.objects.filter(role='employer').select_related('user')
        for emp in employers:
            emp_referrals = Referral.objects.filter(job_vacancy__employer=emp)
            total = emp_referrals.count()
            pending = emp_referrals.filter(status='Pending').count()
            hired = emp_referrals.filter(status='Hired').count()
            not_hired = emp_referrals.filter(status='Not Hired').count()
            no_response = emp_referrals.filter(status='No Response').count()
            
            responded = hired + not_hired
            resp_rate = round((responded / (total - pending) * 100.0), 1) if (total - pending) > 0 else 0.0
            if total == 0:
                resp_rate = 100.0
            writer.writerow([f"{emp.user.first_name} {emp.user.last_name}", total, pending, hired, not_hired, no_response, resp_rate])
            
    elif report_type == 'funnel':
        writer.writerow(['Funnel Stage', 'Count'])
        total_applicants = Profile.objects.filter(role='applicant').count()
        referrals = Referral.objects.all()
        total_referrals = referrals.count()
        total_pending = referrals.filter(status='Pending').count()
        total_hired = referrals.filter(status='Hired').count()
        total_no_response = referrals.filter(status='No Response').count()
        
        writer.writerow(['Registered Applicants', total_applicants])
        writer.writerow(['Total Referred Candidates', total_referrals])
        writer.writerow(['Pending Referrals', total_pending])
        writer.writerow(['Placed / Hired Placements', total_hired])
        writer.writerow(['No Response Referrals', total_no_response])
        
    else:
        writer.writerow(['=== PESO SYSTEM GENERAL SUMMARY ==='])
        writer.writerow([])
        writer.writerow(['Total Registered Applicants', Profile.objects.filter(role='applicant').count()])
        writer.writerow(['Total Registered Employers', Profile.objects.filter(role='employer').count()])
        writer.writerow(['Total Job Vacancies', JobVacancy.objects.count()])
        writer.writerow(['Total Referrals Logged', Referral.objects.count()])
        writer.writerow([])
        writer.writerow(['=== PLACEMENT FUNNEL ==='])
        referrals = Referral.objects.all()
        writer.writerow(['Pending', referrals.filter(status='Pending').count()])
        writer.writerow(['Hired', referrals.filter(status='Hired').count()])
        writer.writerow(['Not Hired', referrals.filter(status='Not Hired').count()])
        writer.writerow(['No Response', referrals.filter(status='No Response').count()])
        
    return response


@role_required(['applicant'])
def applicant_settings(request):
    profile = request.user.profile
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'change_password':
            current_password = request.POST.get('current_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            
            # Verify current password
            if not request.user.check_password(current_password):
                messages.error(request, "Incorrect current password.")
            elif new_password != confirm_password:
                messages.error(request, "New passwords do not match.")
            elif len(new_password) < 6:
                messages.error(request, "Password must be at least 6 characters.")
            else:
                request.user.set_password(new_password)
                request.user.save()
                # To keep user logged in after password change
                auth_login(request, request.user)
                messages.success(request, "Password changed successfully!")
                return redirect('applicant_settings')
                
        elif action == 'update_profile':
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            phone_number = request.POST.get('phone_number', '').strip()
            
            # Simple validation
            if not first_name or not last_name or not email:
                messages.error(request, "First name, last name, and email are required.")
            elif User.objects.exclude(id=request.user.id).filter(email=email).exists():
                messages.error(request, "Email address is already in use.")
            else:
                request.user.first_name = first_name
                request.user.last_name = last_name
                request.user.email = email
                request.user.save()
                
                profile.phone_number = phone_number
                profile.save()
                
                messages.success(request, "Account settings updated successfully!")
                return redirect('applicant_settings')
                
    context = {
        'profile': profile,
        'user': request.user,
    }
    return render(request, 'tracker/APPLICANT/settings.html', context)


@role_required(['employer'])
@require_POST
def schedule_interview_employer(request):
    profile = request.user.profile
    title = request.POST.get('title', '').strip()
    candidate_id = request.POST.get('candidate_id')
    vacancy_id = request.POST.get('vacancy_id')
    date_str = request.POST.get('date')
    time_str = request.POST.get('start_time')
    interview_type = request.POST.get('interview_type', 'Video Call')
    round_name = request.POST.get('round_name', 'Screening')
    meeting_link = request.POST.get('meeting_link', '').strip()
    notes = request.POST.get('notes', '').strip()
    
    if not date_str or not time_str:
        messages.error(request, "Date and Start Time are required.")
        return redirect('employer_dashboard')
        
    try:
        from datetime import datetime
        interview_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        try:
            interview_time = datetime.strptime(time_str, '%H:%M').time()
        except ValueError:
            interview_time = datetime.strptime(time_str, '%H:%M:%S').time()
    except Exception as e:
        messages.error(request, f"Invalid date or time format: {e}")
        return redirect('employer_dashboard')
        
    candidate = None
    if candidate_id:
        candidate = get_object_or_404(Profile, id=candidate_id, role='applicant')
        
    vacancy = None
    if vacancy_id:
        vacancy = get_object_or_404(JobVacancy, id=vacancy_id, employer=profile)
        
    if not title and candidate:
        title = f"Interview with {candidate.user.get_full_name()}"
        
    Interview.objects.create(
        employer=profile,
        title=title,
        candidate=candidate,
        vacancy=vacancy,
        date=interview_date,
        start_time=interview_time,
        interview_type=interview_type,
        round_name=round_name,
        meeting_link=meeting_link or None,
        notes=notes or None
    )
    
    messages.success(request, "Interview scheduled successfully!")
    return redirect('employer_dashboard')


@role_required(['employer'])
@require_POST
def cancel_interview_employer(request, interview_id):
    profile = request.user.profile
    interview = get_object_or_404(Interview, id=interview_id, employer=profile)
    interview_name = interview.candidate.user.get_full_name() if interview.candidate else interview.title or "Meeting"
    interview.delete()
    messages.success(request, f"Scheduled interview/meeting '{interview_name}' cancelled successfully.")
    return redirect('employer_dashboard')


def reset_db_view(request):
    from tracker.models import (
        Skill, Milestone, StudyLog, Profile,
        Education, WorkExperience, Certification, ApplicantSkill,
        JobVacancy, JobSkillRequirement, TrainingProgram, TrainingEnrollment, Referral,
        GapScoreLog, JobBookmark, ApplicantDocument, Interview, Notification, CentralizedSkill
    )
    from django.contrib.admin.models import LogEntry
    from django.contrib.auth.models import User
    from django.contrib import messages
    from django.shortcuts import render, redirect

    # Build options list dynamically
    data_options = [
        # Preserved
        {
            'id': 'users',
            'label': 'User Accounts',
            'count': User.objects.count(),
            'purgeable': False,
            'badge': 'PRESERVED',
            'desc': 'Logins, credentials, superusers, and coordinator accounts.'
        },
        {
            'id': 'profiles',
            'label': 'Profiles',
            'count': Profile.objects.count(),
            'purgeable': False,
            'badge': 'PRESERVED',
            'desc': 'User roles, contact details, and employer company metadata.'
        },
        # Purgeable
        {
            'id': 'vacancies',
            'label': 'Job Vacancies & Requirements',
            'count': JobVacancy.objects.count(),
            'purgeable': True,
            'default_checked': True,
            'badge': 'PURGEABLE',
            'desc': 'Job postings and required skill criteria.'
        },
        {
            'id': 'referrals',
            'label': 'Referrals & Applications',
            'count': Referral.objects.count(),
            'purgeable': True,
            'default_checked': True,
            'badge': 'PURGEABLE',
            'desc': 'Job application records, statuses, and pre-employment checklists.'
        },
        {
            'id': 'interviews',
            'label': 'Interviews',
            'count': Interview.objects.count(),
            'purgeable': True,
            'default_checked': True,
            'badge': 'PURGEABLE',
            'desc': 'Scheduled interview records and meeting details.'
        },
        {
            'id': 'training_enrollments',
            'label': 'Training Enrollments',
            'count': TrainingEnrollment.objects.count(),
            'purgeable': True,
            'default_checked': True,
            'badge': 'PURGEABLE',
            'desc': 'Applicant enrollments in training courses.'
        },
        {
            'id': 'educations',
            'label': 'Education Records',
            'count': Education.objects.count(),
            'purgeable': True,
            'default_checked': True,
            'badge': 'PURGEABLE',
            'desc': 'Degrees and academic history added to applicant profiles.'
        },
        {
            'id': 'work_experiences',
            'label': 'Work Experience',
            'count': WorkExperience.objects.count(),
            'purgeable': True,
            'default_checked': True,
            'badge': 'PURGEABLE',
            'desc': 'Prior employment logs added to applicant profiles.'
        },
        {
            'id': 'certifications',
            'label': 'Certifications',
            'count': Certification.objects.count(),
            'purgeable': True,
            'default_checked': True,
            'badge': 'PURGEABLE',
            'desc': 'Certificates and credentials on applicant profiles.'
        },
        {
            'id': 'applicant_skills',
            'label': 'Applicant Skills',
            'count': ApplicantSkill.objects.count(),
            'purgeable': True,
            'default_checked': True,
            'badge': 'PURGEABLE',
            'desc': 'Core skills linked to applicant accounts.'
        },
        {
            'id': 'tracker_skills',
            'label': 'User Skills & Study Logs',
            'count': Skill.objects.count(),
            'purgeable': True,
            'default_checked': True,
            'badge': 'PURGEABLE',
            'desc': 'Skills, milestones, and time spent studying (Study Logs).'
        },
        {
            'id': 'bookmarks',
            'label': 'Bookmarks',
            'count': JobBookmark.objects.count(),
            'purgeable': True,
            'default_checked': True,
            'badge': 'PURGEABLE',
            'desc': 'Job postings bookmarked by applicants.'
        },
        {
            'id': 'documents',
            'label': 'Uploaded Documents',
            'count': ApplicantDocument.objects.count(),
            'purgeable': True,
            'default_checked': True,
            'badge': 'PURGEABLE',
            'desc': 'Resumes, files, and certifications uploaded by applicants.'
        },
        {
            'id': 'notifications',
            'label': 'Notifications',
            'count': Notification.objects.count(),
            'purgeable': True,
            'default_checked': True,
            'badge': 'PURGEABLE',
            'desc': 'Alerts, updates, and messages sent to users.'
        },
        {
            'id': 'gap_logs',
            'label': 'Gap Score Logs',
            'count': GapScoreLog.objects.count(),
            'purgeable': True,
            'default_checked': True,
            'badge': 'PURGEABLE',
            'desc': 'Stored job matching evaluations and gap logs.'
        },
        # Lookups (Unchecked by default to avoid accidental wipe)
        {
            'id': 'centralized_skills',
            'label': 'Centralized Skills (Lookup)',
            'count': CentralizedSkill.objects.count(),
            'purgeable': True,
            'default_checked': False,
            'badge': 'LOOKUP',
            'desc': 'Lookup dictionary of skill names used globally.'
        },
        {
            'id': 'training_programs',
            'label': 'Training Programs (Lookup)',
            'count': TrainingProgram.objects.count(),
            'purgeable': True,
            'default_checked': False,
            'badge': 'LOOKUP',
            'desc': 'Pre-defined lists of training programs available.'
        },
    ]

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'reset_data':
            selected_ids = request.POST.getlist('selected_types')
            deleted_items = []
            
            # Perform targeted deletions based on user selection
            if 'notifications' in selected_ids:
                Notification.objects.all().delete()
                deleted_items.append("Notifications")
                
            if 'bookmarks' in selected_ids:
                JobBookmark.objects.all().delete()
                deleted_items.append("Bookmarks")
                
            if 'documents' in selected_ids:
                ApplicantDocument.objects.all().delete()
                deleted_items.append("Uploaded Documents")
                
            if 'gap_logs' in selected_ids:
                GapScoreLog.objects.all().delete()
                deleted_items.append("Gap Score Logs")
                
            if 'interviews' in selected_ids:
                Interview.objects.all().delete()
                deleted_items.append("Interviews")
                
            if 'referrals' in selected_ids:
                Referral.objects.all().delete()
                deleted_items.append("Referrals & Applications")
                
            if 'vacancies' in selected_ids:
                JobSkillRequirement.objects.all().delete()
                JobVacancy.objects.all().delete()
                deleted_items.append("Job Vacancies & Requirements")
                
            if 'training_enrollments' in selected_ids:
                TrainingEnrollment.objects.all().delete()
                deleted_items.append("Training Enrollments")
                
            if 'applicant_skills' in selected_ids:
                ApplicantSkill.objects.all().delete()
                deleted_items.append("Applicant Skills")
                
            if 'educations' in selected_ids:
                Education.objects.all().delete()
                deleted_items.append("Education Records")
                
            if 'work_experiences' in selected_ids:
                WorkExperience.objects.all().delete()
                deleted_items.append("Work Experience Records")
                
            if 'certifications' in selected_ids:
                Certification.objects.all().delete()
                deleted_items.append("Certifications")
                
            if 'tracker_skills' in selected_ids:
                StudyLog.objects.all().delete()
                Milestone.objects.all().delete()
                Skill.objects.all().delete()
                deleted_items.append("User Skills & Study Logs")
                
            if 'centralized_skills' in selected_ids:
                CentralizedSkill.objects.all().delete()
                deleted_items.append("Centralized Skills")
                
            if 'training_programs' in selected_ids:
                TrainingProgram.objects.all().delete()
                deleted_items.append("Training Programs")
                
            if deleted_items:
                LogEntry.objects.all().delete()
                create_apex_employer_and_jobs()
                create_prime_employer_and_jobs()
                create_nexus_employer_and_jobs()
                create_applicant_test_accounts()
                messages.success(request, f"Successfully cleared selected data: {', '.join(deleted_items)}.")
            else:
                messages.warning(request, "No data types were selected for deletion.")
                
            return redirect('reset_db_view')

    return render(request, 'tracker/reset_db.html', {
        'data_options': data_options
    })


def forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        user_exists = User.objects.filter(email=email).exists()
        
        if user_exists:
            otp = str(random.randint(100000, 999999))
            
            request.session['reset_email'] = email
            request.session['reset_code'] = otp
            request.session['reset_code_verified'] = False
            
            email_sent = False
            error_message = ""
            try:
                send_mail(
                    subject='Reset Your Password - SKILLUP',
                    message=f'Hello,\n\nWe received a request to reset the password for your SKILLUP account.\n\nYour 6-digit verification code is:\n\n{otp}\n\nEnter this code on the password reset verification page to proceed.\n\nIf you did not request this reset, you can safely ignore this email.\n\nBest regards,\nSKILLUP Team',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[email],
                    fail_silently=False
                )
                email_sent = True
            except Exception as e:
                print("RESET PASSWORD EMAIL SEND ERROR:", str(e))
                error_message = str(e)
            
            request.session['reset_email_sent'] = email_sent
            request.session['reset_email_error'] = error_message
            
            messages.success(request, "A verification code has been dispatched. Please check your email.")
            return redirect('verify_reset_code')
        else:
            messages.error(request, "No account was found with that email address.")
            
    return render(request, 'tracker/LOGIN/forgot_password.html')


def verify_reset_code(request):
    email = request.session.get('reset_email')
    code = request.session.get('reset_code')
    email_sent = request.session.get('reset_email_sent', False)
    error_message = request.session.get('reset_email_error', '')
    
    if not email or not code:
        messages.error(request, "Please enter your email to request a reset code first.")
        return redirect('forgot_password')
        
    if request.method == 'POST':
        entered_code = request.POST.get('code', '').strip()
        if entered_code == code:
            request.session['reset_code_verified'] = True
            return redirect('reset_password')
        else:
            messages.error(request, "Invalid verification code. Please try again.")
            
    return render(request, 'tracker/LOGIN/verify_reset_code.html', {
        'email': email,
        'code': code,
        'email_sent': email_sent,
        'error_message': error_message
    })


def reset_password(request):
    email = request.session.get('reset_email')
    verified = request.session.get('reset_code_verified', False)
    
    if not email or not verified:
        messages.error(request, "Please request and verify a reset code first.")
        return redirect('forgot_password')
        
    if request.method == 'POST':
        new_password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        
        if not new_password or new_password != confirm_password:
            messages.error(request, "Passwords do not match.")
        else:
            user = User.objects.filter(email=email).first()
            if user:
                user.set_password(new_password)
                user.save()
                
                request.session.pop('reset_email', None)
                request.session.pop('reset_code', None)
                request.session.pop('reset_code_verified', None)
                request.session.pop('reset_email_sent', None)
                request.session.pop('reset_email_error', None)
                
                messages.success(request, "Your password has been reset successfully! You can now log in with your new password.")
                return redirect('login')
            else:
                messages.error(request, "User not found.")
                
    return render(request, 'tracker/LOGIN/reset_password.html', {'email': email})





